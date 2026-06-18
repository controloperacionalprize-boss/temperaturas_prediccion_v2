import warnings
warnings.filterwarnings('ignore')
import calendar

import logging
logging.getLogger('prophet').setLevel(logging.ERROR)
logging.getLogger('cmdstanpy').setLevel(logging.ERROR)
import pickle
import hashlib
from pathlib import Path
from scipy.stats import norm
import itertools
import io
import os
import hashlib
import pandas as pd
import numpy as np
from scipy.interpolate import CubicSpline
from prophet import Prophet
import json
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import streamlit as st
import msal
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
try:
    import requests
    REQUESTS_DISPONIBLE = True
except ImportError:
    REQUESTS_DISPONIBLE = False
    
# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Temperatura Fundos — Aquanqa",
    page_icon="🌡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

_APP_DIR      = os.path.dirname(os.path.abspath(__file__))
NORMALES_PATH = os.path.join(_APP_DIR, "assets", "NORMALES_1991_2020.xlsx")
H_MAX_PENDIENTE = 14  # días: más allá de esto, la pendiente deja de extrapolarse

# ══════════════════════════════════════════════════════════════
# ESTILOS CSS (igual a v3.2)
# ══════════════════════════════════════════════════════════════

st.markdown("""
<style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body, [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #f5f7fa 0%, #e8ecf1 100%);
        font-family: 'Segoe UI', 'Arial', sans-serif;
    }
    [data-testid="stSidebar"] {
        background: #FFFFFF;
        color: #333333;
        border-right: 1px solid #E0E0E0;
    }
    [data-testid="stSidebar"] > div:first-child { background: transparent !important; }
    [data-testid="stSidebar"] .stMarkdown { color: #333333; }
    [data-testid="stSidebar"] h3 {
        color: #1565C0;
        font-weight: 700;
        letter-spacing: 0.05em;
        margin-top: 16px;
        margin-bottom: 10px;
    }
    [data-testid="stSidebar"] .stSelectbox > div > div,
    [data-testid="stSidebar"] .stNumberInput input {
        background: #F5F5F5 !important;
        color: #333333 !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 6px !important;
    }
    .header-bar {
        background: linear-gradient(90deg, #0D2340 0%, #1565C0 100%);
        border-radius: 10px;
        padding: 22px 28px;
        margin-bottom: 22px;
        box-shadow: 0 4px 15px rgba(13, 35, 64, 0.2);
    }
    .header-bar h1 {
        color: #FFFFFF;
        font-size: 1.5rem;
        font-weight: 700;
        margin: 0;
        letter-spacing: 0.02em;
    }
    .header-bar p {
        color: #90CAF9;
        font-size: 0.85rem;
        margin: 6px 0 0 0;
    }
    .metric-card {
        background: linear-gradient(135deg, #FFFFFF 0%, #F5F9FC 100%);
        border: 1px solid rgba(200, 210, 225, 0.5);
        border-radius: 10px;
        padding: 16px;
        text-align: center;
        margin-bottom: 10px;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
        transition: all 0.3s ease;
    }
    .metric-card:hover {
        box-shadow: 0 4px 16px rgba(13, 35, 64, 0.1);
        transform: translateY(-2px);
    }
    .metric-card .label {
        font-size: 0.70rem;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 700;
        margin-bottom: 6px;
    }
    .metric-card .value {
        font-size: 1.65rem;
        font-weight: 700;
        color: #0D2340;
        line-height: 1.1;
    }
    .metric-card .sub {
        font-size: 0.72rem;
        color: #94A3B8;
        margin-top: 4px;
    }
    .info-box {
        background: linear-gradient(135deg, #EFF6FF 0%, #E3F2FD 100%);
        border-left: 4px solid #1565C0;
        border-radius: 6px;
        padding: 12px 16px;
        font-size: 0.82rem;
        color: #1E3A5F;
        margin-bottom: 12px;
    }
    .section-divider {
        border: none;
        border-top: 1px solid rgba(13, 35, 64, 0.2);
        margin: 16px 0;
    }
    .footer-note {
        font-size: 0.70rem;
        color: #94A3B8;
        font-style: italic;
        text-align: center;
        margin-top: 28px;
        padding: 16px;
        border-top: 1px solid rgba(13, 35, 64, 0.1);
    }
    .stButton > button {
        background: linear-gradient(135deg, #1565C0 0%, #0D47A1 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 700;
        font-size: 0.85rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(21, 101, 192, 0.3);
        letter-spacing: 0.05em;
    }
    .stButton > button:hover {
        box-shadow: 0 6px 20px rgba(21, 101, 192, 0.5);
        transform: translateY(-2px);
    }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════

SHEET_NAME      = "Prize_Climatology"
MIN_REGISTROS   = 80
ROLLING_DIAS    = 3
SIGMA_CLIM      = 1.5
Z_Q             = 0.6745
DELTA_Q         = round(SIGMA_CLIM * Z_Q, 2)

# Hiperparámetros Prophet — cualquier cambio aquí invalida el pickle automáticamente
PROPHET_PARAMS = {
    "changepoint_prior_scale" : 0.15,
    "changepoint_range"       : 0.95,
    "seasonality_prior_scale" : 2.0,
    "yearly_seasonality"      : 8,
    "fourier_monthly"         : 3,
    "bias_window_dias"        : 21,
    "interval_width"          : 0.90,
    "anomaly_based"           : "v4",
    "clim_halflife_days"      : 90,      # ya no se usa para climatología, puedes dejarlo
    "n_harmonics"             : 4,       # ← NUEVO: armónicos para climatología diaria
    "clim_halflife_anios"     : 2.0,     # ← NUEVO: años anteriores pesan menos (halflife)
}

TMAX_MIN, TMAX_MAX = 10.0, 45.0
TMIN_MIN, TMIN_MAX =  5.0, 35.0

GRID_COLOR      = '#CFD8DC'
BG_TMAX = '#FFFFFF'
BANDA_TMAX      = '#FFCDD2'
CLIM_TMAX       = '#C62828'
REAL_TMAX_COLOR = '#000000'
BG_TMIN = '#FFFFFF'
BANDA_TMIN      = '#BBDEFB'
CLIM_TMIN       = '#1565C0'
REAL_TMIN_COLOR = '#000000'
PRED_COLOR_TMAX = '#FF0000'
PRED_COLOR_TMIN = '#1565C0'
# Colores para la segunda línea climatológica dinámica
CLIM2_COLOR      = '#FF6F00'   # naranja
BANDA2_TMAX      = '#FFE0B2'
BANDA2_TMIN      = '#E3F2FD'   # azul claro distinto al de Tmin

MESES_RAW = [
    'Enero','Febrero','Marzo','Abril','Mayo','Junio',
    'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'
]

DISTRITO_BUSCAR = 'GUADALUPE'
DISTRITO_NINGUNA = '— Ninguna —'

THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

#
#======================ENFEN==============================
# ══════════════════════════════════════════════════════════════
# AJUSTE MANUAL POR EVENTO CLIMÁTICO DECLARADO (ENFEN)
# ══════════════════════════════════════════════════════════════
# 🔔 ACTUALIZAR cuando salga un nuevo Comunicado Oficial ENFEN (enfen.imarpe.gob.pe)
# Última revisión: Comunicado N°10-2026 (29/may/2026) — confirma continuidad
# del N°09-2026, sin cambios. Alerta de El Niño Costero, magnitud moderada
# probable mayo-agosto 2026, evento se prolongaría hasta feb-2027.
AJUSTE_ENFEN = {
    # (mes, año): {variable: ajuste en °C}
    # Tmax calibrado empíricamente con MBE walk-forward (validación may-2026):
    #   Arena Azul +1.84°C, Ayllu Allpa +2.01°C, Vivadis +1.97°C → promedio ≈1.94°C
    (5, 2026): {'Tmax': 1.9, 'Tmin': 1.0},   # ← Tmax actualizado (era 2.5)
    (6, 2026): {'Tmax': 2.5, 'Tmin': 1.0},
    (7, 2026): {'Tmax': 2.5, 'Tmin': 1.0},
    (8, 2026): {'Tmax': 2.5, 'Tmin': 1.0},
}

def obtener_ajuste_enfen(mes, anio, variable):
    return AJUSTE_ENFEN.get((mes, anio), {}).get(variable, 0.0)
# ══════════════════════════════════════════════════════════════
# MONITOR ENFEN — alerta de comunicados nuevos
# ══════════════════════════════════════════════════════════════
ENFEN_URL = "https://enfen.imarpe.gob.pe/comunicados/"
ENFEN_CACHE_FILE = Path(_APP_DIR) / "assets" / "enfen_ultimo_visto.json"


@st.cache_data(ttl=3600 * 12, show_spinner=False)  # revisa máx. cada 12 horas
def chequear_comunicado_enfen():
    """Revisa la página de comunicados ENFEN y extrae el último número/fecha."""
    if not REQUESTS_DISPONIBLE:
        return None
    try:
        resp = requests.get(ENFEN_URL, timeout=10, headers={'User-Agent': 'Mozilla/5.0'})
        resp.raise_for_status()
        texto = resp.text

        match = re.search(
            r'Comunicado\s+Oficial\s+Enfen\s*N[°º]\s*(\d+)\s*-\s*(\d{4}).{0,60}?'
            r'(\d{1,2}\s+\w+,?\s+\d{4})',
            texto, re.IGNORECASE | re.DOTALL
        )
        if match:
            numero, anio, fecha = match.group(1), match.group(2), match.group(3).strip()
            return {'numero': numero, 'anio': anio, 'fecha': fecha, 'id': f"{numero}-{anio}"}
    except Exception:
        return None
    return None


def _leer_ultimo_visto_enfen():
    if ENFEN_CACHE_FILE.exists():
        try:
            return json.loads(ENFEN_CACHE_FILE.read_text())
        except Exception:
            return {'id': None}
    return {'id': None}


def _guardar_ultimo_visto_enfen(comunicado):
    ENFEN_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    ENFEN_CACHE_FILE.write_text(json.dumps(comunicado))
# ══════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════

def _hex_to_rgba(hex_color: str, alpha: float) -> str:
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'rgba({r},{g},{b},{alpha})'

def _normalizar(s) -> str:
    import unicodedata
    import pandas as pd
    
    # Proteger contra NaN, None y floats
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ''
    
    # Convertir a string si no lo es
    s = str(s).strip()
    
    # Descartar "nan" literals
    if s.lower() == 'nan' or s == '':
        return ''
    
    # Normalizar
    return unicodedata.normalize('NFKD', s).encode('ascii', errors='ignore').decode('utf-8').upper().strip()
# ══════════════════════════════════════════════════════════════
# LECTURA OPTIMIZADA
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def leer_meteo_bytes_optimizado(file_bytes: bytes, filename: str, sheet: str = SHEET_NAME) -> pd.DataFrame:
    ext = filename.lower().rsplit('.', 1)[-1]

    if ext == 'csv':
        try:
            df = pd.read_csv(
                io.BytesIO(file_bytes),
                encoding='utf-8-sig',
                sep=None,
                engine='python',
                on_bad_lines='skip',
            )
        except Exception:
            df = pd.read_csv(
                io.BytesIO(file_bytes),
                encoding='latin-1',
                sep=None,
                engine='python',
                on_bad_lines='skip',
            )
        return df

    elif ext == 'xlsx':
        return pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet,
            # ✅ No parsear fechas aquí — lo hacemos en cargar_meteoro
            # para controlar el formato exacto
            parse_dates=False,
        )

    else:
        raise ValueError(f"Extensión no soportada: .{ext}")
# ══════════════════════════════════════════════════════════════
# SPLINE CON CACHE
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner="Cargando polígonos KMZ…")
def load_kmz_bytes(_kmz_bytes: bytes):
    """
    Parsea polígonos de módulos AQ1/AQ2 desde un KMZ (en bytes).
    Devuelve lista de dicts: name, coords, mod_n, fundo_aq.
    """
    import zipfile, io, re
    from lxml import etree

    try:
        with zipfile.ZipFile(io.BytesIO(_kmz_bytes)) as kmz:
            kml_files = [f for f in kmz.namelist() if f.lower().endswith('.kml')]
            if not kml_files:
                return []
            kml_content = kmz.read(kml_files[0])

        parser = etree.XMLParser(recover=True, encoding='utf-8')
        root = etree.fromstring(kml_content, parser=parser)
        nsmap = {'kml': 'http://www.opengis.net/kml/2.2'}

        folders = root.xpath('.//kml:Folder', namespaces=nsmap) \
                  or root.xpath('.//*[local-name()="Folder"]')

        target_folders = []
        for folder in folders:
            fname_xpath = folder.xpath('.//kml:name/text()', namespaces=nsmap) \
                          or folder.xpath('.//*[local-name()="name"]/text()')
            if fname_xpath:
                fname = fname_xpath[0].upper()
                if ('AQ1' in fname or 'AQ2' in fname) and 'MODULO' in fname:
                    target_folders.append((folder, fname_xpath[0]))

        polygons = []
        for target_folder, folder_name in target_folders:
            placemarks = target_folder.xpath('.//kml:Placemark', namespaces=nsmap) \
                         or target_folder.xpath('.//*[local-name()="Placemark"]')

            fundo_match = re.search(r'(AQ\d+)', folder_name, re.IGNORECASE)
            fundo_aq = fundo_match.group(1).upper() if fundo_match else None
            mod_match = re.search(r'MODULO\s*0*(\d+)', folder_name, re.IGNORECASE)
            mod_n = int(mod_match.group(1)) if mod_match else None

            if not fundo_aq or not mod_n:
                continue

            for placemark in placemarks:
                name_xpath = placemark.xpath('.//kml:name/text()', namespaces=nsmap) \
                              or placemark.xpath('.//*[local-name()="name"]/text()')
                name = name_xpath[0].strip() if name_xpath else ""

                coord_xpath = placemark.xpath('.//kml:Polygon//kml:coordinates/text()', namespaces=nsmap) \
                              or placemark.xpath('.//*[local-name()="Polygon"]//*[local-name()="coordinates"]/text()')
                if not coord_xpath:
                    continue

                coords = []
                for pair in coord_xpath[0].strip().split():
                    parts = pair.split(',')
                    if len(parts) >= 2:
                        try:
                            lon, lat = float(parts[0]), float(parts[1])
                            coords.append([lat, lon])
                        except ValueError:
                            pass

                if len(coords) < 3:
                    continue

                polygons.append({
                    "name": name or f"Polígono {len(polygons)+1}",
                    "coords": coords,
                    "mod_n": mod_n,
                    "fundo_aq": fundo_aq,
                })

        return polygons

    except Exception:
        return []
@st.cache_data(show_spinner="Descargando KMZ desde GitHub…", ttl=3600)
def download_kmz_from_github() -> bytes | None:
    import urllib.request, urllib.error
    
    token = st.secrets.get("GITHUB_TOKEN_KMZ", "")
    
    # ✅ OPCIÓN 1: raw.githubusercontent.com (más simple, sin API)
    raw_url = (
        "https://raw.githubusercontent.com/"
        "controloperacionalprize-boss/CAMPO_RENDIMIENTO/"
        "main/MODULOS_PRIZE_PAIJAN.kmz"  # ← Verifica la rama (main, master, etc)
    )
    
    # ✅ OPCIÓN 2: API de GitHub (requiere token)
    api_url = (
        "https://api.github.com/repos/"
        "controloperacionalprize-boss/CAMPO_RENDIMIENTO/"
        "contents/MODULOS_PRIZE_PAIJAN.kmz"
    )
    
    for url, use_token in [(raw_url, False), (api_url, True)]:
        try:
            headers = {"Accept": "application/vnd.github.v3.raw"} if use_token else {}
            if token and use_token:
                headers["Authorization"] = f"token {token}"
            
            req = urllib.request.Request(url, headers=headers)
            resp = urllib.request.urlopen(req, timeout=30)
            kmz_bytes = resp.read()
            
            st.write(f"✅ KMZ descargado: {len(kmz_bytes)} bytes desde {url.split('/')[-2]}")
            return kmz_bytes
            
        except urllib.error.HTTPError as e:
            st.write(f"❌ HTTP {e.code}: {url}")
        except urllib.error.URLError as e:
            st.write(f"❌ URL Error: {e.reason}")
        except Exception as e:
            st.write(f"❌ Error: {str(e)}")
    
    st.error("❌ No se pudo descargar KMZ desde GitHub")
    return None

@st.cache_data(show_spinner=False)
def disolver_modulos(_kmz_polygons):
    """Une los polígonos de cada (fundo_aq, mod_n) en un solo contorno."""
    from shapely.geometry import Polygon as ShPolygon
    from shapely.ops import unary_union

    grupos = {}
    for p in _kmz_polygons:
        key = (p['fundo_aq'], p['mod_n'])
        anillo = [(lon, lat) for lat, lon in p['coords']]  # shapely usa (lon, lat)
        if len(anillo) < 3:
            continue
        try:
            poly = ShPolygon(anillo)
            if not poly.is_valid:
                poly = poly.buffer(0)
            grupos.setdefault(key, []).append(poly)
        except Exception:
            continue

    resultado = []
    for (fundo_aq, mod_n), polys in grupos.items():
        union = unary_union(polys)
        geoms = [union] if union.geom_type == 'Polygon' else list(union.geoms)
        for geom in geoms:
            if geom.is_empty:
                continue
            coords = [[lat, lon] for lon, lat in geom.exterior.coords]
            resultado.append({'fundo_aq': fundo_aq, 'mod_n': mod_n, 'coords': coords})

    return resultado

@st.cache_data(show_spinner=False)
def spline_diario_cached(fecha_inicio_str: str, fecha_fin_str: str, 
                         valores_tuple: tuple) -> pd.DataFrame:
    """Spline con clave de cache determinística."""
    fecha_inicio = pd.Timestamp(fecha_inicio_str)
    fecha_fin = pd.Timestamp(fecha_fin_str)
    valores_mensuales = list(valores_tuple)
    
    fechas_ctrl, vals_ctrl = [], []
    for year in range(fecha_inicio.year - 1, fecha_fin.year + 2):
        for mes_idx, val in enumerate(valores_mensuales):
            fechas_ctrl.append(pd.Timestamp(year=year, month=mes_idx + 1, day=15))
            vals_ctrl.append(val)
    
    fechas_ctrl = pd.DatetimeIndex(fechas_ctrl)
    x_ctrl = (fechas_ctrl - fechas_ctrl[0]).days.values.astype(float)
    cs = CubicSpline(x_ctrl, vals_ctrl)
    
    rango = pd.date_range(fecha_inicio, fecha_fin, freq='D')
    x_eval = (rango - fechas_ctrl[0]).days.values.astype(float)
    
    return pd.DataFrame({'Fecha': rango, 'Valor': cs(x_eval).round(2)})

# ══════════════════════════════════════════════════════════════
# CLIMATOLOGÍA ARMÓNICA MULTI-AÑO
# ══════════════════════════════════════════════════════════════
def calcular_climatologia_armonica(df_clim, n_harmonics, halflife_anios):
    """
    Ajusta climatología diaria vía regresión armónica (Fourier)
    sobre todos los años disponibles, con peso exponencial por antigüedad.
    df_clim: columnas ['ds','y'] — 'y' en temperatura absoluta.
    """
    d = df_clim[['ds', 'y']].dropna().copy()
    doy = d['ds'].dt.dayofyear
    theta = 2 * np.pi * doy / 365.25

    cols = {'const': np.ones(len(d))}
    for k in range(1, n_harmonics + 1):
        cols[f'sin{k}'] = np.sin(k * theta)
        cols[f'cos{k}'] = np.cos(k * theta)
    X = np.column_stack(list(cols.values()))

    # ── Pesos por antigüedad (años) ─────────────────────────────
    edad_anios = ((d['ds'].max() - d['ds']).dt.days / 365.25).to_numpy()
    w = np.exp(-edad_anios / halflife_anios)

    Xw = X * w[:, None]
    coef, *_ = np.linalg.lstsq(Xw, d['y'].to_numpy() * w, rcond=None)
    return coef


def predecir_climatologia_armonica(fechas, coef, n_harmonics):
    doy = pd.to_datetime(fechas).dayofyear.to_numpy().astype(float)
    theta = 2 * np.pi * doy / 365.25
    cols = [np.ones(len(doy))]
    for k in range(1, n_harmonics + 1):
        cols.append(np.sin(k * theta))
        cols.append(np.cos(k * theta))
    X = np.column_stack(cols)
    return X @ coef
# ══════════════════════════════════════════════════════════════
# CARGAR NORMALES (igual a antes)
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def cargar_catalogo_normales(_file_bytes, hoja='TMAX'):
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)

    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break

    if header_row is None:
        return {}

    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_sector   = next((c for c in df.columns if _normalizar(c) == 'SECTOR'), None)
    col_depto    = next((c for c in df.columns if _normalizar(c) == 'DEPARTAMENTO'), None)
    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)

    if col_depto is None or col_distrito is None:
        return {}

    df['_sector']   = df[col_sector].astype(str).apply(_normalizar) if col_sector else 'SIN SECTOR'
    df['_depto']    = df[col_depto].astype(str).apply(_normalizar)
    df['_distrito'] = df[col_distrito].astype(str).apply(_normalizar)

    df = df[
        (df['_depto']    != 'NAN') & (df['_depto']    != '') &
        (df['_distrito'] != 'NAN') & (df['_distrito'] != '') &
        (df['_sector']   != 'NAN') & (df['_sector']   != '')
    ].copy()

    # ── Excluir Guadalupe del catálogo dinámico ────────────
    df = df[~df['_distrito'].str.contains(DISTRITO_BUSCAR, na=False)].copy()

    # catalogo: {sector: {departamento: [distritos]}}
    catalogo = {}
    for sector, g_sector in df.groupby('_sector'):
        catalogo[sector] = {}
        for depto, g_depto in g_sector.groupby('_depto'):
            distritos = sorted(g_depto['_distrito'].unique().tolist())
            catalogo[sector][depto] = distritos

    return catalogo

@st.cache_data(show_spinner=False)
def cargar_normales(_file_bytes, hoja):
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break
    
    if header_row is None:
        raise ValueError(f"No se encontró fila con 'DISTRITO' en hoja '{hoja}'.")
    
    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    
    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
    if col_distrito is None:
        raise ValueError(f"Columna DISTRITO no encontrada.")
    
    df['_dist_norm'] = df[col_distrito].astype(str).apply(_normalizar)
    df_f = df[df['_dist_norm'].str.contains(DISTRITO_BUSCAR, na=False)].copy()
    
    if df_f.empty:
        raise ValueError(f"No se encontraron estaciones del distrito '{DISTRITO_BUSCAR}'.")
    
    meses_encontrados = {}
    for col in df.columns:
        col_norm = _normalizar(col)
        for mes in MESES_RAW:
            if col_norm == _normalizar(mes) and mes not in meses_encontrados:
                meses_encontrados[mes] = col
                break
    
    meses_faltantes = [m for m in MESES_RAW if m not in meses_encontrados]
    if meses_faltantes:
        raise ValueError(f"Columnas de meses no encontradas: {meses_faltantes}")
    
    cols_meses = [meses_encontrados[m] for m in MESES_RAW]
    valores = df_f[cols_meses].apply(pd.to_numeric, errors='coerce').mean(axis=0).values
    
    col_nombre = next((c for c in df.columns if _normalizar(c) == 'NOMBRE ESTACION'), col_distrito)
    col_prov = next((c for c in df.columns if _normalizar(c) == 'PROVINCIA'), col_distrito)
    
    return (
        valores,
        valores - DELTA_Q,
        valores + DELTA_Q,
        df_f[[col_nombre, col_prov, col_distrito]].drop_duplicates()
    )

# ══════════════════════════════════════════════════════════════
# CARGA METEOROLOGÍA OPTIMIZADA
# ══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def cargar_meteoro_optimizado(_file_bytes, filename, sheet, fundos_sel, min_reg):
    """
    Carga y procesa datos meteorológicos cada 15 min a diarios.
    
    Fixes aplicados:
    - Formato de fecha M/DD/YYYY detectado automáticamente
    - 96 registros por día (cada 15 min) como base
    - Rangos amplios 5-45°C para no descartar datos reales
    - Último día por fundo siempre incluido
    - Protección si ET-mm no existe
    """
    df = leer_meteo_bytes_optimizado(_file_bytes, filename, sheet)

    # ── 1. Fechas con detección de formato ────────────────────
    # El Excel tiene formato M/DD/YYYY HH:MM (americano)
    # dayfirst=False asegura que 6/01/2026 = 1 de junio, no 6 de enero
    df['Fecha-Hora'] = pd.to_datetime(
        df['Fecha-Hora'],
        dayfirst=False,   # ← M/DD/YYYY: el primero es el MES
        errors='coerce'
    )
    df = df.dropna(subset=['Fecha-Hora'])

    # ── 2. Numéricos ───────────────────────────────────────────
    for col in ['Temp-C', 'TempAlta-C', 'TempBaja-C']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # ── 3. Filtro de rango — solo elimina basura del sensor ────
    df = df[
        (df['TempAlta-C'] >= TMAX_MIN) & (df['TempAlta-C'] <= TMAX_MAX) &
        (df['TempBaja-C'] >= TMIN_MIN) & (df['TempBaja-C'] <= TMIN_MAX)
    ].copy()

    # ── 4. Protección ET-mm ────────────────────────────────────
    if 'ET-mm' not in df.columns:
        df['ET-mm'] = 0.0

    # ── 5. Fecha diaria normalizada ────────────────────────────
    df['Fecha'] = df['Fecha-Hora'].dt.normalize()

    # ── 6. Conteo de registros por fundo+fecha ─────────────────
    # Con datos cada 15 min → día completo = 96 registros
    # min_reg=80 equivale a tener al menos 20 horas de datos
    reg = df.groupby(['Fundo', 'Fecha'], as_index=False).size()
    reg.columns = ['Fundo', 'Fecha', 'N_registros']
    df = df.merge(reg, on=['Fundo', 'Fecha'], how='left')


    fecha_max_por_fundo = (
        df.groupby('Fundo')['Fecha']
          .max()
          .reset_index()
          .rename(columns={'Fecha': 'Fecha_max_fundo'})
    )
    df = df.merge(fecha_max_por_fundo, on='Fundo', how='left')

    df = df[
        (df['N_registros'] >= min_reg) |
        (df['Fecha'] == df['Fecha_max_fundo'])
    ].copy()

    df = df.drop(columns=['Fecha_max_fundo'])

    # ── 8. Filtro por fundos seleccionados ─────────────────────
    if fundos_sel:
        df = df[df['Fundo'].isin(fundos_sel)]

    # ── 9. Agregación diaria ───────────────────────────────────
    dia = (
        df.groupby(['Empresa', 'Fundo', 'Fecha'], as_index=False)
          .agg(
              Tmax=('TempAlta-C', 'max'),
              Tmin=('TempBaja-C', 'min'),
              ET=('ET-mm',        'sum'),
          )
    )
    dia[['Tmax', 'Tmin', 'ET']] = dia[['Tmax', 'Tmin', 'ET']].round(2)
    dia = dia.sort_values(['Fundo', 'Fecha']).reset_index(drop=True)

    # ── 10. Suavizado rolling 3 días ───────────────────────────
    dia['Tmax_smooth'] = (
        dia.groupby('Fundo')['Tmax']
           .transform(lambda x: x.rolling(ROLLING_DIAS, center=True, min_periods=1).mean())
           .round(2)
    )
    dia['Tmin_smooth'] = (
        dia.groupby('Fundo')['Tmin']
           .transform(lambda x: x.rolling(ROLLING_DIAS, center=True, min_periods=1).mean())
           .round(2)
    )

    return dia
# ══════════════════════════════════════════════════════════════
# PROPHET OPTIMIZADO - con cache por hash
# ══════════════════════════════════════════════════════════════
def _hash_serie(serie_bytes: bytes) -> str:
    """Hash de la serie para cache."""
    return hashlib.md5(serie_bytes).hexdigest()[:12]

@st.cache_data(show_spinner=False)
def entrenar_prophet_opt(_serie_hash: str, _serie_bytes: bytes, dias_pred: int,
                         variable: str, fundo: str):
    """Prophet con historial completo + peso a datos recientes."""
    serie = pd.read_parquet(io.BytesIO(_serie_bytes))

    df = (
        serie.rename(columns={'Fecha': 'ds', 'Valor': 'y'})
             .dropna()
             .sort_values('ds')
             .reset_index(drop=True)
    )
    df['ds'] = pd.to_datetime(df['ds']).dt.tz_localize(None)

    media = df['y'].mean()
    std = df['y'].std()
    df['y'] = np.clip(df['y'], media - 3.0 * std, media + 3.0 * std)

# ── Climatología diaria vía regresión armónica multi-año ─────────────
    coef_clim = calcular_climatologia_armonica(
        df[['ds', 'y']], PROPHET_PARAMS['n_harmonics'], PROPHET_PARAMS['clim_halflife_anios']
    )
    _fi_c = df['ds'].min()
    _ff_c = df['ds'].max() + pd.Timedelta(days=dias_pred + 60)
    _fechas_clim = pd.date_range(_fi_c, _ff_c, freq='D')
    _clim_daily = pd.DataFrame({
        'ds': _fechas_clim,
        'clim': predecir_climatologia_armonica(_fechas_clim, coef_clim, PROPHET_PARAMS['n_harmonics']).round(2)
    })
    df = df.merge(_clim_daily, on='ds', how='left')
    _y_abs_min, _y_abs_max = df['y'].min(), df['y'].max()  # límites antes de restar
    df['y'] = (df['y'] - df['clim']).round(4)              # anomalía

    df['weight'] = 1.0
    mask_reciente = df['ds'] >= df['ds'].max() - pd.Timedelta(days=180)
    df.loc[mask_reciente, 'weight'] = 2.5

    # Validación cruzada walk-forward
    n_total = len(df)
    n_validacion = min(15, max(10, n_total // 8))
    mae_por_h_listas = {h: [] for h in range(1, min(dias_pred + 1, 31))}

    if n_validacion >= 10:
        for i in range(n_validacion):
            corte = n_total - n_validacion + i
            train = df.iloc[:corte].copy()
            test  = df.iloc[corte:corte + dias_pred].copy()

            if len(test) < 1 or len(train) < 30:
                break

            train['weight'] = 1.0
            mask_rec_train = train['ds'] >= train['ds'].max() - pd.Timedelta(days=180)
            train.loc[mask_rec_train, 'weight'] = 2.5

            try:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    m = Prophet(
                        yearly_seasonality=PROPHET_PARAMS['yearly_seasonality'],
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        seasonality_mode='additive',
                        changepoint_prior_scale=PROPHET_PARAMS['changepoint_prior_scale'],
                        changepoint_range=PROPHET_PARAMS['changepoint_range'],   # ← AGREGAR
                        seasonality_prior_scale=PROPHET_PARAMS['seasonality_prior_scale'],
                        interval_width=PROPHET_PARAMS['interval_width'],
                    )
                    m.add_seasonality(name='monthly', period=30.5, fourier_order=PROPHET_PARAMS['fourier_monthly'])

                    m.fit(train)
                    future = m.make_future_dataframe(periods=len(test), freq='D')
                    pred = m.predict(future).tail(len(test))
                    _test_clim = _clim_daily[_clim_daily['ds'].isin(
                    pd.to_datetime(test['ds']).dt.normalize()
                    )].set_index('ds')['clim']

                    for h, (ds_val, real_anom, yhat_anom) in enumerate(
                        zip(test['ds'], test['y'].values, pred['yhat'].values), 1
                    ):
                        ds_norm = pd.Timestamp(ds_val).normalize()
                        clim_val = float(_test_clim.get(ds_norm, _clim_daily['clim'].mean()))
                        real_abs = real_anom + clim_val   # temperatura absoluta real
                        yhat_abs = yhat_anom + clim_val   # temperatura absoluta predicha
                        if h <= 30:
                            mae_por_h_listas[h].append(abs(real_abs - yhat_abs))
               
            except Exception:
                continue

    todos_errores = [e for errs in mae_por_h_listas.values() for e in errs]
    mae_real = round(float(np.mean(todos_errores)), 3) if todos_errores else None

    mae_por_h = {
        h: round(float(np.mean(errs)), 3) if errs else None
        for h, errs in mae_por_h_listas.items()
    }

    # Modelo final
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        modelo = Prophet(
        yearly_seasonality=PROPHET_PARAMS['yearly_seasonality'],
        weekly_seasonality=False,
        daily_seasonality=False,
        seasonality_mode='additive',
        changepoint_prior_scale=PROPHET_PARAMS['changepoint_prior_scale'],
        changepoint_range=PROPHET_PARAMS['changepoint_range'],   # ← AGREGAR
        seasonality_prior_scale=PROPHET_PARAMS['seasonality_prior_scale'],
        interval_width=PROPHET_PARAMS['interval_width'],
    )
        modelo.add_seasonality(name='monthly', period=30.5, fourier_order=PROPHET_PARAMS['fourier_monthly'])
        modelo.fit(df)
        future = modelo.make_future_dataframe(periods=dias_pred, freq='D')
        forecast = modelo.predict(future)

    # ── Añadir climatología → temperatura absoluta ───────────────────
    forecast = forecast.merge(_clim_daily[['ds', 'clim']], on='ds', how='left')
    forecast['clim'] = forecast['clim'].ffill().bfill()
    for col in ['yhat', 'yhat_lower', 'yhat_upper']:
        forecast[col] = (forecast[col] + forecast['clim']).round(2)

    # ── Forecast futuro (solo dias_pred hacia adelante) ────────
    result_futuro = forecast.tail(dias_pred)[
        ['ds', 'yhat', 'yhat_lower', 'yhat_upper']
    ].copy().reset_index(drop=True)

    # ── Forecast histórico COMPLETO (todas las fechas del future) ──
    result_historico = forecast[
        ['ds', 'yhat', 'yhat_lower', 'yhat_upper']
    ].copy().reset_index(drop=True)

    # ✅ FIX — convertir df['y'] (anomalía) a temperatura absoluta antes del merge
    insample = forecast[['ds', 'yhat']].merge(
        df[['ds', 'y', 'clim']].assign(y=lambda d: d['y'] + d['clim']),
        on='ds', how='inner'
    ).sort_values('ds').reset_index(drop=True)
    ventana_bias = min(PROPHET_PARAMS['bias_window_dias'] * 2, len(insample))
    ult = insample.tail(ventana_bias).copy().reset_index(drop=True)
    ult['residuo'] = ult['y'] - ult['yhat']
    ult['t'] = np.arange(len(ult))

    if len(ult) >= 5:
        b_slope, a_int = np.polyfit(ult['t'], ult['residuo'], 1)
    else:
        b_slope, a_int = 0.0, (ult['residuo'].mean() if len(ult) else 0.0)

    h_arr = np.arange(1, len(result_futuro) + 1)
    h_efectivo = np.minimum(h_arr, H_MAX_PENDIENTE)
    # ── Bias estacional multi-año (mismo mes, años anteriores, ponderado) ──
    insample['anio'] = insample['ds'].dt.year
    insample['residuo_total'] = insample['y'] - insample['yhat']

    mes_objetivo  = result_futuro['ds'].iloc[0].month
    anio_objetivo = result_futuro['ds'].iloc[0].year
    datos_mes_prev = insample[
        (insample['ds'].dt.month == mes_objetivo) &
        (insample['anio'] < anio_objetivo)
    ]
    if len(datos_mes_prev) >= 15:
        edad_anios = anio_objetivo - datos_mes_prev['anio']
        pesos = np.exp(-edad_anios / 1.5)
        bias_estacional = float(np.average(datos_mes_prev['residuo_total'], weights=pesos))
    else:
        bias_estacional = 0.0

    bias_h_arr = a_int + b_slope * (len(ult) - 1) + b_slope * h_efectivo + bias_estacional
    ajuste_enfen = obtener_ajuste_enfen(mes_objetivo, anio_objetivo, variable)

    bias_h_arr = (
        a_int + b_slope * (len(ult) - 1) + b_slope * h_efectivo
        + bias_estacional + ajuste_enfen
    )
    for col in ['yhat', 'yhat_lower', 'yhat_upper']:
        result_futuro[col] = (result_futuro[col] + bias_h_arr).round(2)

    margen = 5.0
    for col in ['yhat', 'yhat_lower', 'yhat_upper']:
        result_futuro[col] = np.clip(
            result_futuro[col],
            _y_abs_min - margen,
            _y_abs_max + margen
        )

    result_futuro['ds'] = pd.to_datetime(
        result_futuro['ds']
    ).dt.tz_localize(None).dt.normalize()

    result_historico['ds'] = pd.to_datetime(
        result_historico['ds']
    ).dt.tz_localize(None).dt.normalize()

    return result_futuro, result_historico, mae_real, mae_por_h

def entrenar_todos_optimizado(dia, dias_pred):
    """Entrena todos los modelos con progreso."""
    forecasts = {}
    fundos    = dia['Fundo'].unique().tolist()
    variables = ['Tmax', 'Tmin', 'ET']
    total     = len(fundos) * len(variables)

    prog = st.progress(0, text="Entrenando modelos Prophet...")

    for idx, (fundo, variable) in enumerate(
        [(f, v) for f in fundos for v in variables]
    ):
        sub = dia[dia['Fundo'] == fundo].copy()
        if sub.empty:
            continue

        if variable == 'ET':
            sub = sub[sub['ET'] > 0].copy()
            if len(sub) < 30:
                continue


        if len(sub) < 30:
            continue

        serie = sub[['Fecha', variable]].rename(columns={variable: 'Valor'})
        buf   = io.BytesIO()
        serie.to_parquet(buf, index=False)
        buf.seek(0)
        serie_bytes = buf.getvalue()
        serie_hash  = _hash_serie(serie_bytes)

        # No se corta nada; Prophet predice desde mañana hacia adelante

        # Calcular días necesarios hasta fin del mes actual
        fecha_ultimo_dato = pd.to_datetime(serie['Fecha'].max()).normalize()
        primer_dia_mes    = fecha_ultimo_dato.replace(day=1)
        ultimo_dia_mes    = pd.Timestamp(
            year=primer_dia_mes.year,
            month=primer_dia_mes.month,
            day=calendar.monthrange(primer_dia_mes.year, primer_dia_mes.month)[1]
        )
        dias_pred_real = max(dias_pred, (ultimo_dia_mes - fecha_ultimo_dato).days)

        forecast, forecast_hist, mae_real, mae_por_h = entrenar_prophet_opt(
            serie_hash, serie_bytes, dias_pred_real, variable, fundo
        )

        if variable == 'ET':
            for col in ['yhat', 'yhat_lower', 'yhat_upper']:
                forecast[col]      = forecast[col].clip(lower=0)
                forecast_hist[col] = forecast_hist[col].clip(lower=0)

        forecasts[(fundo, variable)] = {
            'forecast'     : forecast,
            'forecast_hist': forecast_hist,
            'mae'          : mae_real,
            'mae_por_dia'  : mae_por_h,
        }

        prog.progress(
            (idx + 1) / total,
            text=f"Prophet {fundo} — {variable}  ({idx + 1}/{total})..."
        )

    prog.empty()
    return forecasts



def generar_seccion_validacion(variable, dia, forecasts_cache, dias_pred_ui):
    import calendar
    
    # ── Obtener ventanas dinámicas ──
    fecha_max = pd.to_datetime(dia['Fecha'].max()).normalize()
    ultimo_dia_mes_ant = fecha_max.replace(day=1) - pd.Timedelta(days=1)
    primer_dia_mes_ant = ultimo_dia_mes_ant.replace(day=1)
    fecha_corte_1mes = primer_dia_mes_ant - pd.Timedelta(days=1)
    
    fecha_inicio_3m = primer_dia_mes_ant.replace(day=1) - pd.DateOffset(months=2)
    fecha_corte_3m = fecha_inicio_3m - pd.Timedelta(days=1)
    
    ventanas = {
        '1mes': {
            'label': f"{primer_dia_mes_ant.strftime('%B %Y')}",
            'inicio': primer_dia_mes_ant,
            'fin': ultimo_dia_mes_ant,
            'fecha_corte': fecha_corte_1mes,
            'dias': (ultimo_dia_mes_ant - primer_dia_mes_ant).days + 1,
        },
        '3mes': {
            'label': f"{fecha_inicio_3m.strftime('%b %Y')} → {ultimo_dia_mes_ant.strftime('%b %Y')}",
            'inicio': fecha_inicio_3m,
            'fin': ultimo_dia_mes_ant,
            'fecha_corte': fecha_corte_3m,
            'dias': (ultimo_dia_mes_ant - fecha_inicio_3m).days + 1,
        }
    }
    
    with st.expander(f"📊 Validación Prophet — {variable} (walk-forward honesto)", expanded=False):
        
        tab_1mes, tab_3mes = st.tabs([
            f"Último mes",
            f"Últimos 3 meses"
        ])
        
        with tab_1mes:
            st.info(f"Entrenamiento: datos hasta {ventanas['1mes']['fecha_corte'].strftime('%d/%b/%Y')} | Predicción: {ventanas['1mes']['dias']} días")
            _validar_ventana(variable, dia, ventanas['1mes'], '_1mes')
        
        with tab_3mes:
            st.info(f"Entrenamiento: datos hasta {ventanas['3mes']['fecha_corte'].strftime('%d/%b/%Y')} | Predicción: {ventanas['3mes']['dias']} días")
            _validar_ventana(variable, dia, ventanas['3mes'], '_3mes')

    st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    with st.expander("🔬 Calibración de hiperparámetros (grid search)", expanded=False):
            st.caption(
                "Reentrena cada combinación para cada fundo/variable. "
                "El grid de climatología (n_harmonics/halflife) puede tardar más en '3 meses'."
            )

            col_g1, col_g2 = st.columns(2)
            with col_g1:
                ventana_grid_sel = st.radio(
                    "Ventana", options=['1mes', '3mes'], horizontal=True,
                    key=f"grid_ventana_{variable}"
                )
            with col_g2:
                grid_tipo = st.radio(
                    "Tipo de grid",
                    options=['Hiperparámetros Prophet', 'Climatología (n_harmonics/halflife)'],
                    horizontal=True, key=f"grid_tipo_{variable}"
                )

            if st.button(f"Ejecutar grid search — {variable}", key=f"grid_btn_{variable}"):
                if grid_tipo == 'Hiperparámetros Prophet':
                    grid = {
                        'changepoint_prior_scale': [0.05, 0.15, 0.30],
                        'seasonality_prior_scale': [1.0, 2.0, 5.0],
                    }
                else:
                    grid = {
                        'n_harmonics': [2, 3, 4, 6],
                        'clim_halflife_anios': [1.0, 1.5, 2.0, 3.0],
                    }

                df_grid = grid_search_hiperparams(dia, ventanas[ventana_grid_sel], grid)

                if df_grid.empty:
                    st.warning("Sin resultados — revisa que haya suficiente historial.")
                else:
                    st.markdown("##### 📋 Resultados completos")
                    st.dataframe(
                        df_grid.sort_values(['Fundo', 'Variable', 'Precision'], ascending=[True, True, False]),
                        use_container_width=True
                    )
                    mejores = df_grid.loc[df_grid.groupby(['Fundo', 'Variable'])['Precision'].idxmax()]
                    st.markdown("##### 🏆 Mejor combo por fundo/variable")
                    st.dataframe(mejores, use_container_width=True)
                    st.session_state['grid_search_resultado'] = df_grid


def _validar_ventana(variable, dia, ventana, key_suffix):
    import calendar
    
    fundos = dia['Fundo'].unique().tolist()
    fecha_inicio = ventana['inicio']
    fecha_fin = ventana['fin']
    fecha_corte = ventana['fecha_corte']
    
    datos_boxplot = []
    
    for fundo in fundos:
        sub = dia[dia['Fundo'] == fundo].copy().reset_index(drop=True)
        if sub.empty:
            continue
        
        # ── Datos reales en el período ──
        real_periodo = sub[
            (sub['Fecha'] >= fecha_inicio) &
            (sub['Fecha'] <= fecha_fin)
        ].copy().reset_index(drop=True)
        
        if real_periodo.empty:
            continue
        
        # ── Entrenar SOLO con datos ANTES de la ventana ──
        train = sub[sub['Fecha'] <= fecha_corte].copy()
        
        if len(train) < 30:
            st.warning(f"⚠️ {fundo}: historial insuficiente")
            continue
        
        serie_train = train[['Fecha', variable]].rename(
            columns={'Fecha': 'ds', variable: 'y'}
        ).dropna()
        serie_train['ds'] = pd.to_datetime(
            serie_train['ds']
        ).dt.tz_localize(None).dt.normalize()
        
        media = serie_train['y'].mean()
        std = serie_train['y'].std()
        serie_train['y'] = serie_train['y'].clip(lower=media - 3*std, upper=media + 3*std)

         # ── Climatología diaria vía regresión armónica multi-año (validación) ──
        coef_clim_val = calcular_climatologia_armonica(
            serie_train[['ds', 'y']], PROPHET_PARAMS['n_harmonics'], PROPHET_PARAMS['clim_halflife_anios']
        )
        _val_fi = serie_train['ds'].min()
        _val_ff = fecha_fin + pd.Timedelta(days=30)
        _val_fechas = pd.date_range(_val_fi, _val_ff, freq='D')
        _val_clim_daily = pd.DataFrame({
            'ds': _val_fechas,
            'clim': predecir_climatologia_armonica(_val_fechas, coef_clim_val, PROPHET_PARAMS['n_harmonics']).round(2)
        })
        serie_train = serie_train.merge(_val_clim_daily, on='ds', how='left')
        serie_train['y'] = (serie_train['y'] - serie_train['clim']).round(4)

        with st.spinner(f"Entrenando {fundo} {variable}..."):
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    m = Prophet(
                        yearly_seasonality=PROPHET_PARAMS['yearly_seasonality'],
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        seasonality_mode='additive',
                        changepoint_prior_scale=PROPHET_PARAMS['changepoint_prior_scale'],
                        changepoint_range=PROPHET_PARAMS['changepoint_range'],   # ← AGREGAR
                        seasonality_prior_scale=PROPHET_PARAMS['seasonality_prior_scale'],
                        interval_width=PROPHET_PARAMS['interval_width'],
                    )
                    m.add_seasonality(name='monthly', period=30.5, fourier_order=PROPHET_PARAMS['fourier_monthly'])
                    m.fit(serie_train)
                    
                    dias_pred_period = (fecha_fin - fecha_inicio).days + 1
                    future = m.make_future_dataframe(periods=dias_pred_period, freq='D')
                    forecast_wf = m.predict(future)
            
            except Exception as e:
                st.error(f"❌ Error {fundo}: {e}")
                continue
        
        forecast_wf['ds'] = pd.to_datetime(forecast_wf['ds']).dt.tz_localize(None).dt.normalize()

        # ── Añadir climatología ANTES de filtrar pred_periodo ────────────
        forecast_wf = forecast_wf.merge(_val_clim_daily[['ds', 'clim']], on='ds', how='left')
        forecast_wf['clim'] = forecast_wf['clim'].ffill().bfill()
        for col in ['yhat', 'yhat_lower', 'yhat_upper']:
            forecast_wf[col] = (forecast_wf[col] + forecast_wf['clim']).round(2)

        pred_periodo = forecast_wf[
            (forecast_wf['ds'] >= fecha_inicio) &
            (forecast_wf['ds'] <= fecha_fin)
        ][['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy().reset_index(drop=True)


        if pred_periodo.empty:
            continue
        
        # ── Comparación ──
        real_periodo = real_periodo.rename(columns={variable: 'Real'})
        comparacion = real_periodo[['Fecha', 'Real']].merge(
            pred_periodo.rename(columns={
                'ds': 'Fecha',
                'yhat': 'Pred',
                'yhat_lower': 'Pred_Low',
                'yhat_upper': 'Pred_High',
            }),
            on='Fecha', how='inner'
        )
        
        if comparacion.empty:
            continue
        
        # ✅ FIX v2 — bias lineal en función del horizonte (h)
        insample_val = forecast_wf[['ds', 'yhat']].merge(
            serie_train[['ds', 'y', 'clim']].assign(
                y=serie_train['y'] + serie_train['clim']
            ),
            on='ds', how='inner'
        ).sort_values('ds').reset_index(drop=True)

        ventana_bias = min(PROPHET_PARAMS['bias_window_dias'] * 2, len(insample_val))
        ult = insample_val.tail(ventana_bias).copy().reset_index(drop=True)
        ult['residuo'] = ult['y'] - ult['yhat']
        ult['t'] = np.arange(len(ult))

        if len(ult) >= 5:
            b_slope, a_int = np.polyfit(ult['t'], ult['residuo'], 1)
        else:
            b_slope, a_int = 0.0, (ult['residuo'].mean() if len(ult) else 0.0)
        # 1) Bias estacional multi-año
        insample_val['anio'] = insample_val['ds'].dt.year
        insample_val['residuo_total'] = insample_val['y'] - insample_val['yhat']

        mes_objetivo = fecha_inicio.month
        datos_mes_prev = insample_val[
            (insample_val['ds'].dt.month == mes_objetivo) &
            (insample_val['anio'] < fecha_inicio.year)
        ]
        if len(datos_mes_prev) >= 15:
            edad_anios = fecha_inicio.year - datos_mes_prev['anio']
            pesos = np.exp(-edad_anios / 1.5)
            bias_estacional = float(np.average(datos_mes_prev['residuo_total'], weights=pesos))
        else:
            bias_estacional = 0.0

        # 2) h y h_efectivo
        fecha_corte_dt = pd.Timestamp(fecha_corte)
        comparacion = comparacion.copy()
        comparacion['h'] = (comparacion['Fecha'] - fecha_corte_dt).dt.days
        h_efectivo = np.minimum(comparacion['h'].clip(lower=0), H_MAX_PENDIENTE)
        comparacion['ajuste_enfen'] = comparacion['Fecha'].apply(
            lambda f: obtener_ajuste_enfen(f.month, f.year, variable)
        )

        comparacion['bias_h'] = (
            a_int + b_slope * (len(ult) - 1) + b_slope * h_efectivo
            + bias_estacional + comparacion['ajuste_enfen']
        )

        # ── 🩺 Diagnóstico double-counting ──────────────────────
        mbe_sin_correccion = (comparacion['Real'] - comparacion['Pred']).mean()

        tendencia_h = a_int + b_slope * (len(ult) - 1) + b_slope * h_efectivo

        mbe_solo_tendencia = (
            comparacion['Real'] - (comparacion['Pred'] + tendencia_h)
        ).mean()

        mbe_tend_estacional = (
            comparacion['Real'] - (comparacion['Pred'] + tendencia_h + bias_estacional)
        ).mean()

        st.caption(
            f"🩺 Diagnóstico MBE — sin corrección: {mbe_sin_correccion:+.2f}°C | "
            f"+tendencia: {mbe_solo_tendencia:+.2f}°C | "
            f"+estacional: {mbe_tend_estacional:+.2f}°C | "
            f"+ENFEN (final): -- (ver MBE abajo)"
        )

        st.caption(
            f"🔍 Componentes del bias — tendencia (a_int)={a_int:.2f}°C, "
            f"pendiente·h promedio={float((b_slope * h_efectivo).mean()):.2f}°C, "
            f"estacional={bias_estacional:.2f}°C, "
            f"ajuste ENFEN promedio={comparacion['ajuste_enfen'].mean():+.2f}°C "
            f"→ bias_h promedio={comparacion['bias_h'].mean():.2f}°C"
        )

        comparacion['Pred_corr']    = comparacion['Pred'] + comparacion['bias_h']
        comparacion['Error_abs']    = (comparacion['Real'] - comparacion['Pred_corr']).abs().round(2)
        comparacion['Error_signed'] = (comparacion['Real'] - comparacion['Pred_corr']).round(2)
        
        mae_mes = comparacion['Error_abs'].mean().round(2)
        mbe_mes = comparacion['Error_signed'].mean().round(2)
        rmse_mes = round(float(np.sqrt((comparacion['Error_signed']**2).mean())), 2)
        
        dias_dentro   = int((comparacion['Error_abs'] <= 1.5).sum())
        dias_total    = len(comparacion)
        precision_mae = round(dias_dentro / dias_total * 100, 2)
        dias_fuera    = dias_total - dias_dentro
        # ── Techo teórico de precisión (asumiendo bias=0) ───────────────
        std_residual = comparacion['Error_signed'].std()
        if std_residual > 0:
            precision_teorica = (
                norm.cdf(1.5, loc=0, scale=std_residual) -
                norm.cdf(-1.5, loc=0, scale=std_residual)
            ) * 100
        else:
            precision_teorica = 100.0

        brecha = precision_teorica - precision_mae
        diagnostico = (
            "el resto es ruido irreducible — no lo resuelve calibración"
            if abs(mbe_mes) < 0.3 else
            "todavía hay bias corregible — calibración SÍ puede ayudar"
        )

        st.caption(
            f"📐 **Techo teórico (σ={std_residual:.2f}°C, bias=0):** "
            f"{precision_teorica:.0f}% — vs actual {precision_mae:.0f}% "
            f"(brecha {brecha:.0f} pts). {diagnostico}"
        )
        # Ratio RMSE/MAE — si > 1.3 hay días problemáticos
        ratio_rmse_mae = round(rmse_mes / mae_mes, 2) if mae_mes > 0 else None
        
        for _, r in comparacion.iterrows():
            datos_boxplot.append({'Fundo': fundo, 'Error': r['Error_signed']})
        
        if 'bias_correccion' not in st.session_state:
            st.session_state['bias_correccion'] = {}
        st.session_state['bias_correccion'][(fundo, variable)] = float(mbe_mes)
        
        # ── Métricas ──
        st.markdown(f"##### {fundo} — {ventana['label']}")
        
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("MAE", f"{mae_mes:.2f}°C",
        help="Error promedio diario del mes")

        k2.metric("RMSE", f"{rmse_mes:.2f}°C",
                delta=f"ratio {ratio_rmse_mae}x" if ratio_rmse_mae else None,
                help="Si ratio > 1.3: hay días con errores grandes")

        k3.metric(
            "BIAS (MBE)", f"{mbe_mes:+.2f}°C",
            delta="subestima" if mbe_mes > 0 else "sobreestima",
            delta_color="inverse" if mbe_mes > 0 else "normal",
            help="Sesgo sistemático — se corrige automáticamente"
        )

        k4.metric(
            "Precisión (±1.5°C)",
            f"{precision_mae:.1f}%",
            delta=f"{dias_dentro}/{dias_total} días dentro",
            delta_color="normal" if precision_mae >= 60 else "inverse",
            help="% de días con error dentro de ±1.5°C — umbral operacional"
        )
        
        # ── Gráfico error ──
        fig_val = go.Figure()
        
        fig_val.add_hrect(
            y0=-1.5, y1=1.5,
            fillcolor='rgba(76, 175, 80, 0.10)',
            line_width=0,
        )
        
        fig_val.add_hline(y=0, line=dict(color='#546E7A', width=1.5))
        
        for j in range(len(comparacion) - 1):
            x0 = comparacion['Fecha'].iloc[j]
            x1 = comparacion['Fecha'].iloc[j + 1]
            e0 = comparacion['Error_signed'].iloc[j]
            e1 = comparacion['Error_signed'].iloc[j + 1]
            
            if e0 >= 0 and e1 >= 0:
                fig_val.add_trace(go.Scatter(
                    x=[x0, x1, x1, x0], y=[e0, e1, 0, 0],
                    fill='toself', fillcolor='rgba(76,175,80,0.30)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
            elif e0 < 0 and e1 < 0:
                fig_val.add_trace(go.Scatter(
                    x=[x0, x1, x1, x0], y=[e0, e1, 0, 0],
                    fill='toself', fillcolor='rgba(244,67,54,0.30)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
            else:
                frac = abs(e0) / (abs(e0) + abs(e1))
                rango = (x1 - x0).total_seconds()
                x_cross = x0 + pd.Timedelta(seconds=rango * frac)
                
                color_izq = 'rgba(76,175,80,0.30)' if e0 >= 0 else 'rgba(244,67,54,0.30)'
                fig_val.add_trace(go.Scatter(
                    x=[x0, x_cross, x_cross, x0], y=[e0, 0, 0, 0],
                    fill='toself', fillcolor=color_izq,
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
                
                color_der = 'rgba(76,175,80,0.30)' if e1 >= 0 else 'rgba(244,67,54,0.30)'
                fig_val.add_trace(go.Scatter(
                    x=[x_cross, x1, x1, x_cross], y=[0, e1, 0, 0],
                    fill='toself', fillcolor=color_der,
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
        
        fig_val.add_trace(go.Scatter(
            x=comparacion['Fecha'],
            y=comparacion['Error_signed'],
            mode='lines+markers',
            line=dict(color='#1A237E', width=2.5),
            marker=dict(
                size=7,
                color=comparacion['Error_signed'].apply(lambda e: '#4CAF50' if e >= 0 else '#F44336'),
                line=dict(color='white', width=1)
            ),
            name='Error',
            hovertemplate='%{x|%d/%b}<br>Error: %{y:+.2f}°C<extra></extra>'
        ))
        
        fig_val.add_annotation(
            x=comparacion['Fecha'].iloc[len(comparacion) // 2],
            y=2.8,
            text=f"<b>MBE: {mbe_mes:+.2f}°C</b><br>{'🔺 subestima' if mbe_mes > 0 else '🔻 sobreestima'}<br>MAE: {mae_mes:.2f}°C",
            showarrow=False,
            bgcolor='rgba(255,255,255,0.85)',
            bordercolor='#546E7A',
            borderwidth=1,
            font=dict(size=10, color='#1A1A1A'),
            align='center'
        )
        
        fig_val.update_layout(
            height=350,
            title=dict(text=f"<b>Error diario — {fundo} — {variable}</b>", font=dict(size=12), x=0.0),
            xaxis=dict(tickformat='%d/%b', gridcolor='#CFD8DC'),
            yaxis=dict(
                title='Error (°C)',
                ticksuffix='°C',
                gridcolor='#CFD8DC',
                range=[-3.5, 3.5],
                dtick=1,
                zeroline=False,
            ),
            legend=dict(orientation='h', y=-0.15, x=0),
            hovermode='x unified',
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Arial', size=9),
            margin=dict(l=50, r=30, t=60, b=40),
        )
        
        st.plotly_chart(fig_val, use_container_width=True, key=f"fig_val_{variable}_{fundo}{key_suffix}")
        
        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    
    # ── BOX PLOT ──
    if datos_boxplot:
        df_box = pd.DataFrame(datos_boxplot)
        
        fig_box = go.Figure()
        
        colores_box = ['#1565C0', '#C62828', '#2E7D32', '#F57F17', '#6A1B9A', '#00838F']
        
        for idx_f, fundo in enumerate(fundos):
            df_f = df_box[df_box['Fundo'] == fundo]['Error']
            if df_f.empty:
                continue
            
            color = colores_box[idx_f % len(colores_box)]
            
            fig_box.add_trace(go.Box(
                y=df_f,
                name=fundo,
                boxpoints='all',
                jitter=0.4,
                pointpos=0,
                marker=dict(color=color, size=5, opacity=0.6, line=dict(color='white', width=0.5)),
                line=dict(color=color, width=2),
                fillcolor=f'rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.15)',
                hovertemplate=f'<b>{fundo}</b><br>Error: %{{y:+.2f}}°C<extra></extra>'
            ))
        
        fig_box.add_hrect(y0=-1.5, y1=1.5, fillcolor='rgba(76, 175, 80, 0.08)', line_width=0)
        fig_box.add_hline(y=1.5, line=dict(color='#4CAF50', width=1.5, dash='dot'))
        fig_box.add_hline(y=-1.5, line=dict(color='#4CAF50', width=1.5, dash='dot'))
        
        fig_box.update_layout(
            height=420,
            title=dict(text=f"<b>📦 Distribución Error — {variable}</b>", font=dict(size=13, color='#1A1A1A'), x=0.0),
            xaxis=dict(title='Fundo', gridcolor='#CFD8DC'),
            yaxis=dict(
                title='Error (°C)',
                ticksuffix='°C',
                gridcolor='#CFD8DC',
                tickvals=[-3, -2, -1, 0, 1, 2, 3],
                range=[-3.5, 3.5],
            ),
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Arial', size=9, color='#333333'),
            showlegend=False,
            margin=dict(l=60, r=40, t=60, b=40),
        )
        
        st.plotly_chart(fig_box, use_container_width=True, key=f"fig_box_{variable}{key_suffix}")
        
        # ── GRÁFICO APILADO ──
        def categorizar_error(e):
            if e < -1.5:
                return '< -1.5°C'
            elif e < 0:
                return '-1.5 a 0°C'
            elif e <= 1.5:
                return '0 a 1.5°C'
            else:
                return '> 1.5°C'

        df_box['Categoria'] = df_box['Error'].apply(categorizar_error)

        categorias_orden = ['< -1.5°C', '-1.5 a 0°C', '0 a 1.5°C', '> 1.5°C']
        colores_cat = {
            '< -1.5°C': '#E53935',
            '-1.5 a 0°C': '#FFAB40',
            '0 a 1.5°C': '#66BB6A',
            '> 1.5°C': '#1E88E5',
        }
        
        tabla_pct = df_box.groupby(['Fundo', 'Categoria']).size().reset_index(name='N')
        total_por_fundo = tabla_pct.groupby('Fundo')['N'].transform('sum')
        tabla_pct['Pct'] = (tabla_pct['N'] / total_por_fundo * 100).round(1)
        
        fig_stack = go.Figure()
        
        fundos_orden = sorted(df_box['Fundo'].unique().tolist())
        
        for cat in categorias_orden:
            df_cat = tabla_pct[tabla_pct['Categoria'] == cat]
            
            pcts, ns = [], []
            for fundo in fundos_orden:
                row_cat = df_cat[df_cat['Fundo'] == fundo]
                if not row_cat.empty:
                    pcts.append(float(row_cat['Pct'].iloc[0]))
                    ns.append(int(row_cat['N'].iloc[0]))
                else:
                    pcts.append(0.0)
                    ns.append(0)
            
            fig_stack.add_trace(go.Bar(
                name=cat,
                x=fundos_orden,
                y=pcts,
                marker_color=colores_cat[cat],
                text=[f'{p:.1f}%' if p > 4 else '' for p in pcts],
                textposition='inside',
                textfont=dict(size=11, color='white', family='Arial'),
                customdata=ns,
                hovertemplate='<b>%{x}</b><br>' + cat + '<br>Días: %{customdata}<br>%{y:.1f}%<extra></extra>'
            ))
        
        fig_stack.update_layout(
            barmode='stack',
            height=380,
            title=dict(text=f"<b>📊 Rangos Error — {variable}</b>", font=dict(size=13), x=0.0),
            xaxis_title='Fundo',
            yaxis=dict(title='% de días', ticksuffix='%', range=[0, 100], gridcolor='#CFD8DC', dtick=25),
            legend=dict(orientation='h', y=1.02, x=0),
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=60, r=40, t=80, b=40),
        )
        
        st.plotly_chart(fig_stack, use_container_width=True, key=f"fig_stack_{variable}{key_suffix}")




def _score_combo(sub_fundo, variable, ventana, params_combo):
    """Entrena con params_combo y devuelve MAE/Precisión/MBE/Std para la ventana dada."""
    fecha_inicio = ventana['inicio']
    fecha_fin    = ventana['fin']
    fecha_corte  = ventana['fecha_corte']

    train = sub_fundo[sub_fundo['Fecha'] <= fecha_corte].copy()
    if len(train) < 30:
        return None

    serie_train = train[['Fecha', variable]].rename(columns={'Fecha': 'ds', variable: 'y'}).dropna()
    serie_train['ds'] = pd.to_datetime(serie_train['ds']).dt.tz_localize(None).dt.normalize()

    media, std = serie_train['y'].mean(), serie_train['y'].std()
    serie_train['y'] = serie_train['y'].clip(lower=media - 3*std, upper=media + 3*std)

    coef_clim = calcular_climatologia_armonica(
        serie_train[['ds', 'y']], params_combo['n_harmonics'], params_combo['clim_halflife_anios']
    )
    fechas_clim = pd.date_range(serie_train['ds'].min(), fecha_fin + pd.Timedelta(days=30), freq='D')
    clim_daily = pd.DataFrame({
        'ds': fechas_clim,
        'clim': predecir_climatologia_armonica(fechas_clim, coef_clim, params_combo['n_harmonics']).round(2)
    })

    serie_train = serie_train.merge(clim_daily, on='ds', how='left')
    serie_train['y_abs'] = serie_train['y']
    serie_train['y'] = (serie_train['y'] - serie_train['clim']).round(4)

    try:
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            m = Prophet(
                yearly_seasonality=params_combo['yearly_seasonality'],
                weekly_seasonality=False, daily_seasonality=False,
                seasonality_mode='additive',
                changepoint_prior_scale=params_combo['changepoint_prior_scale'],
                changepoint_range=params_combo['changepoint_range'],
                seasonality_prior_scale=params_combo['seasonality_prior_scale'],
                interval_width=params_combo['interval_width'],
            )
            m.add_seasonality(name='monthly', period=30.5, fourier_order=params_combo['fourier_monthly'])
            m.fit(serie_train[['ds', 'y']])
            dias_pred_period = (fecha_fin - fecha_inicio).days + 1
            future = m.make_future_dataframe(periods=dias_pred_period, freq='D')
            forecast = m.predict(future)
    except Exception:
        return None

    forecast['ds'] = pd.to_datetime(forecast['ds']).dt.tz_localize(None).dt.normalize()
    forecast = forecast.merge(clim_daily, on='ds', how='left')
    forecast['clim'] = forecast['clim'].ffill().bfill()
    forecast['yhat'] = (forecast['yhat'] + forecast['clim']).round(2)

    pred_periodo = forecast[
        (forecast['ds'] >= fecha_inicio) & (forecast['ds'] <= fecha_fin)
    ][['ds', 'yhat']]

    real_periodo = sub_fundo[
        (sub_fundo['Fecha'] >= fecha_inicio) & (sub_fundo['Fecha'] <= fecha_fin)
    ][['Fecha', variable]].rename(columns={'Fecha': 'ds', variable: 'Real'})

    comp = real_periodo.merge(pred_periodo, on='ds', how='inner')
    if comp.empty:
        return None

    # ── bias: tendencia reciente (capada) + estacional multi-año ──────
    insample = forecast[['ds', 'yhat']].merge(
        serie_train[['ds', 'y_abs']], on='ds', how='inner'
    ).sort_values('ds').reset_index(drop=True)
    insample['residuo'] = insample['y_abs'] - insample['yhat']

    ventana_bias = min(params_combo['bias_window_dias'] * 2, len(insample))
    ult = insample.tail(ventana_bias).reset_index(drop=True)
    ult['t'] = np.arange(len(ult))
    if len(ult) >= 5:
        b_slope, a_int = np.polyfit(ult['t'], ult['residuo'], 1)
    else:
        b_slope, a_int = 0.0, (ult['residuo'].mean() if len(ult) else 0.0)

    insample['anio'] = insample['ds'].dt.year
    mes_obj = fecha_inicio.month
    datos_mes_prev = insample[
        (insample['ds'].dt.month == mes_obj) & (insample['anio'] < fecha_inicio.year)
    ]
    if len(datos_mes_prev) >= 15:
        edad = fecha_inicio.year - datos_mes_prev['anio']
        pesos = np.exp(-edad / 1.5)
        bias_est = float(np.average(datos_mes_prev['residuo'], weights=pesos))
    else:
        bias_est = 0.0

    comp = comp.copy()
    comp['h'] = (comp['ds'] - pd.Timestamp(fecha_corte)).dt.days
    h_ef = np.minimum(comp['h'].clip(lower=0), H_MAX_PENDIENTE)
    comp['ajuste_enfen'] = comp['ds'].apply(lambda f: obtener_ajuste_enfen(f.month, f.year, variable))
    comp['bias_h'] = a_int + b_slope * (len(ult) - 1) + b_slope * h_ef + bias_est + comp['ajuste_enfen']
    comp['pred_corr'] = comp['yhat'] + comp['bias_h']
    comp['error']     = comp['Real'] - comp['pred_corr']

    return {
        'MAE'      : round(comp['error'].abs().mean(), 2),
        'Precision': round((comp['error'].abs() <= 1.5).mean() * 100, 1),
        'MBE'      : round(comp['error'].mean(), 2),
        'Std'      : round(comp['error'].std(), 2),
    }

def grid_search_hiperparams(dia, ventana, grid):
    fundos    = dia['Fundo'].unique().tolist()
    variables = ['Tmax', 'Tmin']

    nombres = list(grid.keys())
    combos  = list(itertools.product(*grid.values()))

    total = len(fundos) * len(variables) * len(combos)
    prog  = st.progress(0, text="Grid search...")
    c = 0
    filas = []

    for fundo in fundos:
        sub_fundo = dia[dia['Fundo'] == fundo].copy().reset_index(drop=True)
        for variable in variables:
            for combo in combos:
                params_combo = {**PROPHET_PARAMS, **dict(zip(nombres, combo))}
                r = _score_combo(sub_fundo, variable, ventana, params_combo)
                c += 1
                prog.progress(c / total, text=f"{fundo} — {variable} — {dict(zip(nombres, combo))}")
                if r is None:
                    continue
                filas.append({'Fundo': fundo, 'Variable': variable, **dict(zip(nombres, combo)), **r})

    prog.empty()
    return pd.DataFrame(filas)           

def generar_range_plot(variable, historico_df, prediccion_df, fundo_name, row, fig):
    """Agrega Range Plot con Error Bars (barras Tmin-Tmax)."""
    
    # Range Plot de datos reales
    if not historico_df.empty:
        fechas = historico_df['Fecha'].tolist()
        tmax = historico_df['Tmax'].tolist()
        tmin = historico_df['Tmin'].tolist()
        smooth = historico_df[f'{variable}_smooth'].tolist()
        et = historico_df['ET'].tolist()
        
        # Calcular error bars (distancia de smooth a Tmax y Tmin)
        error_plus = [t - s for t, s in zip(tmax, smooth)]
        error_minus = [s - t for t, s in zip(tmin, smooth)]
        
        fig.add_trace(go.Scatter(
            x=fechas,
            y=smooth,
            mode='lines+markers',
            line=dict(color='#000000', width=2),
            marker=dict(size=5, color='#00AA00', line=dict(color='#000000', width=1)),
            error_y=dict(
                type='data',
                symmetric=False,
                array=error_plus,
                arrayminus=error_minus,
                color='#00AA00',
                thickness=2,
                width=3
            ),
            name=f'{variable} REAL',
            showlegend=True,
            legendgroup=fundo_name,
            hovertemplate='%{x|%d/%b/%Y}<br>' + f'{variable}: %{{y:.1f}}°C<br>ET: %{{customdata:.1f}}<extra>{fundo_name}</extra>',
            customdata=et
        ), row=row, col=1)
    
    # Range Plot de predicción
    if not prediccion_df.empty:
        fechas_pred = prediccion_df['ds'].tolist()
        yhat = prediccion_df['yhat'].tolist()
        upper = prediccion_df['yhat_upper'].tolist()
        lower = prediccion_df['yhat_lower'].tolist()
        
        # Calcular error bars para predicción
        error_plus_pred = [u - y for u, y in zip(upper, yhat)]
        error_minus_pred = [y - l for y, l in zip(yhat, lower)]
        
        fig.add_trace(go.Scatter(
            x=fechas_pred,
            y=yhat,
            mode='lines+markers',
            line=dict(color='#FF0000', width=2),
            marker=dict(size=5, color='#FF9800', line=dict(color='#FF0000', width=1)),
            error_y=dict(
                type='data',
                symmetric=False,
                array=error_plus_pred,
                arrayminus=error_minus_pred,
                color='#FF9800',
                thickness=2,
                width=3
            ),
            name=f'Predicción +{len(fechas_pred)}d',
            showlegend=True,
            legendgroup=fundo_name,
            hovertemplate='%{x|%d/%b/%Y}<br>Pred: %{y:.1f}°C<extra>Prophet</extra>'
        ), row=row, col=1)

def generar_figura(variable, dia, media_mensual, q1_mensual, q3_mensual,
                   dias_pred, forecasts_cache, dias_vista=30,
                   dias_pred_mostrar=None, tipo_viz='linea',
                   media_mensual2=None, q1_mensual2=None, q3_mensual2=None,
                   label_clim2=None):
    """Generación de figura optimizada."""
    fundos = dia['Fundo'].unique().tolist()
    n = len(fundos)
    
    if variable == 'Tmax':
        fig_bg, banda_color, clim_color, color_real, pred_color = (
            BG_TMAX, BANDA_TMAX, CLIM_TMAX, REAL_TMAX_COLOR, PRED_COLOR_TMAX
        )
    else:
        fig_bg, banda_color, clim_color, color_real, pred_color = (
            BG_TMIN, BANDA_TMIN, CLIM_TMIN, REAL_TMIN_COLOR, PRED_COLOR_TMIN
        )
    
    fig = make_subplots(
        rows=n, cols=1,
        shared_xaxes=False,
        vertical_spacing=0.08,
        subplot_titles=[f for f in fundos]
    )
    
    rows_export, rows_pred_export = [], []
    anio_actual = pd.Timestamp.today().year
    fecha_ini_anio = pd.Timestamp(year=anio_actual - 1, month=1, day=1)
    
    import calendar
    
    for i, fundo in enumerate(fundos):
        row = i + 1
        
        sub_full = dia[dia['Fundo'] == fundo].copy().reset_index(drop=True)
        if sub_full.empty:
            continue
        
        sub_full = sub_full[sub_full['Fecha'] >= fecha_ini_anio].copy().reset_index(drop=True)
        if sub_full.empty:
            st.warning(f"⚠️ Sin datos en {anio_actual} para **{fundo}**.")
            continue

        # ── Fechas clave ───────────────────────────────────────────
        # fecha_corte_real = último dato disponible (ej. 10 Jun 2026)
        fecha_corte_real    = pd.to_datetime(sub_full['Fecha'].max()).tz_localize(None).normalize()
        primer_dia_mes_actual = fecha_corte_real.replace(day=1)
        # fecha_fin_norm = FDM anterior (31 May) — solo para referencia/export
        fecha_fin_norm      = primer_dia_mes_actual - pd.Timedelta(days=1)
        # Último día del mes actual (30 Jun)
        ultimo_dia_mes_actual = pd.Timestamp(
            year=primer_dia_mes_actual.year,
            month=primer_dia_mes_actual.month,
            day=calendar.monthrange(
                primer_dia_mes_actual.year,
                primer_dia_mes_actual.month
            )[1]
        )

        # ── CAMBIO CLAVE: sub para graficar = TODOS los datos disponibles
        #    (incluyendo 1-10 Jun), NO cortado a FDM ──────────────────
        sub = sub_full.copy().reset_index(drop=True)

        if sub.empty:
            continue
        
        empresa = sub['Empresa'].iloc[0]
        
        zoom_ini = sub['Fecha'].min()
        zoom_fin = ultimo_dia_mes_actual + pd.Timedelta(days=2)
        
        # Splines con cache (sin cambios)
        fecha_ini_clima = sub['Fecha'].min()
        if hasattr(fecha_ini_clima, 'tz_localize'):
            fecha_ini_clima = pd.Timestamp(fecha_ini_clima).tz_localize(None).normalize()
        fecha_fin_clima = pd.Timestamp(zoom_fin)
        
        clim_media = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(media_mensual)
        )
        clim_q1 = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(q1_mensual)
        )
        clim_q3 = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(q3_mensual)
        )
        
        # Merge para export (igual que antes)
        hist_df = sub[['Empresa', 'Fundo', 'Fecha', variable, f'{variable}_smooth']].copy()
        hist_df = hist_df.merge(clim_media.rename(columns={'Valor': 'Clim_MEDIA'}), on='Fecha', how='left')
        hist_df = hist_df.merge(clim_q1.rename(columns={'Valor': 'Clim_Q1'}), on='Fecha', how='left')
        hist_df = hist_df.merge(clim_q3.rename(columns={'Valor': 'Clim_Q3'}), on='Fecha', how='left')
        hist_df['Variable'] = variable
        rows_export.append(hist_df)
        
        # Forecasts
        cache_entry = forecasts_cache.get((fundo, variable), {})
        forecast    = cache_entry.get('forecast', pd.DataFrame())
        mae         = cache_entry.get('mae', None)
        mae_por_dia = cache_entry.get('mae_por_dia', {})
        
        if not forecast.empty:
            forecast = forecast.copy()
            forecast['ds'] = pd.to_datetime(forecast['ds']).dt.tz_localize(None).dt.normalize()

            # ── CAMBIO CLAVE: predicción solo desde el día siguiente al
            #    último dato real hasta fin de mes ─────────────────────
            primer_dia_pred = fecha_corte_real + pd.Timedelta(days=1)

            pred_fc = forecast[
                (forecast['ds'] >= primer_dia_pred) &
                (forecast['ds'] <= ultimo_dia_mes_actual)
            ].reset_index(drop=True)
        else:
            pred_fc = pd.DataFrame()

        # Ajuste de residuos (sin cambios)
        if not pred_fc.empty and len(sub) >= 14:
            ventana     = min(21, len(sub))
            sub_reciente = sub.tail(ventana).copy().reset_index(drop=True)
            residuos    = (sub_reciente[variable] - sub_reciente[f'{variable}_smooth']).values

            from scipy.ndimage import gaussian_filter1d
            residuos_suaves = gaussian_filter1d(residuos, sigma=0.8)

            n_pred   = len(pred_fc)
            n_res    = len(residuos_suaves)
            indices  = [ii % n_res for ii in range(n_pred)]
            patron   = residuos_suaves[indices]

            std_hist = residuos_suaves.std()
            std_pred = pred_fc['yhat'].std()
            factor   = min(1.0, std_hist / (std_pred + 1e-6)) * 0.6

            pred_fc = pred_fc.copy()
            pred_fc['yhat']       = (pred_fc['yhat']       + patron * factor).round(2)
            pred_fc['yhat_lower'] = (pred_fc['yhat_lower'] + patron * factor).round(2)
            pred_fc['yhat_upper'] = (pred_fc['yhat_upper'] + patron * factor).round(2)

            bias_val = st.session_state.get(
                'bias_correccion', {}
            ).get((fundo, variable), 0.0)

            if bias_val != 0.0:
                pred_fc['yhat']       = (pred_fc['yhat']       + bias_val).round(2)
                pred_fc['yhat_lower'] = (pred_fc['yhat_lower'] + bias_val).round(2)
                pred_fc['yhat_upper'] = (pred_fc['yhat_upper'] + bias_val).round(2)

        lg = fundo
        
        # ── Trazas climatología Guadalupe (fija) — sin cambios ────────
        fig.add_trace(go.Scatter(
            x=clim_q3['Fecha'], y=clim_q3['Valor'],
            mode='lines', line=dict(width=0),
            showlegend=False, hoverinfo='skip', legendgroup=lg
        ), row=row, col=1)
        fig.add_trace(go.Scatter(
            x=clim_q1['Fecha'], y=clim_q1['Valor'],
            mode='lines', line=dict(width=0),
            fill='tonexty', fillcolor=_hex_to_rgba(banda_color, 0.55),
            showlegend=(i == 0), legendgroup=lg,
            name=f'Q1-Q3 SENAMHI Guadalupe (±{DELTA_Q}°C)',
            hovertemplate='%{x|%d/%b}<br>Q1: %{y:.1f}°C<extra></extra>'
        ), row=row, col=1)
        
        fig.add_trace(go.Scatter(
            x=clim_media['Fecha'], y=clim_media['Valor'],
            mode='lines', line=dict(color=clim_color, width=2.8),
            showlegend=(i == 0), legendgroup=lg,
            name=f'{variable} climatología SENAMHI Guadalupe',
            hovertemplate='%{x|%d/%b}<br>Clim Guadalupe: %{y:.1f}°C<extra></extra>'
        ), row=row, col=1)

        # ── Segunda climatología dinámica (sin cambios) ───────────────
        if media_mensual2 is not None and label_clim2:
            clim_media2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(media_mensual2)
            )
            clim_q1_2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(q1_mensual2)
            )
            clim_q3_2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(q3_mensual2)
            )

            fig.add_trace(go.Scatter(
                x=clim_q3_2['Fecha'], y=clim_q3_2['Valor'],
                mode='lines', line=dict(width=0),
                showlegend=False, hoverinfo='skip', legendgroup=lg
            ), row=row, col=1)
            fig.add_trace(go.Scatter(
                x=clim_q1_2['Fecha'], y=clim_q1_2['Valor'],
                mode='lines', line=dict(width=0),
                fill='tonexty',
                fillcolor=_hex_to_rgba(
                    BANDA2_TMAX if variable == 'Tmax' else BANDA2_TMIN, 0.15
                ),
                showlegend=(i == 0), legendgroup=lg,
                name=f'Q1-Q3 SENAMHI {label_clim2} (±{DELTA_Q}°C)',
                hovertemplate='%{x|%d/%b}<br>Q1: %{y:.1f}°C<extra></extra>'
            ), row=row, col=1)

            fig.add_trace(go.Scatter(
                x=clim_media2['Fecha'], y=clim_media2['Valor'],
                mode='lines',
                line=dict(color='rgba(255,111,0,0.70)', width=1.8, dash='dot'),
                showlegend=(i == 0), legendgroup=lg,
                name=f'{variable} climatología SENAMHI {label_clim2}',
                hovertemplate='%{x|%d/%b}<br>Clim ' + label_clim2 + ': %{y:.1f}°C<extra></extra>'
            ), row=row, col=1)

        # ── Datos reales — AHORA incluye datos del mes actual ─────────
        if tipo_viz == 'candlestick':
            if not pred_fc.empty:
                if dias_pred_mostrar is not None and dias_pred_mostrar < len(pred_fc):
                    pred_fc = pred_fc.iloc[:dias_pred_mostrar].copy()
                generar_range_plot(variable, sub, pred_fc, fundo, row, fig)
            else:
                generar_range_plot(variable, sub, pd.DataFrame(), fundo, row, fig)
        else:
            fig.add_trace(go.Scatter(
                x=sub['Fecha'], y=sub[f'{variable}_smooth'],
                mode='lines+markers',
                line=dict(color=color_real, width=1.5),
                marker=dict(size=5, color='white', line=dict(color=color_real, width=1.0)),
                customdata=sub['ET'],
                showlegend=(i == 0), legendgroup=lg,
                name=f'{variable} REAL',
                hovertemplate='%{x|%d/%b/%Y}<br>' + f'{variable}: %{{y:.1f}}°C<br>ET: %{{customdata:.1f}}<extra>{fundo}</extra>'
            ), row=row, col=1)

        # Anotación último valor real (ahora apunta a fecha_corte_real)
        ult = sub.iloc[-1]
        ult_real_val = ult[variable]
        fig.add_annotation(
            x=ult['Fecha'], y=ult_real_val,
            text=f"<b>{ult_real_val:.1f}°C</b>",
            showarrow=True, arrowhead=2, arrowcolor=color_real, arrowwidth=1.0,
            ax=14, ay=-18, font=dict(size=9, color=color_real),
            row=row, col=1
        )
        
        # ── CAMBIO CLAVE: línea vertical en el último dato real ───────
        #    (antes era fecha_fin_norm = 31 Mayo)
        fig.add_vline(
            x=fecha_corte_real,
            line=dict(color='#546E7A', width=1.2, dash='dot'),
            row=row, col=1
        )
        
        # ── Predicciones modo línea ────────────────────────────────────
        if tipo_viz == 'linea' and not pred_fc.empty:
            if dias_pred_mostrar is not None and dias_pred_mostrar < len(pred_fc):
                pred_fc = pred_fc.iloc[:dias_pred_mostrar].copy()
            
            # Conector desde último dato real → primer día predicho
            ult_smooth = float(ult[f'{variable}_smooth'])
            
            fig.add_trace(go.Scatter(
                x=[ult['Fecha'], pred_fc['ds'].iloc[0]],
                y=[ult_smooth, float(pred_fc['yhat'].iloc[0])],
                mode='lines',
                line=dict(
                    color=f'rgba({int(pred_color[1:3],16)},{int(pred_color[3:5],16)},{int(pred_color[5:7],16)},0.70)',
                    width=1.5, dash='dot'
                ),
                showlegend=False, hoverinfo='skip', legendgroup=lg
            ), row=row, col=1)
            
            # Banda IC 95%
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat_lower'].tolist(),
                mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'),
                showlegend=False, hoverinfo='skip', legendgroup=lg, name='_lower'
            ), row=row, col=1)
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat_upper'].tolist(),
                mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'),
                fill='tonexty', fillcolor=_hex_to_rgba(pred_color, 0.20),
                showlegend=(i == 0), name='IC 95%',
                legendgroup=lg, hoverinfo='skip'
            ), row=row, col=1)
            
            # Línea predicción
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat'].tolist(),
                mode='lines',
                line=dict(
                    color=f'rgba({int(pred_color[1:3],16)},{int(pred_color[3:5],16)},{int(pred_color[5:7],16)},0.70)',
                    width=1.8, dash='dot'
                ),
                showlegend=(i == 0), legendgroup=lg,
                name='Predicción mes siguiente',
                hovertemplate='%{x|%d/%b/%Y}<br>Pred: %{y:.1f}°C<extra>Predicción</extra>'
            ), row=row, col=1)
            
            # Anotación último valor predicho
            ult_pred_dt  = pred_fc['ds'].iloc[-1]
            ult_pred_val = float(pred_fc['yhat'].iloc[-1])
            fig.add_annotation(
                x=ult_pred_dt, y=ult_pred_val,
                text=f"<b>{ult_pred_val:.1f}°C</b>",
                showarrow=True, arrowhead=2, arrowcolor=pred_color, arrowwidth=1.0,
                ax=-28, ay=-18, font=dict(size=9, color=pred_color),
                row=row, col=1
            )
        
        # ── Título subplot (sin cambios) ───────────────────────────────
        bg_color = '#C62828' if variable == 'Tmax' else '#1565C0'
        fig.layout.annotations[i].update(
            text=f"  <b>{fundo}</b>  ",
            font=dict(size=12, color='#FFFFFF'),
            bgcolor=bg_color,
            borderpad=4,
            bordercolor=bg_color,
            borderwidth=2,
            x=0.0, xanchor='left'
        )
        
        xk = 'xaxis' if row == 1 else f'xaxis{row}'
        fig.layout[xk].update(
            range=[zoom_ini, zoom_fin],
            tickformat='%d/%b', ticklabelmode='period',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
            autorange=False, fixedrange=False,
            rangeslider=dict(visible=False)
        )
        yk = 'yaxis' if row == 1 else f'yaxis{row}'
        fig.layout[yk].update(
            ticksuffix='°C', gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
        )

    fig.update_layout(
        height=320 * n,
        paper_bgcolor=fig_bg,
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Arial', size=9, color='#333333'),
        title=dict(
            text=f"",
            font=dict(size=14, color='#1A1A1A'),
            x=0.0, xanchor='left', pad=dict(t=10, l=10)
        ),
        margin=dict(l=60, r=40, t=60, b=40),
        legend=dict(
            orientation='h', yanchor='top', y=-0.08,
            xanchor='left', x=0, font=dict(size=8),
            bgcolor='rgba(0,0,0,0)', borderwidth=0
        ),
        hovermode='x unified'
    )
    
    df_hist = pd.concat(rows_export, ignore_index=True) if rows_export else pd.DataFrame()
    df_pred = pd.concat(rows_pred_export, ignore_index=True) if rows_pred_export else pd.DataFrame()
    
    return fig, df_hist, df_pred

@st.cache_data(show_spinner=False)
def obtener_distritos_senamhi(_file_bytes, hoja='TMAX'):
    """Devuelve un set con los nombres de distrito (normalizados) presentes en el Excel SENAMHI."""
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break
    if header_row is None:
        return set()

    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
    if col_distrito is None:
        return set()

    distritos = df[col_distrito].dropna().astype(str).apply(_normalizar)
    distritos = distritos[(distritos != 'NAN') & (distritos != '')]
    return set(distritos.unique())


# ══════════════════════════════════════════════════════════════
# EXPORTAR EXCEL (igual que antes)
# ══════════════════════════════════════════════════════════════
def _generar_pred_horaria(df_pred: pd.DataFrame) -> pd.DataFrame:
    if df_pred.empty:
        return pd.DataFrame()
    
    rows = []
    grupos = df_pred.groupby(['Empresa', 'Fundo', 'Variable'])
    
    for (empresa, fundo, variable), grp in grupos:
        grp = grp.sort_values('Fecha').reset_index(drop=True)
        fechas = pd.to_datetime(grp['Fecha'])
        preds = grp['Pred'].values.astype(float)
        lows = grp['Pred_Low'].values.astype(float)
        highs = grp['Pred_High'].values.astype(float)
        
        n = len(fechas)
        x_ctrl = np.arange(n, dtype=float) * 24.0
        
        if n < 2:
            x_ctrl = np.array([0.0, 24.0])
            preds = np.array([preds[0], preds[0]])
            lows = np.array([lows[0], lows[0]])
            highs = np.array([highs[0], highs[0]])
        
        cs_pred = CubicSpline(x_ctrl, preds)
        cs_low = CubicSpline(x_ctrl, lows)
        cs_high = CubicSpline(x_ctrl, highs)
        
        horas_rango = pd.date_range(
            start=fechas.iloc[0].normalize(),
            end=fechas.iloc[-1].normalize() + pd.Timedelta(hours=23),
            freq='h'
        )
        
        for h_dt in horas_rango:
            x_h = (h_dt - fechas.iloc[0].normalize()).total_seconds() / 3600.0
            x_h = float(np.clip(x_h, x_ctrl[0], x_ctrl[-1]))
            hora = h_dt.hour
            
            base = float(cs_pred(x_h))
            lo = float(cs_low(x_h))
            hi = float(cs_high(x_h))
            ampl = max((hi - lo) / 4.0, 0.0)
            
            if variable == 'Tmax':
                offset = ampl * np.sin(2 * np.pi * (hora - 5) / 24)
            else:
                offset = -ampl * np.sin(2 * np.pi * (hora - 5) / 24)
            
            rows.append({
                'Empresa': empresa,
                'Fundo': fundo,
                'Fecha': h_dt.strftime('%Y/%m/%d'),
                'Hora': h_dt.strftime('%H:00'),
                'FechaHora': h_dt,
                'Variable': variable,
                'max_pred': round(base + offset, 1) if variable == 'Tmax' else None,
                'min_pred': round(base + offset, 1) if variable == 'Tmin' else None,
                'Tipo': 'PREDICCION',
            })
    
    return pd.DataFrame(rows)

def exportar_excel(df_hist: pd.DataFrame, df_pred: pd.DataFrame) -> bytes:
    HDR_HIST = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
    HDR_HORA = PatternFill('solid', start_color='1B5E20', end_color='1B5E20')
    HDR_DET = PatternFill('solid', start_color='37474F', end_color='37474F')
    REAL_F = PatternFill('solid', start_color='FFFF00', end_color='FFFF00')
    PRED_F = PatternFill('solid', start_color='FFE0B2', end_color='FFE0B2')
    ALT_F = PatternFill('solid', start_color='F3F4F6', end_color='F3F4F6')
    
    HDR_FNT_W = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    REAL_FNT = Font(name='Arial', size=8, color='000000')
    PRED_FNT = Font(name='Arial', size=8, color='BF360C', bold=True)
    CENT = Alignment(horizontal='center', vertical='center')
    
    def _autowidth(ws, max_w=22):
        for col in ws.columns:
            maxl = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxl + 3, max_w)
    
    wb = Workbook()
    
    # HOJA 1: Por Fechas
    ws_dia = wb.active
    ws_dia.title = 'Por_Fechas'
    ws_dia.sheet_properties.tabColor = '1E3A5F'
    
    filas_dia = []
    # Obtener fundos de ambos DataFrames, validando que no estén vacíos
    fundos_hist = list(df_hist['Fundo'].unique()) if not df_hist.empty and 'Fundo' in df_hist.columns else []
    fundos_pred = list(df_pred['Fundo'].unique()) if not df_pred.empty and 'Fundo' in df_pred.columns else []
    fundos_todos = sorted(set(fundos_hist + fundos_pred))
    
    for fundo in fundos_todos:
        h = df_hist[df_hist['Fundo'] == fundo].copy() if not df_hist.empty and 'Fundo' in df_hist.columns else pd.DataFrame()
        p = df_pred[df_pred['Fundo'] == fundo].copy() if not df_pred.empty and 'Fundo' in df_pred.columns else pd.DataFrame()
        
        empresa = (
            h['Empresa'].iloc[0] if not h.empty and 'Empresa' in h.columns else
            (p['Empresa'].iloc[0] if not p.empty and 'Empresa' in p.columns else '')
        )
        
        if 'Tmax' in h.columns and 'Tmin' in h.columns:
            hist_piv = (
                h[['Fecha', 'Tmax', 'Tmin']].drop_duplicates()
                 .rename(columns={'Tmax': 'max', 'Tmin': 'min'})
            )
        else:
            hist_piv = pd.DataFrame()
        
        if 'Pred' in p.columns:
            p_max = p[p['Variable'] == 'Tmax'][['Fecha', 'Pred']].rename(columns={'Pred': 'max pred'})
            p_min = p[p['Variable'] == 'Tmin'][['Fecha', 'Pred']].rename(columns={'Pred': 'min pred'})
            pred_piv = (
                p_max.merge(p_min, on='Fecha', how='outer') if not p_max.empty else
                pd.DataFrame(columns=['Fecha', 'max pred', 'min pred'])
            )
        else:
            pred_piv = pd.DataFrame()
        
        # Merge seguro: validar que ambos tengan 'Fecha' antes de hacer merge
        if not hist_piv.empty and not pred_piv.empty and 'Fecha' in hist_piv.columns and 'Fecha' in pred_piv.columns:
            tabla = hist_piv.merge(pred_piv, on='Fecha', how='outer')
        elif not hist_piv.empty:
            tabla = hist_piv.copy()
        elif not pred_piv.empty:
            tabla = pred_piv.copy()
        else:
            tabla = pd.DataFrame()
        
        if not tabla.empty:
            tabla['Fecha'] = pd.to_datetime(tabla['Fecha'])
            tabla = tabla.sort_values('Fecha').reset_index(drop=True)
            tabla['Empresa'] = empresa
            tabla['Fundo'] = fundo
            
            def _status(r):
                v = r.get('max', np.nan)
                try:
                    return 'real' if pd.notna(v) and not np.isnan(float(v)) else 'predicha'
                except Exception:
                    return 'predicha'
            tabla['status'] = tabla.apply(_status, axis=1)
            filas_dia.append(tabla)
    
    tabla_dia = pd.concat(filas_dia, ignore_index=True) if filas_dia else pd.DataFrame()
    tabla_dia = tabla_dia.sort_values(['Fundo', 'Fecha']).reset_index(drop=True)
    
    cols_dia = ['Empresa', 'Fundo', 'Fecha', 'max', 'min', 'max pred', 'min pred', 'status']
    for ci, h_txt in enumerate(cols_dia, 1):
        c = ws_dia.cell(row=1, column=ci)
        c.value = h_txt; c.fill = HDR_HIST
        c.font = HDR_FNT_W; c.alignment = CENT; c.border = THIN
    
    if not tabla_dia.empty:
        for ri, row_data in tabla_dia.iterrows():
            er = ri + 2
            is_real = row_data.get('status', 'predicha') == 'real'
            fill = REAL_F if is_real else PRED_F
            font = REAL_FNT if is_real else PRED_FNT
            
            def _v(key):
                val = row_data.get(key, np.nan)
                try:
                    return round(float(val), 1) if pd.notna(val) else None
                except Exception:
                    return None
            
            vals = [
                row_data.get('Empresa', ''),
                row_data.get('Fundo', ''),
                row_data['Fecha'].strftime('%Y/%m/%d') if pd.notna(row_data.get('Fecha')) else '',
                _v('max'), _v('min'),
                _v('max pred'), _v('min pred'),
                row_data.get('status', 'predicha'),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws_dia.cell(row=er, column=ci)
                c.value = v; c.fill = fill; c.font = font
                c.alignment = CENT; c.border = THIN
    
    _autowidth(ws_dia)
    ws_dia.freeze_panes = 'A2'
    
    # HOJA 2: Por Horas (reducida)
    ws_hora = wb.create_sheet('Por_Horas')
    ws_hora.sheet_properties.tabColor = '1B5E20'
    
    pred_hora_df = _generar_pred_horaria(df_pred)
    
    if not pred_hora_df.empty:
        sub_max = pred_hora_df[pred_hora_df['Variable'] == 'Tmax'][
            ['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora', 'max_pred']
        ].copy()
        sub_min = pred_hora_df[pred_hora_df['Variable'] == 'Tmin'][
            ['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora', 'min_pred']
        ].copy()
        
        if not sub_max.empty and not sub_min.empty:
            tabla_h = sub_max.merge(
                sub_min, on=['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora'], how='outer'
            )
        elif not sub_max.empty:
            tabla_h = sub_max.copy()
            if 'min_pred' not in tabla_h.columns:
                tabla_h['min_pred'] = None
        elif not sub_min.empty:
            tabla_h = sub_min.copy()
            if 'max_pred' not in tabla_h.columns:
                tabla_h['max_pred'] = None
        else:
            tabla_h = pd.DataFrame()
        
        if not tabla_h.empty:
            tabla_h = tabla_h.sort_values(['Fundo', 'FechaHora']).reset_index(drop=True)
            
            cols_hora = ['Empresa', 'Fundo', 'Fecha', 'Hora', 'max pred', 'min pred', 'status']
            for ci, h_txt in enumerate(cols_hora, 1):
                c = ws_hora.cell(row=1, column=ci)
                c.value = h_txt; c.fill = HDR_HORA
                c.font = HDR_FNT_W; c.alignment = CENT; c.border = THIN
            
            for ri, row_data in tabla_h.iterrows():
                er = ri + 2
                max_v = row_data.get('max_pred')
                min_v = row_data.get('min_pred')
                vals_h = [
                    row_data.get('Empresa', ''),
                    row_data.get('Fundo', ''),
                    row_data.get('Fecha', ''),
                    row_data.get('Hora', ''),
                    round(float(max_v), 1) if max_v is not None and not pd.isna(max_v) else None,
                    round(float(min_v), 1) if min_v is not None and not pd.isna(min_v) else None,
                    'predicha',
                ]
                for ci, v in enumerate(vals_h, 1):
                    c = ws_hora.cell(row=er, column=ci)
                    c.value = v; c.fill = PRED_F; c.font = PRED_FNT
                    c.alignment = CENT; c.border = THIN
            
            _autowidth(ws_hora)
            ws_hora.freeze_panes = 'A2'
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
# ══════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def _leer_normales_desde_disco(path: str):
    with open(path, 'rb') as f:
        return f.read()

@st.cache_data(show_spinner=False)
def cargar_geojson_peru():
    """Descarga GeoJSON de distritos de Perú desde GitHub."""
    url = "https://raw.githubusercontent.com/juaneladio/peru-geojson/master/peru_distrital_simple.geojson"
    try:
        import urllib.request
        with urllib.request.urlopen(url) as r:
            return json.loads(r.read().decode('utf-8'))
    except Exception as e:
        st.warning(f"⚠️ No se pudo cargar el mapa de Perú: {e}")
        return None
def agregar_leyenda_vertical(m, vmin, vmax, colors, caption, n_ticks=6,
                              position='topleft',
                              height_px=260, width_px=22,
                              font_caption=13, font_ticks=12,
                              decimales=0):
    """
    Leyenda de gradiente VERTICAL (vmax arriba, vmin abajo) como control
    nativo de Leaflet -> se mantiene visible también en pantalla completa.
    position: 'topleft' | 'topright' | 'bottomleft' | 'bottomright'
    """
    from branca.element import MacroElement
    from jinja2 import Template

    colores_arriba_abajo = list(reversed(colors))
    n = len(colores_arriba_abajo)
    stops = [f"{c} {i/(n-1)*100:.0f}%" for i, c in enumerate(colores_arriba_abajo)]
    gradiente = ", ".join(stops)

    ticks_html = ""
    for i in range(n_ticks):
        frac = i / (n_ticks - 1)
        valor = vmax - frac * (vmax - vmin)
        formato = f"{{valor:.{decimales}f}}"
        valor_str = formato.format(valor=valor)
        ticks_html += (
            f'<div style="position:absolute;left:{width_px + 6}px;top:{frac*100:.1f}%;'
            f'transform:translateY(-50%);font-size:{font_ticks}px;font-weight:600;'
            f'color:#263238;white-space:nowrap;">{valor_str}</div>'
        )

    leyenda = MacroElement()
    leyenda._template = Template(f"""
    {{% macro script(this, kwargs) %}}
    var legend_{{{{ this.get_name() }}}} = L.control({{position: '{position}'}});
    legend_{{{{ this.get_name() }}}}.onAdd = function (map) {{
        var div = L.DomUtil.create('div', 'info legend');
        div.style.background = 'rgba(255,255,255,0.88)';
        div.style.padding = '10px 32px 10px 10px';
        div.style.borderRadius = '10px';
        div.style.boxShadow = '0 2px 8px rgba(0,0,0,0.25)';
        div.style.fontFamily = "'Segoe UI', Arial, sans-serif";
        div.innerHTML = `
            <div style="font-size:{font_caption}px;font-weight:700;color:#263238;
                         margin-bottom:8px;max-width:130px;">{caption}</div>
            <div style="position:relative;width:{width_px}px;height:{height_px}px;
                         background:linear-gradient(to bottom, {gradiente});
                         border-radius:4px;border:1px solid #B0BEC5;">
                {ticks_html}
            </div>
        `;
        return div;
    }};
    legend_{{{{ this.get_name() }}}}.addTo({{{{ this._parent.get_name() }}}});
    {{% endmacro %}}
    """)

    m.add_child(leyenda)

COLORES_TEMP_SENAMHI = [
    '#0D47A1', '#1976D2', '#42A5F5', '#81D4FA',
    '#FFF59D', '#FFB74D', '#FB8C00', '#E64A19', '#B71C1C',
]
DIAS_ET_PROMEDIO = 7
COLORES_ET = ['#E1F5FE', '#81D4FA', '#29B6F6', '#0288D1', '#01579B']

DIAS_RIESGO_VENTANA = 7  # horizonte "próx. días" para Tmax/Tmin predicha

RIESGO_COLOR = {
    'Bajo': '#4CAF50', 'Medio': '#FFC107', 'Alto': '#FB8C00',
    'Muy alto': '#C62828', 'Sin datos': '#90A4AE',
}
# ✅ NUEVO - Azul suave (Bajo) → Azul fuerte (Muy alto)
RIESGO_COLOR_ET = {
    'Bajo': '#E1F5FE',      # Azul MÁS SUAVE (casi blanco)
    'Medio': '#81D4FA',     # Azul claro
    'Alto': '#42A5F5',      # Azul intermedio
    'Muy alto': '#0D47A1',  # Azul MUY FUERTE (oscuro)
    'Sin datos': '#90A4AE',
}
def calcular_umbrales_riesgo(media_clim):
    """
    P50/P75/P95 vía PERCENTIL.INC (interpolación lineal, igual que Excel)
    sobre los 12 valores de climatología mensual SENAMHI Guadalupe.
    """
    p50, p75, p95 = np.percentile(media_clim, [50, 75, 95])
    return {'p50': float(p50), 'p75': float(p75), 'p95': float(p95)}


def clasificar_riesgo(valor, umbrales):
    if valor is None or (isinstance(valor, float) and np.isnan(valor)):
        return 'Sin datos'
    if valor <= umbrales['p50']:
        return 'Bajo'
    elif valor <= umbrales['p75']:
        return 'Medio'
    elif valor <= umbrales['p95']:
        return 'Alto'
    else:
        return 'Muy alto'


def calcular_metricas_riesgo(forecasts_cache, fundo, variable, umbrales,
                              dias_ventana=DIAS_RIESGO_VENTANA):
    """Promedio Tmax/Tmin próx. N días + nivel de riesgo según `umbrales`."""
    cache_entry = forecasts_cache.get((fundo, variable), {}) if forecasts_cache else {}
    forecast = cache_entry.get('forecast', pd.DataFrame())

    if forecast.empty:
        return {'valor_prox': None, 'nivel': 'Sin datos',
                'color': RIESGO_COLOR['Sin datos'], 'dias_ventana': None}

    fc = forecast.sort_values('ds').reset_index(drop=True)
    ventana = fc.head(dias_ventana)
    valor_prox = round(float(ventana['yhat'].mean()), 1)
    nivel = clasificar_riesgo(valor_prox, umbrales)

    return {
        'valor_prox': valor_prox,
        'nivel': nivel,
        'color': RIESGO_COLOR.get(nivel, RIESGO_COLOR['Sin datos']),
        'dias_ventana': len(ventana),
    }

def calcular_riesgo_real_fundo(dia, fundo, variable):
    """
    Nivel de Riesgo del último día con dato real (Tmax o Tmin) de un fundo,
    comparado contra los percentiles P50/P75/P95 (PERCENTIL.INC) de su
    propio historial.
    """
    sub = dia[dia['Fundo'] == fundo].sort_values('Fecha')
    serie = sub[variable].dropna()

    if serie.empty:
        return {'valor_actual': None, 'fecha': None, 'nivel': 'Sin datos',
                'color': RIESGO_COLOR['Sin datos'], 'umbrales': None}

    umbrales = calcular_umbrales_riesgo(serie.values)  # P50/P75/P95 de su propia serie
    valor_actual = float(serie.iloc[-1])
    fecha_actual = sub.loc[serie.index[-1], 'Fecha']
    nivel = clasificar_riesgo(valor_actual, umbrales)

    return {
        'valor_actual': round(valor_actual, 1),
        'fecha': fecha_actual,
        'nivel': nivel,
        'color': RIESGO_COLOR.get(nivel, RIESGO_COLOR['Sin datos']),
        'umbrales': umbrales,
    }
@st.cache_data(show_spinner=False)
@st.cache_data(show_spinner=False)
def calcular_riesgo_et_fundo(df_et_mensual, fundo):
    """Nivel de Riesgo de ET mensual usando escala de AZULES."""
    sub_et = df_et_mensual[df_et_mensual['Fundo'] == fundo]['EToPromedioDiaria'].dropna()
    
    if sub_et.empty:
        return {
            'valor_actual': None,
            'nivel': 'Sin datos',
            'color': RIESGO_COLOR_ET['Sin datos'],  # ✅ CAMBIO
            'umbrales': None
        }
    
    umbrales = {
        'p25': float(np.percentile(sub_et.values, 25)),  # Q1 ← NUEVO
        'p50': float(np.percentile(sub_et.values, 50)),  # Q2 (Mediana)
        'p75': float(np.percentile(sub_et.values, 75)),  # Q3
        'p95': float(np.percentile(sub_et.values, 95)),  # Extra
    }
    
    valor_actual = float(sub_et.iloc[-1])
    nivel = clasificar_riesgo(valor_actual, umbrales)
    
    return {
        'valor_actual': round(valor_actual, 2),
        'nivel': nivel,
        'color': RIESGO_COLOR_ET.get(nivel, RIESGO_COLOR_ET['Sin datos']),  # ✅ CAMBIO
        'umbrales': umbrales,
    }

def agregar_leyenda_riesgo(m, riesgos_por_fundo=None, unidad='°C', position='bottomleft',
                            titulo='Riesgo térmico'):
    """Leyenda con las 4 categorías de Riesgo (colores)."""
    from branca.element import MacroElement
    from jinja2 import Template

    filas_color = [
        ('Bajo',     RIESGO_COLOR['Bajo']),
        ('Medio',    RIESGO_COLOR['Medio']),
        ('Alto',     RIESGO_COLOR['Alto']),
        ('Muy alto', RIESGO_COLOR['Muy alto']),
    ]

    items_html = ""
    for nivel, color in filas_color:
        items_html += (
            f'<div style="display:flex;align-items:center;gap:6px;margin:2px 0;">'
            f'<span style="width:11px;height:11px;border-radius:50%;'
            f'background:{color};display:inline-block;"></span>'
            f'<span style="font-size:11px;color:#263238;">{nivel}</span></div>'
        )

    leyenda = MacroElement()
    leyenda._template = Template(f"""
    {{% macro script(this, kwargs) %}}
    var legend_{{{{ this.get_name() }}}} = L.control({{position: '{position}'}});
    legend_{{{{ this.get_name() }}}}.onAdd = function (map) {{
        var div = L.DomUtil.create('div', 'info legend');
        div.style.background = 'rgba(255,255,255,0.88)';
        div.style.padding = '8px 12px';
        div.style.borderRadius = '8px';
        div.style.boxShadow = '0 2px 8px rgba(0,0,0,0.25)';
        div.style.fontFamily = "'Segoe UI', Arial, sans-serif";
        div.innerHTML = `
            <div style="font-size:11px;font-weight:700;color:#263238;margin-bottom:4px;">{titulo}</div>
            {items_html}
        `;
        return div;
    }};
    legend_{{{{ this.get_name() }}}}.addTo({{{{ this._parent.get_name() }}}});
    {{% endmacro %}}
    """)
    m.add_child(leyenda)

def agregar_leyenda_riesgo_et(m, position='bottomleft'):
    """Leyenda con las 4 categorías de Riesgo ET (con colores AZULES)."""
    from branca.element import MacroElement
    from jinja2 import Template

    filas_color = [
        ('Bajo',     RIESGO_COLOR_ET['Bajo']),
        ('Medio',    RIESGO_COLOR_ET['Medio']),
        ('Alto',     RIESGO_COLOR_ET['Alto']),
        ('Muy alto', RIESGO_COLOR_ET['Muy alto']),
    ]

    items_html = ""
    for nivel, color in filas_color:
        items_html += (
            f'<div style="display:flex;align-items:center;gap:6px;margin:2px 0;">'
            f'<span style="width:11px;height:11px;border-radius:50%;'
            f'background:{color};display:inline-block;border:1px solid white;"></span>'
            f'<span style="font-size:11px;color:#263238;font-weight:600;">{nivel}</span></div>'
        )

    leyenda = MacroElement()
    leyenda._template = Template(f"""
    {{% macro script(this, kwargs) %}}
    var legend_{{{{ this.get_name() }}}} = L.control({{position: '{position}'}});
    legend_{{{{ this.get_name() }}}}.onAdd = function (map) {{
        var div = L.DomUtil.create('div', 'info legend');
        div.style.background = 'rgba(255,255,255,0.92)';
        div.style.padding = '8px 12px';
        div.style.borderRadius = '8px';
        div.style.boxShadow = '0 2px 8px rgba(0,0,0,0.25)';
        div.style.fontFamily = "'Segoe UI', Arial, sans-serif";
        div.innerHTML = `
            <div style="font-size:11px;font-weight:700;color:#00838F;margin-bottom:4px;">⚠️ Riesgo ET</div>
            {items_html}
        `;
        return div;
    }};
    legend_{{{{ this.get_name() }}}}.addTo({{{{ this._parent.get_name() }}}});
    {{% endmacro %}}
    """)
    m.add_child(leyenda)
def agregar_leyenda_et_discreta(m, et_por_fundo_sorted, vmin, vmax, position='topleft'):
    """
    Leyenda DISCRETA mostrando solo los 3 fundos + sus valores ET.
    et_por_fundo_sorted: lista [(fundo, valor), ...]
    """
    from branca.element import MacroElement
    from jinja2 import Template

    filas_html = ""
    for fundo, valor in et_por_fundo_sorted:
        # Normalizar valor al rango [0,1] para obtener color del colormap
        if vmax > vmin:
            norm_val = (valor - vmin) / (vmax - vmin)
        else:
            norm_val = 0.5
        
        # Obtener color de la escala ET
        # Interpolación lineal en COLORES_ET
        idx_color = norm_val * (len(COLORES_ET) - 1)
        idx_bajo = int(np.floor(idx_color))
        idx_alto = int(np.ceil(idx_color))
        if idx_bajo == idx_alto:
            color = COLORES_ET[idx_bajo]
        else:
            # Interpolación simple (bonus)
            frac = idx_color - idx_bajo
            color = COLORES_ET[min(idx_bajo, len(COLORES_ET)-1)]
        
        filas_html += (
            f'<div style="display:flex;align-items:center;gap:8px;margin:4px 0;">'
            f'<span style="width:14px;height:14px;border-radius:3px;'
            f'background:{color};display:inline-block;border:1px solid white;"></span>'
            f'<span style="font-size:11px;color:#263238;font-weight:600;">'
            f'{fundo}: <b>{valor:.2f}</b> mm</span></div>'
        )

    leyenda = MacroElement()
    leyenda._template = Template(f"""
    {{% macro script(this, kwargs) %}}
    var legend_{{{{ this.get_name() }}}} = L.control({{position: '{position}'}});
    legend_{{{{ this.get_name() }}}}.onAdd = function (map) {{
        var div = L.DomUtil.create('div', 'info legend');
        div.style.background = 'rgba(255,255,255,0.92)';
        div.style.padding = '10px 14px';
        div.style.borderRadius = '8px';
        div.style.boxShadow = '0 2px 8px rgba(0,0,0,0.25)';
        div.style.fontFamily = "'Segoe UI', Arial, sans-serif";
        div.innerHTML = `
            <div style="font-size:12px;font-weight:700;color:#00838F;margin-bottom:6px;">
                💧 ET Semanal
            </div>
            {filas_html}
            <div style="font-size:9px;color:#78909C;margin-top:6px;font-style:italic;">
                Últimos {DIAS_ET_PROMEDIO} días
            </div>
        `;
        return div;
    }};
    legend_{{{{ this.get_name() }}}}.addTo({{{{ this._parent.get_name() }}}});
    {{% endmacro %}}
    """)
    
    m.add_child(leyenda)
def calcular_et_suma_semanal_fundos(dia, dias=DIAS_ET_PROMEDIO):
    """{fundo: suma de ET (mm) de los últimos `dias` días con ET>0}."""
    resultado = {}
    for fundo in dia['Fundo'].unique():
        sub = dia[dia['Fundo'] == fundo].sort_values('Fecha')
        et_vals = sub['ET'].dropna()
        et_vals = et_vals[et_vals > 0].tail(dias)
        if not et_vals.empty:
            resultado[fundo] = round(float(et_vals.sum()), 2)
    return resultado

def crear_colormap_y_leyenda_et(et_por_fundo):
    """Crea colormap continuo + retorna datos para leyenda discreta (3 fundos)."""
    from branca.colormap import LinearColormap

    if not et_por_fundo:
        return None, None

    valores = np.array(list(et_por_fundo.values()))
    vmin = float(np.floor(valores.min() * 10) / 10)
    vmax = float(np.ceil(valores.max() * 10) / 10)
    if vmin == vmax:
        vmin = max(0.0, vmin - 0.5)
        vmax = vmax + 0.5

    colormap = LinearColormap(
        colors=COLORES_ET,
        vmin=vmin, vmax=vmax,
        caption=f'ET suma semanal (últ. {DIAS_ET_PROMEDIO}d) [mm]'
    )
    
    # ✅ NUEVO: retorna datos para leyenda discreta
    leyenda_data = {
        'vmin': vmin,
        'vmax': vmax,
        'fundos': sorted(et_por_fundo.items())  # [(fundo, valor), ...]
    }
    
    return colormap, leyenda_data

def generar_mapa_distritos(geojson, distrito_fijo, distrito_din, variable,
                            dia=None, mes_sel=None,
                            media_clim_fijo=None, media_clim_din=None,
                            label_clim2=None, modulos_kmz=None,
                            distritos_senamhi=None,
                            temps_distritos=None, temp_colormap=None,
                            forecasts_cache=None):
    if geojson is None:
        return None
    import folium
    from branca.colormap import LinearColormap
    color_fijo = '#C62828' if variable == 'Tmax' else '#1565C0'
    color_din  = '#FF6F00'
    dist_fijo_norm = _normalizar(distrito_fijo)
    dist_din_norm  = _normalizar(distrito_din) if distrito_din else None
    m = folium.Map(
        location=[-7.27, -79.45],
        zoom_start=7,
        tiles=None,
        control_scale=False,
        zoom_control=False,
    )
    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri WorldImagery', name='Satellite', overlay=False, control=False,
    ).add_to(m)
    folium.TileLayer(
        tiles='https://services.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}',
        attr='Esri Labels', name='Labels', overlay=True, control=False, opacity=0.7,
    ).add_to(m)
    from folium.plugins import Fullscreen
    Fullscreen(
        position='topright', title='Pantalla completa',
        title_cancel='Salir', force_separate_button=True,
    ).add_to(m)

    # ── Distritos SENAMHI — pintados por temperatura climatológica (fondo) ──
    if distritos_senamhi:
        for feat in geojson['features']:
            p = feat['properties']
            nombre = _normalizar(p['NOMBDIST'])
            if nombre == dist_fijo_norm:
                continue
            if dist_din_norm and nombre == dist_din_norm:
                continue
            if nombre not in distritos_senamhi:
                continue
            label_txt = p['NOMBDIST'] + ' — ' + p.get('NOMBDEP', '')
            valor_temp = (temps_distritos or {}).get(nombre)
            if valor_temp is not None and temp_colormap is not None:
                fill_color   = temp_colormap(valor_temp)
                fill_opacity = 0.70
                borde        = '#FFFFFF'
                grosor       = 0.8
                tooltip_txt  = (
                    f"🌡️ {label_txt}<br>"
                    f"{variable} clim.: {valor_temp:.1f}°C "
                    f"({MESES_RAW[mes_sel-1] if mes_sel else ''})"
                )
            else:
                fill_color   = '#90A4AE'
                fill_opacity = 0.06
                borde        = '#90A4AE'
                grosor       = 1
                tooltip_txt  = f"📡 {label_txt}"
            folium.GeoJson(
                feat,
                style_function=(lambda f, fc=fill_color, op=fill_opacity,
                                 bc=borde, w=grosor: {
                    'fillColor': fc,
                    'color': bc,
                    'weight': w,
                    'dashArray': None if op > 0.1 else '3,3',
                    'fillOpacity': op,
                }),
                highlight_function=lambda f: {'fillOpacity': 0.75, 'weight': 1.5},
                tooltip=folium.Tooltip(tooltip_txt, sticky=False),
            ).add_to(m)
        if temp_colormap is not None:
            agregar_leyenda_vertical(
                m,
                vmin=temp_colormap.vmin, vmax=temp_colormap.vmax,
                colors=COLORES_TEMP_SENAMHI,
                caption=temp_colormap.caption,
                position='bottomleft',

            )

    # ── Distritos SENAMHI fijo / dinámico ──────────────────────────────
    def _style_fijo(f):
        return {'fillColor': color_fijo, 'color': 'white', 'weight': 1.5, 'fillOpacity': 0.55}
    def _style_din(f):
        return {'fillColor': color_din, 'color': 'white', 'weight': 1.5, 'fillOpacity': 0.55}
    for feat in geojson['features']:
        p      = feat['properties']
        nombre = _normalizar(p['NOMBDIST'])
        es_fijo = nombre == dist_fijo_norm
        es_din  = dist_din_norm and nombre == dist_din_norm
        if not es_fijo and not es_din:
            continue
        label_txt = p['NOMBDIST'] + ' — ' + p.get('NOMBDEP', '')
        if mes_sel is not None:
            valor_clim = None
            if es_fijo and media_clim_fijo is not None:
                valor_clim = float(media_clim_fijo[mes_sel - 1])
            elif es_din and media_clim_din is not None:
                valor_clim = float(media_clim_din[mes_sel - 1])
            if valor_clim is not None:
                label_txt += (
                    f"<br>🌡️ {variable} clim.: {valor_clim:.1f}°C "
                    f"({MESES_RAW[mes_sel-1]})"
                )
        folium.GeoJson(
            feat,
            style_function=_style_fijo if es_fijo else _style_din,
            tooltip=folium.Tooltip(label_txt, sticky=False),
        ).add_to(m)
        geom = feat['geometry']
        coords = []
        if geom['type'] == 'Polygon':
            coords = geom['coordinates'][0]
        elif geom['type'] == 'MultiPolygon':
            coords = geom['coordinates'][0][0]
        if coords:
            lon_c = sum(c[0] for c in coords) / len(coords)
            lat_c = sum(c[1] for c in coords) / len(coords)
            folium.Marker(
                location=[lat_c, lon_c],
                popup=folium.Popup(label_txt, max_width=200),
                icon=folium.Icon(
                    color='red' if es_fijo else 'orange',
                    icon='tint' if variable == 'Tmax' else 'info-sign',
                    prefix='glyphicon',
                ),
            ).add_to(m)
    if modulos_kmz:
        from shapely.geometry import Polygon as ShPolygon
        from shapely.ops import unary_union
        for mod in modulos_kmz:
            mod['fundo'] = asignar_fundo(mod['fundo_aq'], mod['mod_n'])
        for mod in modulos_kmz:
            folium.Polygon(
                locations=mod['coords'],
                color='#FFFFFF',
                weight=2,
                fill=True,
                fill_color='#FFFFFF',
                fill_opacity=0.15,
            ).add_to(m)
        grupos_fundo = {}
        for mod in modulos_kmz:
            anillo = [(lon, lat) for lat, lon in mod['coords']]
            try:
                poly = ShPolygon(anillo)
                if not poly.is_valid:
                    poly = poly.buffer(0)
                grupos_fundo.setdefault(mod['fundo'], []).append(poly)
            except Exception:
                continue
        COLOR_FUNDO = {
            'Arena Azul':   '#1565C0',
            'Vivadis':      '#2E7D32',
            'Santa Teresa': '#F57F17',
            'Ayllu Allpa':  '#6A1B9A',
        }
        for fundo_nombre, polys in grupos_fundo.items():
            union = unary_union(polys)
            geoms = [union] if union.geom_type == 'Polygon' else list(union.geoms)
            centroide = union.centroid
            lat_c, lon_c = centroide.y, centroide.x
            color_acento = COLOR_FUNDO.get(fundo_nombre, '#37474F')

            # ── Calcular riesgo ──
            riesgo = (
                calcular_riesgo_real_fundo(dia, fundo_nombre, variable)
                if dia is not None else
                {'valor_actual': None, 'fecha': None, 'nivel': 'Sin datos',
                 'color': RIESGO_COLOR['Sin datos'], 'umbrales': None}
            )
            color_riesgo = riesgo['color']
            nivel = riesgo['nivel']
            valor_txt = f"{riesgo['valor_actual']:.1f}°C" if riesgo['valor_actual'] is not None else "—"
            fecha_txt = riesgo['fecha'].strftime('%d/%b') if riesgo['fecha'] is not None else ""

            tooltip_html = (
                f"<b>{fundo_nombre}</b><br>"
                f"{variable} ({fecha_txt}): <b>{valor_txt}</b><br>"
                f"<span style='color:{color_riesgo};font-weight:700;'>Riesgo: {nivel.upper()}</span>"
            )

            popup_html = f"<b>{fundo_nombre}</b><br>"
            if riesgo['valor_actual'] is not None:
                popup_html += f"{variable} ({fecha_txt}): <b>{riesgo['valor_actual']:.1f}°C</b><br>"
            if riesgo['umbrales'] is not None:
                u = riesgo['umbrales']
                popup_html += f"<small>Percentiles propios — P50={u['p50']:.1f} | P75={u['p75']:.1f} | P95={u['p95']:.1f}</small><br>"
            popup_html += f"<b>Riesgo: <span style='color:{color_riesgo}'>{nivel.upper()}</span></b>"

            # ✅ NUEVO: Colorear polígonos KMZ por riesgo
            for geom in geoms:
                if geom.is_empty:
                    continue
                coords = [[lat, lon] for lon, lat in geom.exterior.coords]
                
                folium.Polygon(
                    locations=coords,
                    color='white',              # borde blanco
                    weight=2.5,
                    fill=True,
                    fill_color=color_riesgo,   # ✨ RELLENO por riesgo
                    fill_opacity=0.60,          # semitransparente
                    tooltip=folium.Tooltip(tooltip_html, sticky=False),
                    popup=folium.Popup(popup_html, max_width=250),
                ).add_to(m)

            # ── Marker con nombre (sin cambios) ──
            folium.map.Marker(
                [lat_c, lon_c],
                icon=folium.DivIcon(
                    icon_size=(0, 0),
                    icon_anchor=(0, 0),
                    html=(
                        '<div style="'
                        'transform: translate(-50%, -50%);'
                        'display: inline-flex; align-items: center; gap: 6px;'
                        'background: rgba(255,255,255,0.92);'
                        'padding: 4px 10px 4px 8px;'
                        'border-radius: 14px;'
                        f'border-left: 4px solid {color_acento};'
                        'box-shadow: 0 2px 6px rgba(0,0,0,0.25);'
                        'font-family: "Segoe UI", Arial, sans-serif;'
                        'font-size: 12px; font-weight: 700;'
                        'color: #263238; letter-spacing: 0.02em;'
                        'white-space: nowrap; pointer-events: auto;'
                        '">'
                        f'{fundo_nombre}'
                        '</div>'
                    )
                ),
                tooltip=folium.Tooltip(tooltip_html, sticky=False),
                popup=folium.Popup(popup_html, max_width=250),
            ).add_to(m)

        # ── Agregar leyenda de riesgo ──
        agregar_leyenda_riesgo(m, position='bottomleft')

    return m

def generar_mapa_et(modulos_kmz, et_por_fundo, colormap_et=None, df_et_mensual=None, riesgos_et=None):
    """Mapa con los módulos KMZ coloreados según ET promedio por fundo."""
    import folium
    from shapely.geometry import Polygon as ShPolygon
    from shapely.ops import unary_union

    m = folium.Map(
        location=[-7.65, -79.36],
        zoom_start=12,
        tiles=None,
        control_scale=False,
        zoom_control=False,
    )

    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri WorldImagery', name='Satellite', overlay=False, control=False,
    ).add_to(m)

    folium.TileLayer(
        tiles='https://services.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}',
        attr='Esri Labels', name='Labels', overlay=True, control=False, opacity=0.7,
    ).add_to(m)

    from folium.plugins import Fullscreen
    Fullscreen(
        position='topright', title='Pantalla completa',
        title_cancel='Salir', force_separate_button=True,
    ).add_to(m)

    if not modulos_kmz:
        return m

    for mod in modulos_kmz:
        if 'fundo' not in mod:
            mod['fundo'] = asignar_fundo(mod['fundo_aq'], mod['mod_n'])

    grupos_fundo = {}
    for mod in modulos_kmz:
        anillo = [(lon, lat) for lat, lon in mod['coords']]
        try:
            poly = ShPolygon(anillo)
            if not poly.is_valid:
                poly = poly.buffer(0)
            grupos_fundo.setdefault(mod['fundo'], []).append(poly)
        except Exception:
            continue

    bounds_all = []

    for fundo_nombre, polys in grupos_fundo.items():
        union = unary_union(polys)
        geoms = [union] if union.geom_type == 'Polygon' else list(union.geoms)

        et_val = et_por_fundo.get(fundo_nombre)
        if et_val is not None and colormap_et is not None:
            fill_color   = colormap_et(et_val)
            fill_opacity = 0.55
        else:
            fill_color   = '#90A4AE'
            fill_opacity = 0.15

        tooltip_txt = f"<b>{fundo_nombre}</b>"
        tooltip_txt += (
            f"<br>💧 ET semanal: {et_val:.2f} mm" if et_val is not None
            else "<br>Sin datos de ET"
        )

        minx, miny, maxx, maxy = union.bounds
        bounds_all.append((miny, minx, maxy, maxx))

        centroide = union.centroid
        et_txt = f"{et_val:.2f} mm/día" if et_val is not None else "—"
        
        # ✅ USA riesgo pre-calculado (no lo recalcules)
        if riesgos_et and fundo_nombre in riesgos_et:
            riesgo_et = riesgos_et[fundo_nombre]
        else:
            riesgo_et = calcular_riesgo_et_fundo(df_et_mensual, fundo_nombre)
        
        color_riesgo_et = riesgo_et['color']
        nivel_et = riesgo_et['nivel']
        valor_et_txt = f"{riesgo_et['valor_actual']:.2f} mm/día" if riesgo_et['valor_actual'] is not None else "—"

        tooltip_et = (
            f"<b>{fundo_nombre}</b><br>"
            f"💧 ET promedio: <b>{valor_et_txt}</b><br>"
            f"<span style='color:{color_riesgo_et};font-weight:700;'>Riesgo: {nivel_et.upper()}</span>"
        )

        # ✅ CONSTRUIR POPUP COMPLETO CON TODOS LOS DATOS
        popup_et_html_content = f"<b>{fundo_nombre}</b><br>"
        
        # ET semanal
        if et_val is not None:
            popup_et_html_content += f"💧 <b>ET semanal:</b> {et_val:.2f} mm<br>"
        
        # ET promedio diario
        popup_et_html_content += f"💧 <b>ET promedio:</b> {valor_et_txt}<br>"
        
        # Percentiles/Cuartiles
        if riesgo_et['umbrales'] is not None:
            u = riesgo_et['umbrales']
            popup_et_html_content += (
                f"<small><b>Percentiles propios:</b><br>"
                f"Q1={u.get('p25', 'N/A'):.2f} | Q2={u['p50']:.2f} | Q3={u['p75']:.2f} | P95={u['p95']:.2f}</small><br>"
            )
        
        # Riesgo ET
        popup_et_html_content += (
            f"<b>Riesgo: <span style='color:{color_riesgo_et};font-weight:700;'>"
            f"{nivel_et.upper()}</span></b>"
        )
        
        # Envolver con fondo semi-transparente
        popup_et_html = f"""
        <div style="background-color: rgba(255,255,255,0.75); padding: 8px; border-radius: 4px; font-family: Arial, sans-serif;">
        {popup_et_html_content}
        </div>
        """

        # ✅ UN SOLO POLÍGONO CON TODOS LOS DATOS
        for geom in geoms:
            if geom.is_empty:
                continue
            coords = [[lat, lon] for lon, lat in geom.exterior.coords]
            
            folium.Polygon(
                locations=coords,
                color='white', 
                weight=2,
                fill=True,
                fill_color=color_riesgo_et,
                fill_opacity=0.55,
                tooltip=folium.Tooltip(tooltip_et, sticky=False),
                popup=folium.Popup(popup_et_html, max_width=280),
            ).add_to(m)

        # Marker con nombre del fundo
        folium.map.Marker(
            [centroide.y, centroide.x],
            icon=folium.DivIcon(
                icon_size=(0, 0), 
                icon_anchor=(0, 0),
                html=(
                    f'<div style="'
                    f'font-size:10px;'
                    f'font-weight:700;'
                    f'color:#263238;'
                    f'text-align:center;'
                    f'pointer-events:none;">'
                    f'{fundo_nombre}'
                    f'</div>'
                )
            ),
            tooltip=folium.Tooltip(et_txt, sticky=False)
        ).add_to(m)

    if bounds_all:
        lat_min = min(b[0] for b in bounds_all)
        lon_min = min(b[1] for b in bounds_all)
        lat_max = max(b[2] for b in bounds_all)
        lon_max = max(b[3] for b in bounds_all)
        m.fit_bounds([[lat_min, lon_min], [lat_max, lon_max]])

    # Leyendas
    if colormap_et is not None and et_por_fundo:
        agregar_leyenda_et_discreta(
            m, 
            sorted(et_por_fundo.items()),
            vmin=colormap_et.vmin,
            vmax=colormap_et.vmax,
            position='topleft'
        )
    
    agregar_leyenda_riesgo_et(m, position='bottomleft')

    return m

def generar_tab_et(dia, forecasts_cache, dias_pred_ui, modulos_kmz=None, 
                   riesgos_et_pre=None):  # ← RECIBE pre-calculado
    """Gráfico ET diaria + predicción mes siguiente por fundo."""
    import calendar

    fundos = dia['Fundo'].unique().tolist()
    anio_actual = pd.Timestamp.today().year
    fecha_ini_anio = pd.Timestamp(year=anio_actual - 1, month=1, day=1)

    n = len(fundos)
    altura_et = 300 * n   # mismo alto que usa fig.update_layout(height=300*n, ...)

    col_graf_et, col_mapa_et = st.columns([3, 1])
    fig = make_subplots(
        rows=n, cols=1,
        shared_xaxes=False,
        vertical_spacing=0.08,
        subplot_titles=[f for f in fundos]
    )

    for i, fundo in enumerate(fundos):
        row = i + 1

        sub = dia[
            (dia['Fundo'] == fundo) &
            (dia['Fecha'] >= fecha_ini_anio)
        ].copy().reset_index(drop=True)

        if sub.empty or 'ET' not in sub.columns:
            continue

        et_df = sub[['Fecha', 'ET']].dropna().copy()
        if et_df.empty or et_df['ET'].sum() == 0:
            continue

        empresa  = sub['Empresa'].iloc[0]
        et_total = et_df['ET'].sum().round(2)
        et_prom  = et_df['ET'].mean().round(2)

        # ── Fechas clave ───────────────────────────────────────
        fecha_fin_raw         = pd.to_datetime(et_df['Fecha'].max()).normalize()
        primer_dia_mes_actual = fecha_fin_raw.replace(day=1)
        fecha_fin_norm        = primer_dia_mes_actual - pd.Timedelta(days=1)  # 31 mayo
        ultimo_dia_mes_actual = pd.Timestamp(
            year=primer_dia_mes_actual.year,
            month=primer_dia_mes_actual.month,
            day=calendar.monthrange(
                primer_dia_mes_actual.year,
                primer_dia_mes_actual.month
            )[1]
        )  # 30 junio

        # ── Datos reales — TODOS los disponibles ──────────────
        et_real = et_df[et_df['Fecha'] <= fecha_fin_norm].copy()


        # ── Línea real ─────────────────────────────────────────
        fig.add_trace(go.Scatter(
            x=et_real['Fecha'],
            y=et_real['ET'],
            mode='lines+markers',
            line=dict(color='#1565C0', width=1.8),
            marker=dict(size=4, color='white', line=dict(color='#1565C0', width=1.2)),
            name='ET real',
            showlegend=(i == 0),
            fill='tozeroy',
            fillcolor='rgba(21, 101, 192, 0.08)',
            hovertemplate='%{x|%d/%b/%Y}<br>ET: %{y:.2f} mm<extra>' + fundo + '</extra>'
        ), row=row, col=1)


        # ✅ NUEVO - Agregar líneas de cuartiles
        if not et_df.empty:
            # Calcular cuartiles
            q1 = et_df['ET'].quantile(0.25)
            q2 = et_df['ET'].quantile(0.50)
            q3 = et_df['ET'].quantile(0.75)
            p95 = et_df['ET'].quantile(0.95)
            
            # Q1 (Percentil 25) - línea azul clara
            fig.add_hline(
                y=q1,
                line=dict(color='#B3E5FC', width=1, dash='dash'),
                annotation_text=f"Q1: {q1:.2f}",
                annotation_position="right",
                annotation_font=dict(size=7, color='#0288D1'),
                row=row, col=1
            )
            
           
            
            # Q3 (Percentil 75) - línea azul oscura
            fig.add_hline(
                y=q3,
                line=dict(color='#01579B', width=1, dash='dash'),
                annotation_text=f"Q3: {q3:.2f}",
                annotation_position="right",
                annotation_font=dict(size=7, color='#01579B'),
                row=row, col=1
            )
            

        # ── Anotación último valor real ────────────────────────
        fig.add_annotation(
            x=et_real['Fecha'].iloc[-1],
            y=float(et_real['ET'].iloc[-1]),
            text=f"<b>{float(et_real['ET'].iloc[-1]):.2f}</b>",
            showarrow=True, arrowhead=2,
            arrowcolor='#1565C0', arrowwidth=1.0,
            ax=14, ay=-18,
            font=dict(size=9, color='#1565C0'),
            row=row, col=1
        )

        # ── Línea vertical 31 mayo ─────────────────────────────
        fig.add_vline(
            x=fecha_fin_norm,
            line=dict(color='#546E7A', width=1.2, dash='dot'),
            row=row, col=1
        )

        # ── Predicción Prophet ─────────────────────────────────
        cache_entry = forecasts_cache.get((fundo, 'ET'), {})
        forecast_et = cache_entry.get('forecast', pd.DataFrame())

        if not forecast_et.empty:
            forecast_et = forecast_et.copy()
            forecast_et['ds'] = pd.to_datetime(
                forecast_et['ds']
            ).dt.tz_localize(None).dt.normalize()

            # Mes completo desde 1 junio hasta 30 junio
            pred_et = forecast_et[
                (forecast_et['ds'] >= primer_dia_mes_actual) &
                (forecast_et['ds'] <= ultimo_dia_mes_actual)
            ].reset_index(drop=True)
            # ── Calcular amplitud histórica del mismo mes ──────────
            mes_actual = primer_dia_mes_actual.month
            et_mismo_mes = et_df[
                pd.to_datetime(et_df['Fecha']).dt.month == mes_actual
            ]['ET']

            if len(et_mismo_mes) >= 5:
                    std_historica = et_mismo_mes.std()
                    amp_max = et_mismo_mes.max()
                    amp_min = et_mismo_mes.min()
            else:
                    std_historica = et_df['ET'].std()
                    amp_max = et_df['ET'].max()
                    amp_min = et_df['ET'].min()

                # ← AQUÍ, fuera del else, siempre se ejecuta
            pred_et = pred_et.copy()
            ventana_patron = min(30, len(et_real))
            et_ventana = et_real['ET'].values[-ventana_patron:]
            media_ventana = et_ventana.mean()
            desviacion = et_ventana - media_ventana

            n_pred = len(pred_et)
            indices = [j % len(desviacion) for j in range(n_pred)]
            patron_ciclado = desviacion[indices]

            std_actual = np.std(desviacion)
            factor = std_historica / (std_actual + 1e-6)
            factor = min(factor, 1.5)

            pred_et['yhat'] = (
                pred_et['yhat'] + patron_ciclado * factor
            ).clip(lower=0).round(2)

            pred_et['yhat_upper'] = (pred_et['yhat'] + std_historica * 0.8).clip(upper=float(amp_max)).round(2)
            pred_et['yhat_lower'] = (pred_et['yhat'] - std_historica * 0.8).clip(lower=0.0).round(2)
            if not pred_et.empty:
                # Conector desde 31 mayo → 1 junio predicho
                et_val_ultimo = float(et_real['ET'].iloc[-1])
                fig.add_trace(go.Scatter(
                    x=[et_real['Fecha'].iloc[-1], pred_et['ds'].iloc[0]],
                    y=[et_val_ultimo, float(pred_et['yhat'].iloc[0])],
                    mode='lines',
                    line=dict(color='#43A047', width=1.8, dash='dash'),
                    showlegend=False, hoverinfo='skip'
                ), row=row, col=1)

                # Banda IC 95%
                fig.add_trace(go.Scatter(
                    x=pred_et['ds'].tolist() + pred_et['ds'].tolist()[::-1],
                    y=pred_et['yhat_upper'].tolist() + pred_et['yhat_lower'].tolist()[::-1],
                    fill='toself',
                    fillcolor='rgba(67, 160, 71, 0.12)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ), row=row, col=1)

                # Línea predicción
                fig.add_trace(go.Scatter(
                    x=pred_et['ds'],
                    y=pred_et['yhat'],
                    mode='lines+markers',
                    line=dict(color='#43A047', width=2.5, dash='dash'),
                    marker=dict(size=4, color='#43A047'),
                    name='ET predicha',
                    showlegend=(i == 0),
                    hovertemplate='%{x|%d/%b/%Y}<br>ET pred: %{y:.2f} mm<extra>Prophet</extra>'
                ), row=row, col=1)

                # Anotación último valor predicho
                fig.add_annotation(
                    x=pred_et['ds'].iloc[-1],
                    y=float(pred_et['yhat'].iloc[-1]),
                    text=f"<b>{float(pred_et['yhat'].iloc[-1]):.2f}</b>",
                    showarrow=True, arrowhead=2,
                    arrowcolor='#43A047', arrowwidth=1.0,
                    ax=-28, ay=-18,
                    font=dict(size=9, color='#43A047'),
                    row=row, col=1
                )

        fig.layout.annotations[i].update(
            text=f"<b>💧 {fundo}</b>",
            font=dict(size=13, color='#00838F'),
            bgcolor='rgba(0,0,0,0)',
            borderpad=4,
            bordercolor='rgba(0,0,0,0)',
            x=0.0, xanchor='left'
        )

        xk = 'xaxis' if row == 1 else f'xaxis{row}'
        fig.layout[xk].update(
            range=[
                et_real['Fecha'].min(),
                ultimo_dia_mes_actual + pd.Timedelta(days=2)
            ],
            tickformat='%d/%b',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
        )
        yk = 'yaxis' if row == 1 else f'yaxis{row}'
        fig.layout[yk].update(
            title='mm/día',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=True,
            rangemode='tozero'
        )

    fig.update_layout(
        height=300 * n,
        paper_bgcolor='#FFFFFF',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Arial', size=9, color='#333333'),
        title=dict(
            text="<b>💧 Evapotranspiración diaria — FUNDOS AQUANQA</b>",
            font=dict(size=14, color='#1A1A1A'),
            x=0.0, xanchor='left', pad=dict(t=10, l=10)
        ),
        margin=dict(l=60, r=40, t=60, b=40),
        legend=dict(
            orientation='h', yanchor='top', y=-0.05,
            xanchor='left', x=0, font=dict(size=8),
            bgcolor='rgba(0,0,0,0)', borderwidth=0
        ),
        hovermode='x unified'
    )

    with col_graf_et:
            st.plotly_chart(fig, use_container_width=True, key="fig_et_main")

    with col_mapa_et:
        et_por_fundo = calcular_et_suma_semanal_fundos(dia)
        colormap_et, leyenda_et_data = crear_colormap_y_leyenda_et(et_por_fundo)
        mapa_et = generar_mapa_et(modulos_kmz, et_por_fundo, colormap_et, 
                                   df_et_mensual, riesgos_et_pre)  # ← USA pre-calculado
        if mapa_et is not None:
            from streamlit_folium import st_folium
            st_folium(mapa_et, width=None, height=altura_et,
                    returned_objects=[], key='mapa_et')    
    # ── KPIs resumen ───────────────────────────────────────────
    st.markdown("#### 📊 Resumen ET por fundo")
    cols = st.columns(len(fundos))
    for i, fundo in enumerate(fundos):
        sub_et = dia[
            (dia['Fundo'] == fundo) &
            (dia['Fecha'] >= fecha_ini_anio)
        ]['ET'].dropna()
        if not sub_et.empty:
            cols[i].metric(
                fundo,
                f"{sub_et.sum():.1f} mm",
                f"Prom {sub_et.mean():.2f} mm/día"
            )




def asignar_fundo(fundo_aq, mod_n):
    if fundo_aq == 'AQ1':
        return 'Arena Azul'
    if fundo_aq == 'AQ2':
        if 1 <= mod_n <= 5:
            return 'Vivadis'
        if mod_n in (6, 7, 8, 9, 10, 11, 16, 17, 18):
            return 'Santa Teresa'
        return 'Ayllu Allpa'
    return 'Desconocido'


@st.cache_data(show_spinner=False)
def cargar_temperaturas_distritos(_file_bytes, hoja, mes_sel):
    """
    Devuelve dict {distrito_normalizado: temperatura_promedio_mes}
    para TODOS los distritos del Excel SENAMHI, para el mes dado.
    """
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break
    if header_row is None:
        return {}

    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
    if col_distrito is None:
        return {}

    mes_nombre = MESES_RAW[mes_sel - 1]
    col_mes = next((c for c in df.columns if _normalizar(c) == _normalizar(mes_nombre)), None)
    if col_mes is None:
        return {}

    df = df.copy()
    df['_dist_norm'] = df[col_distrito].astype(str).apply(_normalizar)
    df['_valor'] = pd.to_numeric(df[col_mes], errors='coerce')
    df = df[(df['_dist_norm'] != 'NAN') & (df['_dist_norm'] != '') & df['_valor'].notna()]

    return df.groupby('_dist_norm')['_valor'].mean().round(1).to_dict()


def crear_colormap_temperatura(temps_dict, variable, mes_sel):
    from branca.colormap import LinearColormap
    import numpy as np

    if not temps_dict:
        return None

    valores = np.array(list(temps_dict.values()))
    # Usar percentiles en vez de min/max absoluto → evita que la selva
    # (mucho más caliente) aplaste la escala de la costa
    vmin = np.floor(np.percentile(valores, 5))
    vmax = np.ceil(np.percentile(valores, 95))
    if vmin == vmax:
        vmin -= 1
        vmax += 1

    return LinearColormap(
    colors=COLORES_TEMP_SENAMHI,
    vmin=vmin, vmax=vmax,
    caption=f'{variable} climatología SENAMHI — {MESES_RAW[mes_sel-1]} [°C]'
)

def cargar_normales_dinamico(_file_bytes, distrito_sel: str):
    """
    Carga normales SENAMHI para el distrito seleccionado dinámicamente.
    Reemplaza completamente a cargar_normales_fijo().
    """
    try:
        # Función interna que usa distrito dinámico (no la constante DISTRITO_BUSCAR)
        def _cargar_hoja(hoja):
            raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
            header_row = None
            for idx, row in raw.iterrows():
                vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
                if 'DISTRITO' in vals:
                    header_row = idx
                    break
 
            if header_row is None:
                raise ValueError(f"No se encontró fila con 'DISTRITO' en hoja '{hoja}'.")
 
            df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
            df.columns = [str(c).strip() for c in df.columns]
 
            col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
            if col_distrito is None:
                raise ValueError("Columna DISTRITO no encontrada.")
 
            # Filtrar por distrito seleccionado dinámicamente
            df['_dist_norm'] = df[col_distrito].astype(str).apply(_normalizar)
            df_f = df[df['_dist_norm'].str.contains(distrito_sel, na=False)].copy()
 
            if df_f.empty:
                raise ValueError(f"No se encontraron datos para el distrito '{distrito_sel}'.")
 
            # Extraer valores mensuales
            meses_encontrados = {}
            for col in df.columns:
                col_norm = _normalizar(col)
                for mes in MESES_RAW:
                    if col_norm == _normalizar(mes) and mes not in meses_encontrados:
                        meses_encontrados[mes] = col
                        break
 
            meses_faltantes = [m for m in MESES_RAW if m not in meses_encontrados]
            if meses_faltantes:
                raise ValueError(f"Columnas de meses no encontradas: {meses_faltantes}")
 
            cols_meses = [meses_encontrados[m] for m in MESES_RAW]
            valores = df_f[cols_meses].apply(pd.to_numeric, errors='coerce').mean(axis=0).values
 
            col_nombre = next((c for c in df.columns if _normalizar(c) == 'NOMBRE ESTACION'), col_distrito)
            col_prov   = next((c for c in df.columns if _normalizar(c) == 'PROVINCIA'), col_distrito)
 
            return (
                valores,
                valores - DELTA_Q,
                valores + DELTA_Q,
                df_f[[col_nombre, col_prov, col_distrito]].drop_duplicates()
            )
 
        MEDIA_TMAX, Q1_TMAX, Q3_TMAX, estaciones_tmax = _cargar_hoja('TMAX')
        MEDIA_TMIN, Q1_TMIN, Q3_TMIN, _               = _cargar_hoja('TMIN')
 
        return MEDIA_TMAX, Q1_TMAX, Q3_TMAX, MEDIA_TMIN, Q1_TMIN, Q3_TMIN, estaciones_tmax
 
    except Exception as e:
        st.error(f"Error al cargar normales para '{distrito_sel}': {e}")
        return None
 

# ──── SIDEBAR MEJORADO ────
norm_bytes_sidebar = _leer_normales_desde_disco(NORMALES_PATH) if os.path.exists(NORMALES_PATH) else None
with st.sidebar:
    st.markdown("### 🌊 Monitor ENFEN")
    comunicado_actual = chequear_comunicado_enfen()

    if comunicado_actual is None:
        st.caption("⚠️ No se pudo verificar ENFEN (sin conexión o cambió el formato de la página).")
    else:
        ultimo_visto = _leer_ultimo_visto_enfen()
        if comunicado_actual['id'] != ultimo_visto.get('id'):
            st.warning(
                f"📢 **Nuevo comunicado ENFEN**\n\n"
                f"N°{comunicado_actual['numero']}-{comunicado_actual['anio']} "
                f"({comunicado_actual['fecha']})\n\n"
                f"Revisa si `AJUSTE_ENFEN` necesita actualizarse."
            )
            st.markdown(f"[📄 Ver comunicados ENFEN]({ENFEN_URL})")
            if st.button("✅ Marcar como revisado"):
                _guardar_ultimo_visto_enfen(comunicado_actual)
                st.rerun()
        else:
            st.caption(
                f"✅ Al día — último: N°{comunicado_actual['numero']}-{comunicado_actual['anio']} "
                f"({comunicado_actual['fecha']})"
            )
with st.sidebar:
    st.markdown("## ⚙️ CONFIGURACIÓN")

    with st.expander("📂 **Archivo meteorológico**", expanded=True):
        st.markdown("### 📊 **Archivo meteorológico**")
        st.info("✅ Leyendo: `assets/Metereologia_Prize.xlsx`")

    st.divider()

    st.markdown("### 🌿 **Fundos**")
    fundos_disponibles_placeholder = st.empty()

    st.divider()
    # Estado corrección BIAS
    bias_dict = st.session_state.get('bias_correccion', {})
    if bias_dict:
        st.markdown("### 🎯 Corrección BIAS")
        for (f, v), b in bias_dict.items():
            icono = '🔺' if b > 0 else '🔻'
            st.caption(f"{icono} {f} {v}: {b:+.2f}°C")
    else:
        st.caption("⚠️ Abre validación para activar BIAS")
    st.markdown("### 🌦️ **Climatología SENAMHI**")

    # ── Guadalupe siempre fijo ─────────────────────────────────
    st.info(f"📌 Referencia fija: **{DISTRITO_BUSCAR}**")

    # ── Selector dinámico (sin Guadalupe) ─────────────────────
    st.markdown("**Comparar con otra estación:**")

    if norm_bytes_sidebar is not None:
        catalogo = cargar_catalogo_normales(norm_bytes_sidebar, 'TMAX')

        if catalogo:
            sectores_lista = sorted(catalogo.keys())

            sector_sel = st.selectbox(
                "Sector",
                options=sectores_lista,
                index=0,
                key='sector_sel'
            )

            deptos_lista = sorted(catalogo.get(sector_sel, {}).keys())

            depto_sel = st.selectbox(
                "Departamento",
                options=deptos_lista,
                index=0,
                key='depto_sel'
            )

            distritos_lista = [DISTRITO_NINGUNA] + catalogo.get(sector_sel, {}).get(depto_sel, [])

            distrito_sel2 = st.selectbox(
                "Estación / Distrito",
                options=distritos_lista,
                index=0,   # ← "— Ninguna —" por defecto
                key='distrito_sel2'
            )

            if distrito_sel2 != DISTRITO_NINGUNA:
                st.success(f"✅ Comparando: **{distrito_sel2}**")
            else:
                st.caption("Sin comparación activa.")
        else:
            st.error("❌ No se pudo leer el catálogo.")
            distrito_sel2 = DISTRITO_NINGUNA
    else:
        st.error("❌ Normales SENAMHI no encontradas")
        distrito_sel2 = DISTRITO_NINGUNA


# ── GeoJSON Perú ──────────────────────────────────────────────
geojson_peru = cargar_geojson_peru()  # ← AQUÍ
# ── Polígonos KMZ de módulos (solo contorno de referencia) ──
distritos_senamhi = obtener_distritos_senamhi(norm_bytes_sidebar, 'TMAX') if norm_bytes_sidebar is not None else set()
# ── Polígonos KMZ de módulos (con DEBUG) ──
distritos_senamhi = obtener_distritos_senamhi(norm_bytes_sidebar, 'TMAX') if norm_bytes_sidebar is not None else set()

_kmz_bytes_modulos = download_kmz_from_github()
st.write(f"🔍 DEBUG: _kmz_bytes_modulos = {type(_kmz_bytes_modulos)}, tamaño = {len(_kmz_bytes_modulos) if _kmz_bytes_modulos else 'None'}")

kmz_polygons = load_kmz_bytes(_kmz_bytes_modulos) if _kmz_bytes_modulos else []
st.write(f"🔍 DEBUG: kmz_polygons = {len(kmz_polygons)} polígonos")

modulos_kmz = disolver_modulos(kmz_polygons) if kmz_polygons else []
st.write(f"🔍 DEBUG: modulos_kmz = {len(modulos_kmz)} módulos")

if not modulos_kmz:
    st.warning("⚠️ No se cargaron KMZ desde GitHub")

# ──── HEADER ────
st.markdown("""
<div class="header-bar">
    <h1>🌡️ Temperatura Fundos — Aquanqa</h1>
    <p>Climatología SENAMHI · Predicción Prophet · OPTIMIZADO</p>
</div>
""", unsafe_allow_html=True)

# ──── VALORES POR DEFECTO ────
dias_vista_num = 30  # Mostrar últimos 30 días
variable_sel = 'Ambas'  # Mostrar Tmax y Tmin
min_reg_ui = MIN_REGISTROS  # Registros mínimos por día (valor de constante)
dias_pred_ui = 30  # Días de predicción Prophet

# ──── CARGAR NORMALES ────────────────────────────────────────

# Guadalupe siempre fijo
normales_guadalupe = cargar_normales_dinamico(norm_bytes_sidebar, DISTRITO_BUSCAR)
if normales_guadalupe is None:
    st.stop()

MEDIA_TMAX, Q1_TMAX, Q3_TMAX, MEDIA_TMIN, Q1_TMIN, Q3_TMIN, estaciones_tmax = normales_guadalupe

# Normales dinámicas (solo si se seleccionó algo)
normales_din = None
if distrito_sel2 != DISTRITO_NINGUNA:
    normales_din = cargar_normales_dinamico(norm_bytes_sidebar, distrito_sel2)

if normales_din is not None:
    MEDIA_TMAX2, Q1_TMAX2, Q3_TMAX2, MEDIA_TMIN2, Q1_TMIN2, Q3_TMIN2, _ = normales_din
else:
    MEDIA_TMAX2 = Q1_TMAX2 = Q3_TMAX2 = None
    MEDIA_TMIN2 = Q1_TMIN2 = Q3_TMIN2 = None
    
def conectar_fabric():
    """Conecta a Fabric con MSAL en local e cloud."""
    import os
    import pyodbc
    import msal
    
    SQL_SERVER = st.secrets["SQL_SERVER"]
    SQL_DB = st.secrets["SQL_DB"]
    SQL_USER = st.secrets["SQL_USER"]
    
    # Detectar entorno
    es_streamlit_cloud = os.environ.get("STREAMLIT_SERVER_HEADLESS") == "true"
    
    try:
        app = msal.PublicClientApplication(
            client_id="04b07795-8ddb-461a-bbee-02f9e1bf7b46",
            authority="https://login.microsoftonline.com/common"
        )
        
        if es_streamlit_cloud:
            # ── CLOUD: Device Code Flow ──────────────────────────
            st.info("🔐 Autenticación Azure AD (Device Code)")
            
            flow = app.initiate_device_flow(
                scopes=["https://database.windows.net/.default"]
            )
            
            st.warning(
                f"📱 **Abre este link en otro dispositivo:**\n\n"
                f"[{flow.get('verification_uri')}]({flow.get('verification_uri')})\n\n"
                f"**Código:** `{flow.get('user_code')}`\n\n"
                f"Esperando autenticación..."
            )
            
            result = app.acquire_token_by_device_flow(flow)
        
        else:
            # ── LOCAL: Abre navegador automático ──────────────────
            st.info("🌐 Abriendo navegador para autenticación...")
            
            result = app.acquire_token_interactive(
                scopes=["https://database.windows.net/.default"]
            )
        
        # Verificar si obtuvo token
        if "access_token" not in result:
            error_msg = result.get('error_description', 'Token no obtenido')
            st.error(f"❌ Autenticación fallida: {error_msg}")
            return None
        
        st.success("✅ Token obtenido correctamente")
        
        # Conectar a Fabric con credenciales
        connection_string = (
            f'Driver={{ODBC Driver 17 for SQL Server}};'
            f'Server={SQL_SERVER};'
            f'Database={SQL_DB};'
            f'UID={SQL_USER};'
            f'PWD={st.secrets["SQL_PASS"]};'
            f'Encrypt=yes;'
            f'TrustServerCertificate=no;'
            f'Connection Timeout=30;'
        )
        
        conn = pyodbc.connect(connection_string)
        st.success("✅ Conectado a Fabric SQL")
        return conn
        
    except Exception as e:
        st.error(f"❌ Error conexión: {str(e)}")
        st.info(
            "💡 **Soluciones:**\n"
            "- Verifica que `msal` esté en requirements.txt\n"
            "- En local: abre el navegador que aparezca\n"
            "- En Cloud: sigue el código de dispositivo"
        )
        return None


# ── CONECTAR ──────────────────────────────────────────────────
with st.spinner("Conectando a Fabric..."):
    conn_fabric = conectar_fabric()
    
    if conn_fabric is None:
        st.stop()
# ── LLAMAR LA FUNCIÓN ────────────────────────────────────────
with st.spinner("Conectando a Fabric..."):
    conn_fabric = conectar_fabric()
    
    if conn_fabric is None:
        st.error("No se pudo conectar a Fabric")
        st.stop()
# ──── CARGAR ET MENSUAL PROMEDIO ────
@st.cache_data(ttl=3600, show_spinner=False)
def cargar_et_mensual_promedio(_conn_fabric):  # ✅ guion bajo
    """Carga ET promedio diario mensual por fundo desde Fabric."""
    try:
        query = """
        WITH EtoMensual AS
        (
            SELECT
                Fundo,
                YEAR(CONVERT(date,[Fecha-Hora])) AS Anio,
                MONTH(CONVERT(date,[Fecha-Hora])) AS Mes,
                SUM([ET-mm]) AS EToMensual,
                COUNT(DISTINCT CONVERT(date,[Fecha-Hora])) AS DiasConDatos
            FROM [dbo].[Clima]
            GROUP BY
                Fundo,
                YEAR(CONVERT(date,[Fecha-Hora])),
                MONTH(CONVERT(date,[Fecha-Hora]))
        )
        SELECT
            Fundo,
            Mes,
            ROUND(AVG(EToMensual * 1.0 / DiasConDatos),2) AS EToPromedioDiaria
        FROM EtoMensual
        GROUP BY
            Fundo,
            Mes
        ORDER BY
            Fundo,
            Mes
        """
        
        df_et_mensual = pd.read_sql(query, _conn_fabric)  # ← aquí
        return df_et_mensual
    
    except Exception as e:
        st.error(f"Error al cargar ET mensual: {e}")
        return pd.DataFrame()

# Cargar ET mensual
df_et_mensual = cargar_et_mensual_promedio(conn_fabric)

if not df_et_mensual.empty:
    st.success(f"✅ ET mensual cargado: {len(df_et_mensual)} registros")
else:
    st.warning("⚠️ Sin datos de ET mensual")
# ──── CARGAR DATOS DESDE FABRIC ────
with st.spinner("Leyendo datos desde Fabric..."):
    try:
        query = "SELECT * FROM [dbo].[vw_Clima]"
        df_fabric = pd.read_sql(query, conn_fabric)
        
        if df_fabric.empty:
            st.error("❌ La vista vw_Clima no tiene datos")
            st.stop()
        
        fecha_min = pd.to_datetime(df_fabric['Fecha-Hora']).min()
        fecha_max = pd.to_datetime(df_fabric['Fecha-Hora']).max()
        
        st.info(
            f"✅ {len(df_fabric):,} registros desde Fabric\n"
            f"📅 {fecha_min.strftime('%d/%m/%Y')} → {fecha_max.strftime('%d/%m/%Y')}"
        )
    except Exception as e:
        st.error(f"Error al leer Fabric: {e}")
        st.stop()

# ──── FUNDOS DISPONIBLES ────
fundos_disponibles = sorted(df_fabric['Fundo'].dropna().unique().tolist())

with fundos_disponibles_placeholder:
    fundos_sel = st.multiselect(
        "Seleccionar fundos",
        options=fundos_disponibles,
        default=fundos_disponibles,
    )

fundos_activos = fundos_sel if fundos_sel else fundos_disponibles

# ──── PROCESAR DATOS ────
with st.spinner("Procesando datos desde Fabric..."):
    try:
        df = df_fabric.copy()
        
        # Fechas
        df['Fecha-Hora'] = pd.to_datetime(df['Fecha-Hora'], dayfirst=False, errors='coerce')
        df = df.dropna(subset=['Fecha-Hora'])
        
        # Numéricos
        for col in ['TempAlta-C', 'TempBaja-C']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Filtro rango
        df = df[
            (df['TempAlta-C'] >= TMAX_MIN) & (df['TempAlta-C'] <= TMAX_MAX) &
            (df['TempBaja-C'] >= TMIN_MIN) & (df['TempBaja-C'] <= TMIN_MAX)
        ].copy()
        
        # ET
        if 'ET-mm' not in df.columns:
            df['ET-mm'] = 0.0
        df['ET-mm'] = pd.to_numeric(df['ET-mm'], errors='coerce').fillna(0.0)
        
        # Fecha normalizada
        df['Fecha'] = df['Fecha-Hora'].dt.normalize()
        
        # Conteo de registros
        reg = df.groupby(['Fundo', 'Fecha'], as_index=False).size()
        reg.columns = ['Fundo', 'Fecha', 'N_registros']
        df = df.merge(reg, on=['Fundo', 'Fecha'], how='left')
        
        # Última fecha por fundo
        fecha_max_por_fundo = (
            df.groupby('Fundo')['Fecha']
              .max()
              .reset_index()
              .rename(columns={'Fecha': 'Fecha_max_fundo'})
        )
        df = df.merge(fecha_max_por_fundo, on='Fundo', how='left')
        
        # Filtro (mínimo 80 o último día)
        df = df[
            (df['N_registros'] >= min_reg_ui) |
            (df['Fecha'] == df['Fecha_max_fundo'])
        ].copy()
        df = df.drop(columns=['Fecha_max_fundo'])
        
        # Filtro fundos
        df = df[df['Fundo'].isin(fundos_activos)]
        
        # Agregación diaria
        dia_full = (
            df.groupby(['Empresa', 'Fundo', 'Fecha'], as_index=False)
              .agg(
                  Tmax=('TempAlta-C', 'max'),
                  Tmin=('TempBaja-C', 'min'),
                  ET=('ET-mm', 'sum'),
              )
        )
        dia_full[['Tmax', 'Tmin', 'ET']] = dia_full[['Tmax', 'Tmin', 'ET']].round(2)
        dia_full = dia_full.sort_values(['Fundo', 'Fecha']).reset_index(drop=True)
        
        # Suavizado 3 días
        dia_full['Tmax_smooth'] = (
            dia_full.groupby('Fundo')['Tmax']
               .transform(lambda x: x.rolling(3, center=True, min_periods=1).mean())
               .round(2)
        )
        dia_full['Tmin_smooth'] = (
            dia_full.groupby('Fundo')['Tmin']
               .transform(lambda x: x.rolling(3, center=True, min_periods=1).mean())
               .round(2)
        )
        
        dia = dia_full.copy()
        
    except Exception as e:
        st.error(f"Error procesando datos: {e}")
        st.stop()

# ──── PROPHET ────
CACHE_DIR  = Path(_APP_DIR) / "assets" / ".prophet_cache"
CACHE_PKL  = CACHE_DIR / "forecasts.pkl"
CACHE_HASH = CACHE_DIR / "data_hash.txt"

CODE_VERSION = "v2.4-fabric-secrets"

def _hash_meteo(df_data):
    params_bytes = json.dumps(PROPHET_PARAMS, sort_keys=True).encode()
    version_bytes = CODE_VERSION.encode()
    try:
        df_hash = hashlib.md5(
            pd.util.hash_pandas_object(dia_full, index=True).values
        ).hexdigest().encode()
    except:
        df_hash = b"fabric_hash"
    return hashlib.md5(df_hash + params_bytes + version_bytes).hexdigest()

def guardar_cache_prophet(forecasts: dict, data_hash: str):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(CACHE_PKL, "wb") as f:
        pickle.dump(forecasts, f)
    CACHE_HASH.write_text(data_hash)

def cargar_cache_prophet(data_hash: str) -> dict | None:
    if not CACHE_PKL.exists() or not CACHE_HASH.exists():
        return None
    if CACHE_HASH.read_text().strip() != data_hash:
        return None
    with open(CACHE_PKL, "rb") as f:
        return pickle.load(f)
    
_data_hash = _hash_meteo(df_fabric)

forecasts_cache = cargar_cache_prophet(_data_hash)

if forecasts_cache is None:
    with st.spinner("Entrenando modelos Prophet..."):
        forecasts_cache = entrenar_todos_optimizado(dia_full, dias_pred_ui)
    guardar_cache_prophet(forecasts_cache, _data_hash)
    st.success("Modelos entrenados y guardados en caché.")
else:
    st.info("Modelos cargados desde caché.")

# ──── MÉTRICAS ────
st.markdown("#### 📊 Resumen de datos")

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Registros</div>
        <div class="value">{len(dia):,}</div>
        <div class="sub">días-fundo</div>
    </div>""", unsafe_allow_html=True)
with col2:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Fundos</div>
        <div class="value">{dia['Fundo'].nunique()}</div>
        <div class="sub">activos</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Desde</div>
        <div class="value" style="font-size:1.15rem">{dia['Fecha'].min().strftime('%d/%m')}</div>
        <div class="sub">{dia['Fecha'].min().year}</div>
    </div>""", unsafe_allow_html=True)
with col4:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Hasta</div>
        <div class="value" style="font-size:1.15rem">{dia['Fecha'].max().strftime('%d/%m')}</div>
        <div class="sub">{dia['Fecha'].max().year}</div>
    </div>""", unsafe_allow_html=True)
with col5:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Tmax prom</div>
        <div class="value">{dia['Tmax'].mean():.1f}°C</div>
        <div class="sub">todos</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)

# ──── TABS ────
mostrar_tmax = variable_sel in ['Ambas', 'Solo Tmax']
mostrar_tmin = variable_sel in ['Ambas', 'Solo Tmin']

dias_pred_mitad = max(7, dias_pred_ui // 2)

if mostrar_tmax and mostrar_tmin:
    tab_tmax, tab_tmin, tab_et, tab_datos = st.tabs(
        ["🔴 Temperatura Máxima", "🔵 Temperatura Mínima", "💧 ET", "📋 Datos"]
    )
elif mostrar_tmax:
    tab_tmax, tab_et, tab_datos = st.tabs(["🔴 Temperatura Máxima", "💧 ET", "📋 Datos"])
    tab_tmin = None
else:
    tab_tmin, tab_et, tab_datos = st.tabs(["🔵 Temperatura Mínima", "💧 ET", "📋 Datos"])
    tab_tmax = None

n_fundos = len(dia['Fundo'].unique())
altura_graficos = 320 * n_fundos
ALTURA_SELECTBOX = 78
# ──── PRE-CALCULAR RIESGOS ET (para evitar re-runs) ────
fundos_activos_list = list(fundos_activos) if fundos_activos else []
riesgos_et_cache = {}
for fundo in fundos_activos_list:
    riesgos_et_cache[fundo] = calcular_riesgo_et_fundo(df_et_mensual, fundo)

if mostrar_tmax and tab_tmax is not None:
    with tab_tmax:
        col1, col2 = st.columns([2, 2])
        with col1:
            st.markdown("#### 📊 Predicción a mostrar")
            dias_pred_mostrar = st.radio(
                "Mostrar predicción de:",
                options=[f"{dias_pred_mitad}d", f"{dias_pred_ui}d"],
                index=1, horizontal=True,
                key='dias_pred_mostrar_tmax'
            )
            dias_pred_mostrar_num = int(dias_pred_mostrar.replace('d', ''))

        tipo_viz_tmax_lower = 'linea'

        with st.spinner("Generando Tmax..."):
            fig_tmax, df_tmax_hist, df_tmax_pred = generar_figura(
                'Tmax', dia, MEDIA_TMAX, Q1_TMAX, Q3_TMAX,
                dias_pred_ui, forecasts_cache, dias_vista_num,
                dias_pred_mostrar_num, tipo_viz_tmax_lower,
                media_mensual2=MEDIA_TMAX2, q1_mensual2=Q1_TMAX2,
                q3_mensual2=Q3_TMAX2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )

        col_graf_tmax, col_mapa_tmax = st.columns([3, 1])
        with col_graf_tmax:
            st.plotly_chart(fig_tmax, use_container_width=True)
        with col_mapa_tmax:
            from streamlit_folium import st_folium

            mes_actual = pd.Timestamp.today().month
            mes_sel_tmax = st.selectbox(
                "Mes",
                options=list(range(1, 13)),
                format_func=lambda mm: MESES_RAW[mm-1],
                index=mes_actual - 1,
                key='mes_sel_anomalia_tmax',
                label_visibility='collapsed'
            )

            temps_distritos_tmax = (
                cargar_temperaturas_distritos(norm_bytes_sidebar, 'TMAX', mes_sel_tmax)
                if norm_bytes_sidebar is not None else {}
            )
            temp_colormap_tmax = crear_colormap_temperatura(temps_distritos_tmax, 'Tmax', mes_sel_tmax)

            mapa_tmax = generar_mapa_distritos(
                geojson_peru,
                distrito_fijo=DISTRITO_BUSCAR,
                distrito_din=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                variable='Tmax',
                dia=dia,
                mes_sel=mes_sel_tmax,
                media_clim_fijo=MEDIA_TMAX,
                media_clim_din=MEDIA_TMAX2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                modulos_kmz=modulos_kmz,
                distritos_senamhi=distritos_senamhi,
                temps_distritos=temps_distritos_tmax,
                temp_colormap=temp_colormap_tmax,
            )
            if mapa_tmax:
                st_folium(mapa_tmax, width=None,
                          height=max(300, altura_graficos - ALTURA_SELECTBOX),
                          returned_objects=[], key='mapa_tmax')

        generar_seccion_validacion('Tmax', dia, forecasts_cache, dias_pred_ui)

else:
    df_tmax_hist = pd.DataFrame()
    df_tmax_pred = pd.DataFrame()

if mostrar_tmin and tab_tmin is not None:
    with tab_tmin:
        col1, col2 = st.columns([2, 2])
        with col1:
            st.markdown("#### 📊 Predicción a mostrar")
            dias_pred_mostrar = st.radio(
                "Mostrar predicción de:",
                options=[f"{dias_pred_mitad}d", f"{dias_pred_ui}d"],
                index=1, horizontal=True,
                key='dias_pred_mostrar_tmin'
            )
            dias_pred_mostrar_num = int(dias_pred_mostrar.replace('d', ''))

        tipo_viz_tmin_lower = 'linea'

        with st.spinner("Generando Tmin..."):
            fig_tmin, df_tmin_hist, df_tmin_pred = generar_figura(
                'Tmin', dia, MEDIA_TMIN, Q1_TMIN, Q3_TMIN,
                dias_pred_ui, forecasts_cache, dias_vista_num,
                dias_pred_mostrar_num, tipo_viz_tmin_lower,
                media_mensual2=MEDIA_TMIN2, q1_mensual2=Q1_TMIN2,
                q3_mensual2=Q3_TMIN2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )

        col_graf_tmin, col_mapa_tmin = st.columns([3, 1])
        with col_graf_tmin:
            st.plotly_chart(fig_tmin, use_container_width=True)
        with col_mapa_tmin:
            from streamlit_folium import st_folium

            mes_actual = pd.Timestamp.today().month
            mes_sel_tmin = st.selectbox(
                "Mes",
                options=list(range(1, 13)),
                format_func=lambda mm: MESES_RAW[mm-1],
                index=mes_actual - 1,
                key='mes_sel_anomalia_tmin',
                label_visibility='collapsed'
            )

            temps_distritos_tmin = (
                cargar_temperaturas_distritos(norm_bytes_sidebar, 'TMIN', mes_sel_tmin)
                if norm_bytes_sidebar is not None else {}
            )
            temp_colormap_tmin = crear_colormap_temperatura(temps_distritos_tmin, 'Tmin', mes_sel_tmin)

            mapa_tmin = generar_mapa_distritos(
                geojson_peru,
                distrito_fijo=DISTRITO_BUSCAR,
                distrito_din=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                variable='Tmin',
                dia=dia,
                mes_sel=mes_sel_tmin,
                media_clim_fijo=MEDIA_TMIN,
                media_clim_din=MEDIA_TMIN2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                modulos_kmz=modulos_kmz,
                distritos_senamhi=distritos_senamhi,
                temps_distritos=temps_distritos_tmin,
                temp_colormap=temp_colormap_tmin,
            )
            if mapa_tmin:
                st_folium(mapa_tmin, width=None,
                          height=max(300, altura_graficos - ALTURA_SELECTBOX),
                          returned_objects=[], key='mapa_tmin')

        generar_seccion_validacion('Tmin', dia, forecasts_cache, dias_pred_ui)

else:
    df_tmin_hist = pd.DataFrame()
    df_tmin_pred = pd.DataFrame()

if tab_et is not None:
    with tab_et:
        generar_tab_et(dia, forecasts_cache, dias_pred_ui, modulos_kmz, riesgos_et_cache)

with tab_datos:
    st.markdown("#### 📋 Tabla de datos diarios")
    
    col_f, col_var = st.columns([2, 2])
    with col_f:
        fundos_tabla = st.multiselect(
            "Filtrar fundo", dia['Fundo'].unique().tolist(),
            default=dia['Fundo'].unique().tolist(), key='tabla_fundo'
        )
    with col_var:
        var_tabla = st.selectbox("Variable", ['Tmax', 'Tmin', 'Ambas'], key='tabla_var')
    
    df_show = dia[dia['Fundo'].isin(fundos_tabla)].copy()
    df_show['Fecha'] = df_show['Fecha'].dt.strftime('%Y-%m-%d')
    
    if var_tabla == 'Tmax':
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmax', 'Tmax_smooth']
    elif var_tabla == 'Tmin':
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmin', 'Tmin_smooth']
    else:
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmax', 'Tmax_smooth', 'Tmin', 'Tmin_smooth']
    
    st.dataframe(df_show[cols_show], use_container_width=True, hide_index=True, height=340)
    
    st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    st.markdown("#### ⬇️ Descargas")
    
    col_d1, col_d2, col_d3 = st.columns(3)
    
    df_hist_all = pd.concat([df_tmax_hist, df_tmin_hist], ignore_index=True)
    df_pred_all = pd.concat([df_tmax_pred, df_tmin_pred], ignore_index=True)
    
    with col_d1:
        if not df_hist_all.empty:
            excel_bytes = exportar_excel(df_hist_all, df_pred_all)
            st.download_button(
                label="📥 Excel completo",
                data=excel_bytes,
                file_name="Temperatura_Aquanqa_BI.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    with col_d2:
        if not df_hist_all.empty:
            st.download_button(
                label="📄 CSV histórico",
                data=df_hist_all.to_csv(index=False).encode('utf-8'),
                file_name="historico_temperatura.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col_d3:
        if not df_pred_all.empty:
            st.download_button(
                label="📄 CSV predicciones",
                data=df_pred_all.to_csv(index=False).encode('utf-8'),
                file_name="predicciones_prophet.csv",
                mime="text/csv",
                use_container_width=True
            )

st.markdown("""
<div class="footer-note">
    Elaborado por la Gerencia de Planificación, Control Operacional y Gestión
</div>
""", unsafe_allow_html=True)