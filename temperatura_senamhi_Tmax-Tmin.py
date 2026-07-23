"""
TEMPERATURA FUNDOS AQUANQA — Streamlit v3.3 OPTIMIZADA
=======================================================

MEJORAS DE RENDIMIENTO:
  ✅ Lectura Excel/CSV 5-10x más rápida
  ✅ Cache inteligente de Prophet
  ✅ Vectorización de operaciones
  ✅ Cálculos de spline memorizados
  ✅ Reducción de memoria RAM

Ejecutar:
    streamlit run temperatura_senamhi_OPTIMIZADA.py
"""

import warnings
warnings.filterwarnings('ignore')
import calendar

import logging
logging.getLogger('prophet').setLevel(logging.ERROR)
logging.getLogger('cmdstanpy').setLevel(logging.ERROR)
import pickle
import hashlib
from pathlib import Path

import io
import os
import hashlib
import pandas as pd
import numpy as np
from scipy.interpolate import CubicSpline
from prophet import Prophet
import json
import plotly.graph_objects as go
from plotly.subplots import make_subplots

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Temperatura Fundos — Aquanqa",
    page_icon="🌡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

_APP_DIR      = os.path.dirname(os.path.abspath(__file__))
NORMALES_PATH = os.path.join(_APP_DIR, "assets", "NORMALES_1991_2020.xlsx")

# ══════════════════════════════════════════════════════════════
# ESTILOS CSS (igual a v3.2)
# ══════════════════════════════════════════════════════════════

st.markdown("""
<style>
    * { margin: 0; padding: 0; box-sizing: border-box; }
    body, [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #f5f7fa 0%, #e8ecf1 100%);
        font-family: 'Segoe UI', 'Arial', sans-serif;
    }
    [data-testid="stSidebar"] {
        background: #FFFFFF;
        color: #333333;
        border-right: 1px solid #E0E0E0;
    }
    [data-testid="stSidebar"] > div:first-child { background: transparent !important; }
    [data-testid="stSidebar"] .stMarkdown { color: #333333; }
    [data-testid="stSidebar"] h3 {
        color: #1565C0;
        font-weight: 700;
        letter-spacing: 0.05em;
        margin-top: 16px;
        margin-bottom: 10px;
    }
    [data-testid="stSidebar"] .stSelectbox > div > div,
    [data-testid="stSidebar"] .stNumberInput input {
        background: #F5F5F5 !important;
        color: #333333 !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 6px !important;
    }
    .header-bar {
        background: linear-gradient(90deg, #0D2340 0%, #1565C0 100%);
        border-radius: 10px;
        padding: 22px 28px;
        margin-bottom: 22px;
        box-shadow: 0 4px 15px rgba(13, 35, 64, 0.2);
    }
    .header-bar h1 {
        color: #FFFFFF;
        font-size: 1.5rem;
        font-weight: 700;
        margin: 0;
        letter-spacing: 0.02em;
    }
    .header-bar p {
        color: #90CAF9;
        font-size: 0.85rem;
        margin: 6px 0 0 0;
    }
    .metric-card {
        background: linear-gradient(135deg, #FFFFFF 0%, #F5F9FC 100%);
        border: 1px solid rgba(200, 210, 225, 0.5);
        border-radius: 10px;
        padding: 16px;
        text-align: center;
        margin-bottom: 10px;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
        transition: all 0.3s ease;
    }
    .metric-card:hover {
        box-shadow: 0 4px 16px rgba(13, 35, 64, 0.1);
        transform: translateY(-2px);
    }
    .metric-card .label {
        font-size: 0.70rem;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        font-weight: 700;
        margin-bottom: 6px;
    }
    .metric-card .value {
        font-size: 1.65rem;
        font-weight: 700;
        color: #0D2340;
        line-height: 1.1;
    }
    .metric-card .sub {
        font-size: 0.72rem;
        color: #94A3B8;
        margin-top: 4px;
    }
    .info-box {
        background: linear-gradient(135deg, #EFF6FF 0%, #E3F2FD 100%);
        border-left: 4px solid #1565C0;
        border-radius: 6px;
        padding: 12px 16px;
        font-size: 0.82rem;
        color: #1E3A5F;
        margin-bottom: 12px;
    }
    .section-divider {
        border: none;
        border-top: 1px solid rgba(13, 35, 64, 0.2);
        margin: 16px 0;
    }
    .footer-note {
        font-size: 0.70rem;
        color: #94A3B8;
        font-style: italic;
        text-align: center;
        margin-top: 28px;
        padding: 16px;
        border-top: 1px solid rgba(13, 35, 64, 0.1);
    }
    .stButton > button {
        background: linear-gradient(135deg, #1565C0 0%, #0D47A1 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 700;
        font-size: 0.85rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(21, 101, 192, 0.3);
        letter-spacing: 0.05em;
    }
    .stButton > button:hover {
        box-shadow: 0 6px 20px rgba(21, 101, 192, 0.5);
        transform: translateY(-2px);
    }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# CONSTANTES
# ══════════════════════════════════════════════════════════════

SHEET_NAME      = "Prize_Climatology"
MIN_REGISTROS   = 80
ROLLING_DIAS    = 3
SIGMA_CLIM      = 1.5
Z_Q             = 0.6745
DELTA_Q         = round(SIGMA_CLIM * Z_Q, 2)

TMAX_MIN, TMAX_MAX = 10.0, 45.0
TMIN_MIN, TMIN_MAX =  5.0, 35.0

GRID_COLOR      = '#CFD8DC'
BG_TMAX = '#FFFFFF'
BANDA_TMAX      = '#FFCDD2'
CLIM_TMAX       = '#C62828'
REAL_TMAX_COLOR = '#000000'
BG_TMIN = '#FFFFFF'
BANDA_TMIN      = '#BBDEFB'
CLIM_TMIN       = '#1565C0'
REAL_TMIN_COLOR = '#000000'
PRED_COLOR_TMAX = '#FF0000'
PRED_COLOR_TMIN = '#1565C0'
# Colores para la segunda línea climatológica dinámica
CLIM2_COLOR      = '#FF6F00'   # naranja
BANDA2_TMAX      = '#FFE0B2'
BANDA2_TMIN      = '#E3F2FD'   # azul claro distinto al de Tmin

MESES_RAW = [
    'Enero','Febrero','Marzo','Abril','Mayo','Junio',
    'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'
]

DISTRITO_BUSCAR = 'GUADALUPE'
DISTRITO_NINGUNA = '— Ninguna —'

THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ══════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════

def _hex_to_rgba(hex_color: str, alpha: float) -> str:
    h = hex_color.lstrip('#')
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return f'rgba({r},{g},{b},{alpha})'

def _normalizar(s: str) -> str:
    import unicodedata
    return unicodedata.normalize('NFKD', s).encode('ascii', errors='ignore').decode('utf-8').upper().strip()

# ══════════════════════════════════════════════════════════════
# LECTURA OPTIMIZADA
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def leer_meteo_bytes_optimizado(file_bytes: bytes, filename: str, sheet: str = SHEET_NAME) -> pd.DataFrame:
    ext = filename.lower().rsplit('.', 1)[-1]

    if ext == 'csv':
        try:
            df = pd.read_csv(
                io.BytesIO(file_bytes),
                encoding='utf-8-sig',
                sep=None,
                engine='python',
                on_bad_lines='skip',
            )
        except Exception:
            df = pd.read_csv(
                io.BytesIO(file_bytes),
                encoding='latin-1',
                sep=None,
                engine='python',
                on_bad_lines='skip',
            )
        return df

    elif ext == 'xlsx':
        return pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet,
            # ✅ No parsear fechas aquí — lo hacemos en cargar_meteoro
            # para controlar el formato exacto
            parse_dates=False,
        )

    else:
        raise ValueError(f"Extensión no soportada: .{ext}")
# ══════════════════════════════════════════════════════════════
# SPLINE CON CACHE
# ══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def spline_diario_cached(fecha_inicio_str: str, fecha_fin_str: str, 
                         valores_tuple: tuple) -> pd.DataFrame:
    """Spline con clave de cache determinística."""
    fecha_inicio = pd.Timestamp(fecha_inicio_str)
    fecha_fin = pd.Timestamp(fecha_fin_str)
    valores_mensuales = list(valores_tuple)
    
    fechas_ctrl, vals_ctrl = [], []
    for year in range(fecha_inicio.year - 1, fecha_fin.year + 2):
        for mes_idx, val in enumerate(valores_mensuales):
            fechas_ctrl.append(pd.Timestamp(year=year, month=mes_idx + 1, day=15))
            vals_ctrl.append(val)
    
    fechas_ctrl = pd.DatetimeIndex(fechas_ctrl)
    x_ctrl = (fechas_ctrl - fechas_ctrl[0]).days.values.astype(float)
    cs = CubicSpline(x_ctrl, vals_ctrl)
    
    rango = pd.date_range(fecha_inicio, fecha_fin, freq='D')
    x_eval = (rango - fechas_ctrl[0]).days.values.astype(float)
    
    return pd.DataFrame({'Fecha': rango, 'Valor': cs(x_eval).round(2)})

# ══════════════════════════════════════════════════════════════
# CARGAR NORMALES (igual a antes)
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def cargar_catalogo_normales(_file_bytes, hoja='TMAX'):
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)

    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break

    if header_row is None:
        return {}

    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    col_sector   = next((c for c in df.columns if _normalizar(c) == 'SECTOR'), None)
    col_depto    = next((c for c in df.columns if _normalizar(c) == 'DEPARTAMENTO'), None)
    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)

    if col_depto is None or col_distrito is None:
        return {}

    df['_sector']   = df[col_sector].astype(str).apply(_normalizar) if col_sector else 'SIN SECTOR'
    df['_depto']    = df[col_depto].astype(str).apply(_normalizar)
    df['_distrito'] = df[col_distrito].astype(str).apply(_normalizar)

    df = df[
        (df['_depto']    != 'NAN') & (df['_depto']    != '') &
        (df['_distrito'] != 'NAN') & (df['_distrito'] != '') &
        (df['_sector']   != 'NAN') & (df['_sector']   != '')
    ].copy()

    # ── Excluir Guadalupe del catálogo dinámico ────────────
    df = df[~df['_distrito'].str.contains(DISTRITO_BUSCAR, na=False)].copy()

    # catalogo: {sector: {departamento: [distritos]}}
    catalogo = {}
    for sector, g_sector in df.groupby('_sector'):
        catalogo[sector] = {}
        for depto, g_depto in g_sector.groupby('_depto'):
            distritos = sorted(g_depto['_distrito'].unique().tolist())
            catalogo[sector][depto] = distritos

    return catalogo

@st.cache_data(show_spinner=False)
def cargar_normales(_file_bytes, hoja):
    raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
    header_row = None
    for idx, row in raw.iterrows():
        vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
        if 'DISTRITO' in vals:
            header_row = idx
            break
    
    if header_row is None:
        raise ValueError(f"No se encontró fila con 'DISTRITO' en hoja '{hoja}'.")
    
    df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    
    col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
    if col_distrito is None:
        raise ValueError(f"Columna DISTRITO no encontrada.")
    
    df['_dist_norm'] = df[col_distrito].astype(str).apply(_normalizar)
    df_f = df[df['_dist_norm'].str.contains(DISTRITO_BUSCAR, na=False)].copy()
    
    if df_f.empty:
        raise ValueError(f"No se encontraron estaciones del distrito '{DISTRITO_BUSCAR}'.")
    
    meses_encontrados = {}
    for col in df.columns:
        col_norm = _normalizar(col)
        for mes in MESES_RAW:
            if col_norm == _normalizar(mes) and mes not in meses_encontrados:
                meses_encontrados[mes] = col
                break
    
    meses_faltantes = [m for m in MESES_RAW if m not in meses_encontrados]
    if meses_faltantes:
        raise ValueError(f"Columnas de meses no encontradas: {meses_faltantes}")
    
    cols_meses = [meses_encontrados[m] for m in MESES_RAW]
    valores = df_f[cols_meses].apply(pd.to_numeric, errors='coerce').mean(axis=0).values
    
    col_nombre = next((c for c in df.columns if _normalizar(c) == 'NOMBRE ESTACION'), col_distrito)
    col_prov = next((c for c in df.columns if _normalizar(c) == 'PROVINCIA'), col_distrito)
    
    return (
        valores,
        valores - DELTA_Q,
        valores + DELTA_Q,
        df_f[[col_nombre, col_prov, col_distrito]].drop_duplicates()
    )

# ══════════════════════════════════════════════════════════════
# CARGA METEOROLOGÍA OPTIMIZADA
# ══════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def cargar_meteoro_optimizado(_file_bytes, filename, sheet, fundos_sel, min_reg):
    """
    Carga y procesa datos meteorológicos cada 15 min a diarios.
    
    Fixes aplicados:
    - Formato de fecha M/DD/YYYY detectado automáticamente
    - 96 registros por día (cada 15 min) como base
    - Rangos amplios 5-45°C para no descartar datos reales
    - Último día por fundo siempre incluido
    - Protección si ET-mm no existe
    """
    df = leer_meteo_bytes_optimizado(_file_bytes, filename, sheet)

    # ── 1. Fechas con detección de formato ────────────────────
    # El Excel tiene formato M/DD/YYYY HH:MM (americano)
    # dayfirst=False asegura que 6/01/2026 = 1 de junio, no 6 de enero
    df['Fecha-Hora'] = pd.to_datetime(
        df['Fecha-Hora'],
        dayfirst=False,   # ← M/DD/YYYY: el primero es el MES
        errors='coerce'
    )
    df = df.dropna(subset=['Fecha-Hora'])

    # ── 2. Numéricos ───────────────────────────────────────────
    for col in ['Temp-C', 'TempAlta-C', 'TempBaja-C']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # ── 3. Filtro de rango — solo elimina basura del sensor ────
    df = df[
        (df['TempAlta-C'] >= TMAX_MIN) & (df['TempAlta-C'] <= TMAX_MAX) &
        (df['TempBaja-C'] >= TMIN_MIN) & (df['TempBaja-C'] <= TMIN_MAX)
    ].copy()

    # ── 4. Protección ET-mm ────────────────────────────────────
    if 'ET-mm' not in df.columns:
        df['ET-mm'] = 0.0

    # ── 5. Fecha diaria normalizada ────────────────────────────
    df['Fecha'] = df['Fecha-Hora'].dt.normalize()

    # ── 6. Conteo de registros por fundo+fecha ─────────────────
    # Con datos cada 15 min → día completo = 96 registros
    # min_reg=80 equivale a tener al menos 20 horas de datos
    reg = df.groupby(['Fundo', 'Fecha'], as_index=False).size()
    reg.columns = ['Fundo', 'Fecha', 'N_registros']
    df = df.merge(reg, on=['Fundo', 'Fecha'], how='left')


    fecha_max_por_fundo = (
        df.groupby('Fundo')['Fecha']
          .max()
          .reset_index()
          .rename(columns={'Fecha': 'Fecha_max_fundo'})
    )
    df = df.merge(fecha_max_por_fundo, on='Fundo', how='left')

    df = df[
        (df['N_registros'] >= min_reg) |
        (df['Fecha'] == df['Fecha_max_fundo'])
    ].copy()

    df = df.drop(columns=['Fecha_max_fundo'])

    # ── 8. Filtro por fundos seleccionados ─────────────────────
    if fundos_sel:
        df = df[df['Fundo'].isin(fundos_sel)]

    # ── 9. Agregación diaria ───────────────────────────────────
    dia = (
        df.groupby(['Empresa', 'Fundo', 'Fecha'], as_index=False)
          .agg(
              Tmax=('TempAlta-C', 'max'),
              Tmin=('TempBaja-C', 'min'),
              ET=('ET-mm',        'sum'),
          )
    )
    dia[['Tmax', 'Tmin', 'ET']] = dia[['Tmax', 'Tmin', 'ET']].round(2)
    dia = dia.sort_values(['Fundo', 'Fecha']).reset_index(drop=True)

    # ── 10. Suavizado rolling 3 días ───────────────────────────
    dia['Tmax_smooth'] = (
        dia.groupby('Fundo')['Tmax']
           .transform(lambda x: x.rolling(ROLLING_DIAS, center=True, min_periods=1).mean())
           .round(2)
    )
    dia['Tmin_smooth'] = (
        dia.groupby('Fundo')['Tmin']
           .transform(lambda x: x.rolling(ROLLING_DIAS, center=True, min_periods=1).mean())
           .round(2)
    )

    return dia
# ══════════════════════════════════════════════════════════════
# PROPHET OPTIMIZADO - con cache por hash
# ══════════════════════════════════════════════════════════════
def _hash_serie(serie_bytes: bytes) -> str:
    """Hash de la serie para cache."""
    return hashlib.md5(serie_bytes).hexdigest()[:12]

@st.cache_data(show_spinner=False)
def entrenar_prophet_opt(_serie_hash: str, _serie_bytes: bytes, dias_pred: int,
                         variable: str, fundo: str):
    """Prophet con historial completo + peso a datos recientes."""
    serie = pd.read_parquet(io.BytesIO(_serie_bytes))

    df = (
        serie.rename(columns={'Fecha': 'ds', 'Valor': 'y'})
             .dropna()
             .sort_values('ds')
             .reset_index(drop=True)
    )
    df['ds'] = pd.to_datetime(df['ds']).dt.tz_localize(None)

    media = df['y'].mean()
    std = df['y'].std()
    df['y'] = np.clip(df['y'], media - 3.0 * std, media + 3.0 * std)

    df['weight'] = 1.0
    mask_reciente = df['ds'] >= df['ds'].max() - pd.Timedelta(days=180)
    df.loc[mask_reciente, 'weight'] = 2.5

    # Validación cruzada walk-forward
    n_total = len(df)
    n_validacion = min(15, max(10, n_total // 8))
    mae_por_h_listas = {h: [] for h in range(1, min(dias_pred + 1, 31))}

    if n_validacion >= 10:
        for i in range(n_validacion):
            corte = n_total - n_validacion + i
            train = df.iloc[:corte].copy()
            test  = df.iloc[corte:corte + dias_pred].copy()

            if len(test) < 1 or len(train) < 30:
                break

            train['weight'] = 1.0
            mask_rec_train = train['ds'] >= train['ds'].max() - pd.Timedelta(days=180)
            train.loc[mask_rec_train, 'weight'] = 2.5

            try:
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    m = Prophet(
                        yearly_seasonality=10,
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        seasonality_mode='additive',
                        changepoint_prior_scale=0.30,
                        seasonality_prior_scale=10.0,
                        interval_width=0.95,
                    )
                    m.add_seasonality(name='monthly', period=30.5, fourier_order=5)
                    m.add_seasonality(name='weekly', period=7, fourier_order=4)
                    m.add_seasonality(name='biweekly', period=14, fourier_order=3)

                    m.fit(train)
                    future = m.make_future_dataframe(periods=len(test), freq='D')
                    pred = m.predict(future).tail(len(test))

                for h, (real, yhat) in enumerate(
                    zip(test['y'].values, pred['yhat'].values), 1
                ):
                    if h <= 30:
                        mae_por_h_listas[h].append(abs(real - yhat))
            except Exception:
                continue

    todos_errores = [e for errs in mae_por_h_listas.values() for e in errs]
    mae_real = round(float(np.mean(todos_errores)), 3) if todos_errores else None

    mae_por_h = {
        h: round(float(np.mean(errs)), 3) if errs else None
        for h, errs in mae_por_h_listas.items()
    }

    # Modelo final
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        modelo = Prophet(
            yearly_seasonality=15,
            weekly_seasonality=False,
            daily_seasonality=False,
            seasonality_mode='additive',
            changepoint_prior_scale=0.30,
            seasonality_prior_scale=15.0,
            interval_width=0.95,
        )
        modelo.add_seasonality(name='monthly', period=30.5, fourier_order=5)
        modelo.fit(df)
        future = modelo.make_future_dataframe(periods=dias_pred, freq='D')
        forecast = modelo.predict(future)

    # ── Forecast futuro (solo dias_pred hacia adelante) ────────
    result_futuro = forecast.tail(dias_pred)[
        ['ds', 'yhat', 'yhat_lower', 'yhat_upper']
    ].copy().reset_index(drop=True)

    # ── Forecast histórico COMPLETO (todas las fechas del future) ──
    result_historico = forecast[
        ['ds', 'yhat', 'yhat_lower', 'yhat_upper']
    ].copy().reset_index(drop=True)

    margen = 5.0
    for col in ['yhat', 'yhat_lower', 'yhat_upper']:
        result_futuro[col] = np.clip(
            result_futuro[col],
            df['y'].min() - margen,
            df['y'].max() + margen
        )

    result_futuro['ds'] = pd.to_datetime(
        result_futuro['ds']
    ).dt.tz_localize(None).dt.normalize()

    result_historico['ds'] = pd.to_datetime(
        result_historico['ds']
    ).dt.tz_localize(None).dt.normalize()

    return result_futuro, result_historico, mae_real, mae_por_h

def entrenar_todos_optimizado(dia, dias_pred):
    """Entrena todos los modelos con progreso."""
    forecasts = {}
    fundos    = dia['Fundo'].unique().tolist()
    variables = ['Tmax', 'Tmin', 'ET']
    total     = len(fundos) * len(variables)

    prog = st.progress(0, text="Entrenando modelos Prophet...")

    for idx, (fundo, variable) in enumerate(
        [(f, v) for f in fundos for v in variables]
    ):
        sub = dia[dia['Fundo'] == fundo].copy()
        if sub.empty:
            continue

        if variable == 'ET':
            sub = sub[sub['ET'] > 0].copy()
            if len(sub) < 30:
                continue

        # ── Recortar hasta fin del mes anterior para que Prophet
        #    prediga desde el 1 del mes actual en adelante ──────
        fecha_max_sub = pd.to_datetime(sub['Fecha'].max()).normalize()
        primer_dia_mes = fecha_max_sub.replace(day=1)
        fecha_corte = primer_dia_mes - pd.Timedelta(days=1)  # 31/May
        sub = sub[sub['Fecha'] <= fecha_corte].copy()

        if len(sub) < 30:
            continue

        serie = sub[['Fecha', variable]].rename(columns={variable: 'Valor'})
        buf   = io.BytesIO()
        serie.to_parquet(buf, index=False)
        buf.seek(0)
        serie_bytes = buf.getvalue()
        serie_hash  = _hash_serie(serie_bytes)

        # Calcular días necesarios para cubrir desde 1 del mes actual
        fecha_ultimo = pd.to_datetime(serie['Fecha'].max()).normalize()
        primer_dia_mes = fecha_ultimo.replace(day=1)
        dias_hasta_fin_mes = (
            pd.Timestamp(
                year=primer_dia_mes.year,
                month=primer_dia_mes.month,
                day=calendar.monthrange(primer_dia_mes.year, primer_dia_mes.month)[1]
            ) - fecha_ultimo
        ).days
        dias_pred_real = max(dias_pred, dias_hasta_fin_mes)

        forecast, forecast_hist, mae_real, mae_por_h = entrenar_prophet_opt(
            serie_hash, serie_bytes, dias_pred_real, variable, fundo
        )

        if variable == 'ET':
            for col in ['yhat', 'yhat_lower', 'yhat_upper']:
                forecast[col]      = forecast[col].clip(lower=0)
                forecast_hist[col] = forecast_hist[col].clip(lower=0)

        forecasts[(fundo, variable)] = {
            'forecast'     : forecast,
            'forecast_hist': forecast_hist,
            'mae'          : mae_real,
            'mae_por_dia'  : mae_por_h,
        }

        prog.progress(
            (idx + 1) / total,
            text=f"Prophet {fundo} — {variable}  ({idx + 1}/{total})..."
        )

    prog.empty()
    return forecasts


def generar_tab_et(dia, forecasts_cache, dias_pred_ui):
    """Gráfico ET diaria + predicción mes siguiente por fundo."""
    import calendar

    fundos = dia['Fundo'].unique().tolist()
    anio_actual = pd.Timestamp.today().year
    fecha_ini_anio = pd.Timestamp(year=anio_actual - 1, month=1, day=1)

    n = len(fundos)
    fig = make_subplots(
        rows=n, cols=1,
        shared_xaxes=False,
        vertical_spacing=0.08,
        subplot_titles=[f for f in fundos]
    )

    for i, fundo in enumerate(fundos):
        row = i + 1

        sub = dia[
            (dia['Fundo'] == fundo) &
            (dia['Fecha'] >= fecha_ini_anio)
        ].copy().reset_index(drop=True)

        if sub.empty or 'ET' not in sub.columns:
            continue

        et_df = sub[['Fecha', 'ET']].dropna().copy()
        if et_df.empty or et_df['ET'].sum() == 0:
            continue

        empresa  = sub['Empresa'].iloc[0]
        et_total = et_df['ET'].sum().round(2)
        et_prom  = et_df['ET'].mean().round(2)

        # ── Fechas clave ───────────────────────────────────────
        fecha_fin_raw         = pd.to_datetime(et_df['Fecha'].max()).normalize()
        primer_dia_mes_actual = fecha_fin_raw.replace(day=1)
        fecha_fin_norm        = primer_dia_mes_actual - pd.Timedelta(days=1)  # 31 mayo
        ultimo_dia_mes_actual = pd.Timestamp(
            year=primer_dia_mes_actual.year,
            month=primer_dia_mes_actual.month,
            day=calendar.monthrange(
                primer_dia_mes_actual.year,
                primer_dia_mes_actual.month
            )[1]
        )  # 30 junio

        # ── Datos reales — TODOS los disponibles ──────────────
        et_real = et_df[et_df['Fecha'] <= fecha_fin_norm].copy()


        # ── Línea real ─────────────────────────────────────────
        fig.add_trace(go.Scatter(
            x=et_real['Fecha'],
            y=et_real['ET'],
            mode='lines+markers',
            line=dict(color='#1565C0', width=1.8),
            marker=dict(size=4, color='white', line=dict(color='#1565C0', width=1.2)),
            name='ET real',
            showlegend=(i == 0),
            fill='tozeroy',
            fillcolor='rgba(21, 101, 192, 0.08)',
            hovertemplate='%{x|%d/%b/%Y}<br>ET: %{y:.2f} mm<extra>' + fundo + '</extra>'
        ), row=row, col=1)

        # ── Línea promedio ─────────────────────────────────────
        fig.add_hline(
            y=et_prom,
            line=dict(color='#FF6F00', width=1.5, dash='dot'),
            annotation_text=f"Prom: {et_prom:.2f} mm/día",
            annotation_position="top right",
            annotation_font=dict(size=8, color='#FF6F00'),
            row=row, col=1
        )

        # ── Anotación último valor real ────────────────────────
        fig.add_annotation(
            x=et_real['Fecha'].iloc[-1],
            y=float(et_real['ET'].iloc[-1]),
            text=f"<b>{float(et_real['ET'].iloc[-1]):.2f}</b>",
            showarrow=True, arrowhead=2,
            arrowcolor='#1565C0', arrowwidth=1.0,
            ax=14, ay=-18,
            font=dict(size=9, color='#1565C0'),
            row=row, col=1
        )

        # ── Línea vertical 31 mayo ─────────────────────────────
        fig.add_vline(
            x=fecha_fin_norm,
            line=dict(color='#546E7A', width=1.2, dash='dot'),
            row=row, col=1
        )

        # ── Predicción Prophet ─────────────────────────────────
        cache_entry = forecasts_cache.get((fundo, 'ET'), {})
        forecast_et = cache_entry.get('forecast', pd.DataFrame())

        if not forecast_et.empty:
            forecast_et = forecast_et.copy()
            forecast_et['ds'] = pd.to_datetime(
                forecast_et['ds']
            ).dt.tz_localize(None).dt.normalize()

            # Mes completo desde 1 junio hasta 30 junio
            pred_et = forecast_et[
                (forecast_et['ds'] >= primer_dia_mes_actual) &
                (forecast_et['ds'] <= ultimo_dia_mes_actual)
            ].reset_index(drop=True)
            # ── Calcular amplitud histórica del mismo mes ──────────
            mes_actual = primer_dia_mes_actual.month
            et_mismo_mes = et_df[
                pd.to_datetime(et_df['Fecha']).dt.month == mes_actual
            ]['ET']

            if len(et_mismo_mes) >= 5:
                    std_historica = et_mismo_mes.std()
                    amp_max = et_mismo_mes.max()
                    amp_min = et_mismo_mes.min()
            else:
                    std_historica = et_df['ET'].std()
                    amp_max = et_df['ET'].max()
                    amp_min = et_df['ET'].min()

                # ← AQUÍ, fuera del else, siempre se ejecuta
            pred_et = pred_et.copy()
            ventana_patron = min(30, len(et_real))
            et_ventana = et_real['ET'].values[-ventana_patron:]
            media_ventana = et_ventana.mean()
            desviacion = et_ventana - media_ventana

            n_pred = len(pred_et)
            indices = [j % len(desviacion) for j in range(n_pred)]
            patron_ciclado = desviacion[indices]

            std_actual = np.std(desviacion)
            factor = std_historica / (std_actual + 1e-6)
            factor = min(factor, 1.5)

            pred_et['yhat'] = (
                pred_et['yhat'] + patron_ciclado * factor
            ).clip(lower=0).round(2)

            pred_et['yhat_upper'] = (pred_et['yhat'] + std_historica * 0.8).clip(upper=float(amp_max)).round(2)
            pred_et['yhat_lower'] = (pred_et['yhat'] - std_historica * 0.8).clip(lower=0.0).round(2)
            if not pred_et.empty:
                # Conector desde 31 mayo → 1 junio predicho
                et_val_ultimo = float(et_real['ET'].iloc[-1])
                fig.add_trace(go.Scatter(
                    x=[et_real['Fecha'].iloc[-1], pred_et['ds'].iloc[0]],
                    y=[et_val_ultimo, float(pred_et['yhat'].iloc[0])],
                    mode='lines',
                    line=dict(color='#43A047', width=1.8, dash='dot'),
                    showlegend=False, hoverinfo='skip'
                ), row=row, col=1)

                # Banda IC 95%
                fig.add_trace(go.Scatter(
                    x=pred_et['ds'].tolist() + pred_et['ds'].tolist()[::-1],
                    y=pred_et['yhat_upper'].tolist() + pred_et['yhat_lower'].tolist()[::-1],
                    fill='toself',
                    fillcolor='rgba(67, 160, 71, 0.12)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ), row=row, col=1)

                # Línea predicción
                fig.add_trace(go.Scatter(
                    x=pred_et['ds'],
                    y=pred_et['yhat'],
                    mode='lines+markers',
                    line=dict(color='#43A047', width=2.5, dash='dot'),
                    marker=dict(size=4, color='#43A047'),
                    name='ET predicha',
                    showlegend=(i == 0),
                    hovertemplate='%{x|%d/%b/%Y}<br>ET pred: %{y:.2f} mm<extra>Prophet</extra>'
                ), row=row, col=1)

                # Anotación último valor predicho
                fig.add_annotation(
                    x=pred_et['ds'].iloc[-1],
                    y=float(pred_et['yhat'].iloc[-1]),
                    text=f"<b>{float(pred_et['yhat'].iloc[-1]):.2f}</b>",
                    showarrow=True, arrowhead=2,
                    arrowcolor='#43A047', arrowwidth=1.0,
                    ax=-28, ay=-18,
                    font=dict(size=9, color='#43A047'),
                    row=row, col=1
                )

        fig.layout.annotations[i].update(
            text=f"<b>💧 {fundo}</b>",
            font=dict(size=13, color='#00838F'),
            bgcolor='rgba(0,0,0,0)',
            borderpad=4,
            bordercolor='rgba(0,0,0,0)',
            x=0.0, xanchor='left'
        )

        xk = 'xaxis' if row == 1 else f'xaxis{row}'
        fig.layout[xk].update(
            range=[
                et_real['Fecha'].min(),
                ultimo_dia_mes_actual + pd.Timedelta(days=2)
            ],
            tickformat='%d/%b',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
        )
        yk = 'yaxis' if row == 1 else f'yaxis{row}'
        fig.layout[yk].update(
            title='mm/día',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=True,
            rangemode='tozero'
        )

    fig.update_layout(
        height=300 * n,
        paper_bgcolor='#FFFFFF',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Arial', size=9, color='#333333'),
        title=dict(
            text="<b>💧 Evapotranspiración diaria — FUNDOS AQUANQA</b>",
            font=dict(size=14, color='#1A1A1A'),
            x=0.0, xanchor='left', pad=dict(t=10, l=10)
        ),
        margin=dict(l=60, r=40, t=60, b=40),
        legend=dict(
            orientation='h', yanchor='top', y=-0.05,
            xanchor='left', x=0, font=dict(size=8),
            bgcolor='rgba(0,0,0,0)', borderwidth=0
        ),
        hovermode='x unified'
    )

    st.plotly_chart(fig, use_container_width=True, key="fig_et_main")
    
    # ── KPIs resumen ───────────────────────────────────────────
    st.markdown("#### 📊 Resumen ET por fundo")
    cols = st.columns(len(fundos))
    for i, fundo in enumerate(fundos):
        sub_et = dia[
            (dia['Fundo'] == fundo) &
            (dia['Fecha'] >= fecha_ini_anio)
        ]['ET'].dropna()
        if not sub_et.empty:
            cols[i].metric(
                fundo,
                f"{sub_et.sum():.1f} mm",
                f"Prom {sub_et.mean():.2f} mm/día"
            )



def generar_seccion_validacion(variable, dia, forecasts_cache, dias_pred_ui):
    import calendar
    
    # ── Obtener ventanas dinámicas ──
    fecha_max = pd.to_datetime(dia['Fecha'].max()).normalize()
    ultimo_dia_mes_ant = fecha_max.replace(day=1) - pd.Timedelta(days=1)
    primer_dia_mes_ant = ultimo_dia_mes_ant.replace(day=1)
    fecha_corte_1mes = primer_dia_mes_ant - pd.Timedelta(days=1)
    
    fecha_inicio_3m = primer_dia_mes_ant.replace(day=1) - pd.DateOffset(months=2)
    fecha_corte_3m = fecha_inicio_3m - pd.Timedelta(days=1)
    
    ventanas = {
        '1mes': {
            'label': f"{primer_dia_mes_ant.strftime('%B %Y')}",
            'inicio': primer_dia_mes_ant,
            'fin': ultimo_dia_mes_ant,
            'fecha_corte': fecha_corte_1mes,
            'dias': (ultimo_dia_mes_ant - primer_dia_mes_ant).days + 1,
        },
        '3mes': {
            'label': f"{fecha_inicio_3m.strftime('%b %Y')} → {ultimo_dia_mes_ant.strftime('%b %Y')}",
            'inicio': fecha_inicio_3m,
            'fin': ultimo_dia_mes_ant,
            'fecha_corte': fecha_corte_3m,
            'dias': (ultimo_dia_mes_ant - fecha_inicio_3m).days + 1,
        }
    }
    
    with st.expander(f"📊 Validación Prophet — {variable} (walk-forward honesto)", expanded=False):
        
        tab_1mes, tab_3mes = st.tabs([
            f"Último mes",
            f"Últimos 3 meses"
        ])
        
        # ════════════════════════════════════════════════════════
        # TAB 1: ÚLTIMO MES
        # ════════════════════════════════════════════════════════
        with tab_1mes:
            st.info(f"Entrenamiento: datos hasta {ventanas['1mes']['fecha_corte'].strftime('%d/%b/%Y')} | Predicción: {ventanas['1mes']['dias']} días")
            _validar_ventana(variable, dia, ventanas['1mes'], '_1mes')
        
        # ════════════════════════════════════════════════════════
        # TAB 2: ÚLTIMOS 3 MESES
        # ════════════════════════════════════════════════════════
        with tab_3mes:
            st.info(f"Entrenamiento: datos hasta {ventanas['3mes']['fecha_corte'].strftime('%d/%b/%Y')} | Predicción: {ventanas['3mes']['dias']} días")
            _validar_ventana(variable, dia, ventanas['3mes'], '_3mes')


def _validar_ventana(variable, dia, ventana, key_suffix):
    import calendar
    
    fundos = dia['Fundo'].unique().tolist()
    fecha_inicio = ventana['inicio']
    fecha_fin = ventana['fin']
    fecha_corte = ventana['fecha_corte']
    
    datos_boxplot = []
    
    for fundo in fundos:
        sub = dia[dia['Fundo'] == fundo].copy().reset_index(drop=True)
        if sub.empty:
            continue
        
        # ── Datos reales en el período ──
        real_periodo = sub[
            (sub['Fecha'] >= fecha_inicio) &
            (sub['Fecha'] <= fecha_fin)
        ].copy().reset_index(drop=True)
        
        if real_periodo.empty:
            continue
        
        # ── Entrenar SOLO con datos ANTES de la ventana ──
        train = sub[sub['Fecha'] <= fecha_corte].copy()
        
        if len(train) < 30:
            st.warning(f"⚠️ {fundo}: historial insuficiente")
            continue
        
        serie_train = train[['Fecha', variable]].rename(
            columns={'Fecha': 'ds', variable: 'y'}
        ).dropna()
        serie_train['ds'] = pd.to_datetime(
            serie_train['ds']
        ).dt.tz_localize(None).dt.normalize()
        
        media = serie_train['y'].mean()
        std = serie_train['y'].std()
        serie_train['y'] = serie_train['y'].clip(lower=media - 3*std, upper=media + 3*std)
        
        with st.spinner(f"Entrenando {fundo} {variable}..."):
            try:
                import warnings
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    m = Prophet(
                        yearly_seasonality=15,
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        seasonality_mode='additive',
                        changepoint_prior_scale=0.30,
                        seasonality_prior_scale=15.0,
                        interval_width=0.95,
                    )
                    m.add_seasonality(name='monthly', period=30.5, fourier_order=5)
                    m.fit(serie_train)
                    
                    dias_pred_period = (fecha_fin - fecha_inicio).days + 1
                    future = m.make_future_dataframe(periods=dias_pred_period, freq='D')
                    forecast_wf = m.predict(future)
            
            except Exception as e:
                st.error(f"❌ Error {fundo}: {e}")
                continue
        
        forecast_wf['ds'] = pd.to_datetime(forecast_wf['ds']).dt.tz_localize(None).dt.normalize()
        
        pred_periodo = forecast_wf[
            (forecast_wf['ds'] >= fecha_inicio) &
            (forecast_wf['ds'] <= fecha_fin)
        ][['ds', 'yhat', 'yhat_lower', 'yhat_upper']].copy().reset_index(drop=True)
        
        if pred_periodo.empty:
            continue
        
        # ── Comparación ──
        real_periodo = real_periodo.rename(columns={variable: 'Real'})
        comparacion = real_periodo[['Fecha', 'Real']].merge(
            pred_periodo.rename(columns={
                'ds': 'Fecha',
                'yhat': 'Pred',
                'yhat_lower': 'Pred_Low',
                'yhat_upper': 'Pred_High',
            }),
            on='Fecha', how='inner'
        )
        
        if comparacion.empty:
            continue
        
        comparacion['Error_abs'] = (comparacion['Real'] - comparacion['Pred']).abs().round(2)
        comparacion['Error_signed'] = (comparacion['Real'] - comparacion['Pred']).round(2)
        
        mae_mes = comparacion['Error_abs'].mean().round(2)
        mbe_mes = comparacion['Error_signed'].mean().round(2)
        rmse_mes = round(float(np.sqrt((comparacion['Error_signed']**2).mean())), 2)
        
        dias_dentro   = int((comparacion['Error_abs'] <= 1.0).sum())
        dias_total    = len(comparacion)
        precision_mae = round(dias_dentro / dias_total * 100, 2)
        dias_fuera    = dias_total - dias_dentro

        # Ratio RMSE/MAE — si > 1.3 hay días problemáticos
        ratio_rmse_mae = round(rmse_mes / mae_mes, 2) if mae_mes > 0 else None
        
        for _, r in comparacion.iterrows():
            datos_boxplot.append({'Fundo': fundo, 'Error': r['Error_signed']})
        
        if 'bias_correccion' not in st.session_state:
            st.session_state['bias_correccion'] = {}
        st.session_state['bias_correccion'][(fundo, variable)] = float(mbe_mes)
        
        # ── Métricas ──
        st.markdown(f"##### {fundo} — {ventana['label']}")
        
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("MAE", f"{mae_mes:.2f}°C",
        help="Error promedio diario del mes")

        k2.metric("RMSE", f"{rmse_mes:.2f}°C",
                delta=f"ratio {ratio_rmse_mae}x" if ratio_rmse_mae else None,
                help="Si ratio > 1.3: hay días con errores grandes")

        k3.metric(
            "BIAS (MBE)", f"{mbe_mes:+.2f}°C",
            delta="subestima" if mbe_mes > 0 else "sobreestima",
            delta_color="inverse" if mbe_mes > 0 else "normal",
            help="Sesgo sistemático — se corrige automáticamente"
        )

        k4.metric(
            "Precisión (±1°C)",
            f"{precision_mae:.1f}%",
            delta=f"{dias_dentro}/{dias_total} días dentro",
            delta_color="normal" if precision_mae >= 60 else "inverse",
            help="% de días con error dentro de ±1°C — umbral operacional"
        )
        
        # ── Gráfico error ──
        fig_val = go.Figure()
        
        fig_val.add_hrect(
            y0=-1, y1=1,
            fillcolor='rgba(76, 175, 80, 0.10)',
            line_width=0,
        )
        
        fig_val.add_hline(y=0, line=dict(color='#546E7A', width=1.5))
        
        for j in range(len(comparacion) - 1):
            x0 = comparacion['Fecha'].iloc[j]
            x1 = comparacion['Fecha'].iloc[j + 1]
            e0 = comparacion['Error_signed'].iloc[j]
            e1 = comparacion['Error_signed'].iloc[j + 1]
            
            if e0 >= 0 and e1 >= 0:
                fig_val.add_trace(go.Scatter(
                    x=[x0, x1, x1, x0], y=[e0, e1, 0, 0],
                    fill='toself', fillcolor='rgba(76,175,80,0.30)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
            elif e0 < 0 and e1 < 0:
                fig_val.add_trace(go.Scatter(
                    x=[x0, x1, x1, x0], y=[e0, e1, 0, 0],
                    fill='toself', fillcolor='rgba(244,67,54,0.30)',
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
            else:
                frac = abs(e0) / (abs(e0) + abs(e1))
                rango = (x1 - x0).total_seconds()
                x_cross = x0 + pd.Timedelta(seconds=rango * frac)
                
                color_izq = 'rgba(76,175,80,0.30)' if e0 >= 0 else 'rgba(244,67,54,0.30)'
                fig_val.add_trace(go.Scatter(
                    x=[x0, x_cross, x_cross, x0], y=[e0, 0, 0, 0],
                    fill='toself', fillcolor=color_izq,
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
                
                color_der = 'rgba(76,175,80,0.30)' if e1 >= 0 else 'rgba(244,67,54,0.30)'
                fig_val.add_trace(go.Scatter(
                    x=[x_cross, x1, x1, x_cross], y=[0, e1, 0, 0],
                    fill='toself', fillcolor=color_der,
                    line=dict(color='rgba(0,0,0,0)'),
                    showlegend=False, hoverinfo='skip'
                ))
        
        fig_val.add_trace(go.Scatter(
            x=comparacion['Fecha'],
            y=comparacion['Error_signed'],
            mode='lines+markers',
            line=dict(color='#1A237E', width=2.5),
            marker=dict(
                size=7,
                color=comparacion['Error_signed'].apply(lambda e: '#4CAF50' if e >= 0 else '#F44336'),
                line=dict(color='white', width=1)
            ),
            name='Error',
            hovertemplate='%{x|%d/%b}<br>Error: %{y:+.2f}°C<extra></extra>'
        ))
        
        fig_val.add_annotation(
            x=comparacion['Fecha'].iloc[len(comparacion) // 2],
            y=2.8,
            text=f"<b>MBE: {mbe_mes:+.2f}°C</b><br>{'🔺 subestima' if mbe_mes > 0 else '🔻 sobreestima'}<br>MAE: {mae_mes:.2f}°C",
            showarrow=False,
            bgcolor='rgba(255,255,255,0.85)',
            bordercolor='#546E7A',
            borderwidth=1,
            font=dict(size=10, color='#1A1A1A'),
            align='center'
        )
        
        fig_val.update_layout(
            height=350,
            title=dict(text=f"<b>Error diario — {fundo} — {variable}</b>", font=dict(size=12), x=0.0),
            xaxis=dict(tickformat='%d/%b', gridcolor='#CFD8DC'),
            yaxis=dict(
                title='Error (°C)',
                ticksuffix='°C',
                gridcolor='#CFD8DC',
                range=[-3.5, 3.5],
                dtick=1,
                zeroline=False,
            ),
            legend=dict(orientation='h', y=-0.15, x=0),
            hovermode='x unified',
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Arial', size=9),
            margin=dict(l=50, r=30, t=60, b=40),
        )
        
        st.plotly_chart(fig_val, use_container_width=True, key=f"fig_val_{variable}_{fundo}{key_suffix}")
        
        st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    
    # ── BOX PLOT ──
    if datos_boxplot:
        df_box = pd.DataFrame(datos_boxplot)
        
        fig_box = go.Figure()
        
        colores_box = ['#1565C0', '#C62828', '#2E7D32', '#F57F17', '#6A1B9A', '#00838F']
        
        for idx_f, fundo in enumerate(fundos):
            df_f = df_box[df_box['Fundo'] == fundo]['Error']
            if df_f.empty:
                continue
            
            color = colores_box[idx_f % len(colores_box)]
            
            fig_box.add_trace(go.Box(
                y=df_f,
                name=fundo,
                boxpoints='all',
                jitter=0.4,
                pointpos=0,
                marker=dict(color=color, size=5, opacity=0.6, line=dict(color='white', width=0.5)),
                line=dict(color=color, width=2),
                fillcolor=f'rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.15)',
                hovertemplate=f'<b>{fundo}</b><br>Error: %{{y:+.2f}}°C<extra></extra>'
            ))
        
        fig_box.add_hrect(y0=-1, y1=1, fillcolor='rgba(76, 175, 80, 0.08)', line_width=0)
        fig_box.add_hline(y=1.0, line=dict(color='#4CAF50', width=1.5, dash='dot'))
        fig_box.add_hline(y=-1.0, line=dict(color='#4CAF50', width=1.5, dash='dot'))
        
        fig_box.update_layout(
            height=420,
            title=dict(text=f"<b>📦 Distribución Error — {variable}</b>", font=dict(size=13, color='#1A1A1A'), x=0.0),
            xaxis=dict(title='Fundo', gridcolor='#CFD8DC'),
            yaxis=dict(
                title='Error (°C)',
                ticksuffix='°C',
                gridcolor='#CFD8DC',
                tickvals=[-3, -2, -1, 0, 1, 2, 3],
                range=[-3.5, 3.5],
            ),
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Arial', size=9, color='#333333'),
            showlegend=False,
            margin=dict(l=60, r=40, t=60, b=40),
        )
        
        st.plotly_chart(fig_box, use_container_width=True, key=f"fig_box_{variable}{key_suffix}")
        
        # ── GRÁFICO APILADO ──
        def categorizar_error(e):
            if e < -1:
                return '< -1°C'
            elif e < 0:
                return '-1 a 0°C'
            elif e <= 1:
                return '0 a 1°C'
            else:
                return '> 1°C'
        
        df_box['Categoria'] = df_box['Error'].apply(categorizar_error)
        
        categorias_orden = ['< -1°C', '-1 a 0°C', '0 a 1°C', '> 1°C']
        colores_cat = {
            '< -1°C': '#E53935',
            '-1 a 0°C': '#FFAB40',
            '0 a 1°C': '#66BB6A',
            '> 1°C': '#1E88E5',
        }
        
        tabla_pct = df_box.groupby(['Fundo', 'Categoria']).size().reset_index(name='N')
        total_por_fundo = tabla_pct.groupby('Fundo')['N'].transform('sum')
        tabla_pct['Pct'] = (tabla_pct['N'] / total_por_fundo * 100).round(1)
        
        fig_stack = go.Figure()
        
        fundos_orden = sorted(df_box['Fundo'].unique().tolist())
        
        for cat in categorias_orden:
            df_cat = tabla_pct[tabla_pct['Categoria'] == cat]
            
            pcts, ns = [], []
            for fundo in fundos_orden:
                row_cat = df_cat[df_cat['Fundo'] == fundo]
                if not row_cat.empty:
                    pcts.append(float(row_cat['Pct'].iloc[0]))
                    ns.append(int(row_cat['N'].iloc[0]))
                else:
                    pcts.append(0.0)
                    ns.append(0)
            
            fig_stack.add_trace(go.Bar(
                name=cat,
                x=fundos_orden,
                y=pcts,
                marker_color=colores_cat[cat],
                text=[f'{p:.1f}%' if p > 4 else '' for p in pcts],
                textposition='inside',
                textfont=dict(size=11, color='white', family='Arial'),
                customdata=ns,
                hovertemplate='<b>%{x}</b><br>' + cat + '<br>Días: %{customdata}<br>%{y:.1f}%<extra></extra>'
            ))
        
        fig_stack.update_layout(
            barmode='stack',
            height=380,
            title=dict(text=f"<b>📊 Rangos Error — {variable}</b>", font=dict(size=13), x=0.0),
            xaxis_title='Fundo',
            yaxis=dict(title='% de días', ticksuffix='%', range=[0, 100], gridcolor='#CFD8DC', dtick=25),
            legend=dict(orientation='h', y=1.02, x=0),
            paper_bgcolor='#FFFFFF',
            plot_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=60, r=40, t=80, b=40),
        )
        
        st.plotly_chart(fig_stack, use_container_width=True, key=f"fig_stack_{variable}{key_suffix}")
            
def generar_range_plot(variable, historico_df, prediccion_df, fundo_name, row, fig):
    """Agrega Range Plot con Error Bars (barras Tmin-Tmax)."""
    
    # Range Plot de datos reales
    if not historico_df.empty:
        fechas = historico_df['Fecha'].tolist()
        tmax = historico_df['Tmax'].tolist()
        tmin = historico_df['Tmin'].tolist()
        smooth = historico_df[f'{variable}_smooth'].tolist()
        et = historico_df['ET'].tolist()
        
        # Calcular error bars (distancia de smooth a Tmax y Tmin)
        error_plus = [t - s for t, s in zip(tmax, smooth)]
        error_minus = [s - t for t, s in zip(tmin, smooth)]
        
        fig.add_trace(go.Scatter(
            x=fechas,
            y=smooth,
            mode='lines+markers',
            line=dict(color='#000000', width=2),
            marker=dict(size=5, color='#00AA00', line=dict(color='#000000', width=1)),
            error_y=dict(
                type='data',
                symmetric=False,
                array=error_plus,
                arrayminus=error_minus,
                color='#00AA00',
                thickness=2,
                width=3
            ),
            name=f'{variable} REAL',
            showlegend=True,
            legendgroup=fundo_name,
            hovertemplate='%{x|%d/%b/%Y}<br>' + f'{variable}: %{{y:.1f}}°C<br>ET: %{{customdata:.1f}}<extra>{fundo_name}</extra>',
            customdata=et
        ), row=row, col=1)
    
    # Range Plot de predicción
    if not prediccion_df.empty:
        fechas_pred = prediccion_df['ds'].tolist()
        yhat = prediccion_df['yhat'].tolist()
        upper = prediccion_df['yhat_upper'].tolist()
        lower = prediccion_df['yhat_lower'].tolist()
        
        # Calcular error bars para predicción
        error_plus_pred = [u - y for u, y in zip(upper, yhat)]
        error_minus_pred = [y - l for y, l in zip(yhat, lower)]
        
        fig.add_trace(go.Scatter(
            x=fechas_pred,
            y=yhat,
            mode='lines+markers',
            line=dict(color='#FF0000', width=2),
            marker=dict(size=5, color='#FF9800', line=dict(color='#FF0000', width=1)),
            error_y=dict(
                type='data',
                symmetric=False,
                array=error_plus_pred,
                arrayminus=error_minus_pred,
                color='#FF9800',
                thickness=2,
                width=3
            ),
            name=f'Predicción +{len(fechas_pred)}d',
            showlegend=True,
            legendgroup=fundo_name,
            hovertemplate='%{x|%d/%b/%Y}<br>Pred: %{y:.1f}°C<extra>Prophet</extra>'
        ), row=row, col=1)

def generar_figura(variable, dia, media_mensual, q1_mensual, q3_mensual,
                   dias_pred, forecasts_cache, dias_vista=30,
                   dias_pred_mostrar=None, tipo_viz='linea',
                   media_mensual2=None, q1_mensual2=None, q3_mensual2=None,
                   label_clim2=None):
    """Generación de figura optimizada."""
    fundos = dia['Fundo'].unique().tolist()
    n = len(fundos)
    
    if variable == 'Tmax':
        fig_bg, banda_color, clim_color, color_real, pred_color = (
            BG_TMAX, BANDA_TMAX, CLIM_TMAX, REAL_TMAX_COLOR, PRED_COLOR_TMAX
        )
    else:
        fig_bg, banda_color, clim_color, color_real, pred_color = (
            BG_TMIN, BANDA_TMIN, CLIM_TMIN, REAL_TMIN_COLOR, PRED_COLOR_TMIN
        )
    
    fig = make_subplots(
        rows=n, cols=1,
        shared_xaxes=False,
        vertical_spacing=0.08,
        subplot_titles=[f for f in fundos]
    )
    
    rows_export, rows_pred_export = [], []
    anio_actual = pd.Timestamp.today().year
    fecha_ini_anio = pd.Timestamp(year=anio_actual - 1, month=1, day=1)
    
    import calendar
    
    for i, fundo in enumerate(fundos):
        row = i + 1
        
        sub_full = dia[dia['Fundo'] == fundo].copy().reset_index(drop=True)
        if sub_full.empty:
            continue
        
        sub_full = sub_full[sub_full['Fecha'] >= fecha_ini_anio].copy().reset_index(drop=True)
        if sub_full.empty:
            st.warning(f"⚠️ Sin datos en {anio_actual} para **{fundo}**.")
            continue

        # ── Fechas clave ───────────────────────────────────────────
        # fecha_corte_real = último dato disponible (ej. 10 Jun 2026)
        fecha_corte_real    = pd.to_datetime(sub_full['Fecha'].max()).tz_localize(None).normalize()
        primer_dia_mes_actual = fecha_corte_real.replace(day=1)
        # fecha_fin_norm = FDM anterior (31 May) — solo para referencia/export
        fecha_fin_norm      = primer_dia_mes_actual - pd.Timedelta(days=1)
        # Último día del mes actual (30 Jun)
        ultimo_dia_mes_actual = pd.Timestamp(
            year=primer_dia_mes_actual.year,
            month=primer_dia_mes_actual.month,
            day=calendar.monthrange(
                primer_dia_mes_actual.year,
                primer_dia_mes_actual.month
            )[1]
        )

        # ── CAMBIO CLAVE: sub para graficar = TODOS los datos disponibles
        #    (incluyendo 1-10 Jun), NO cortado a FDM ──────────────────
        sub = sub_full.copy().reset_index(drop=True)

        if sub.empty:
            continue
        
        empresa = sub['Empresa'].iloc[0]
        
        zoom_ini = sub['Fecha'].min()
        zoom_fin = ultimo_dia_mes_actual + pd.Timedelta(days=2)
        
        # Splines con cache (sin cambios)
        fecha_ini_clima = sub['Fecha'].min()
        if hasattr(fecha_ini_clima, 'tz_localize'):
            fecha_ini_clima = pd.Timestamp(fecha_ini_clima).tz_localize(None).normalize()
        fecha_fin_clima = pd.Timestamp(zoom_fin)
        
        clim_media = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(media_mensual)
        )
        clim_q1 = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(q1_mensual)
        )
        clim_q3 = spline_diario_cached(
            fecha_ini_clima.strftime('%Y-%m-%d'),
            fecha_fin_clima.strftime('%Y-%m-%d'),
            tuple(q3_mensual)
        )
        
        # Merge para export (igual que antes)
        hist_df = sub[['Empresa', 'Fundo', 'Fecha', variable, f'{variable}_smooth']].copy()
        hist_df = hist_df.merge(clim_media.rename(columns={'Valor': 'Clim_MEDIA'}), on='Fecha', how='left')
        hist_df = hist_df.merge(clim_q1.rename(columns={'Valor': 'Clim_Q1'}), on='Fecha', how='left')
        hist_df = hist_df.merge(clim_q3.rename(columns={'Valor': 'Clim_Q3'}), on='Fecha', how='left')
        hist_df['Variable'] = variable
        rows_export.append(hist_df)
        
        # Forecasts
        cache_entry = forecasts_cache.get((fundo, variable), {})
        forecast    = cache_entry.get('forecast', pd.DataFrame())
        mae         = cache_entry.get('mae', None)
        mae_por_dia = cache_entry.get('mae_por_dia', {})
        
        if not forecast.empty:
            forecast = forecast.copy()
            forecast['ds'] = pd.to_datetime(forecast['ds']).dt.tz_localize(None).dt.normalize()

            # ── CAMBIO CLAVE: predicción solo desde el día siguiente al
            #    último dato real hasta fin de mes ─────────────────────
            primer_dia_pred = fecha_corte_real + pd.Timedelta(days=1)

            pred_fc = forecast[
                (forecast['ds'] >= primer_dia_pred) &
                (forecast['ds'] <= ultimo_dia_mes_actual)
            ].reset_index(drop=True)
        else:
            pred_fc = pd.DataFrame()

        # Ajuste de residuos (sin cambios)
        if not pred_fc.empty and len(sub) >= 14:
            ventana     = min(21, len(sub))
            sub_reciente = sub.tail(ventana).copy().reset_index(drop=True)
            residuos    = (sub_reciente[variable] - sub_reciente[f'{variable}_smooth']).values

            from scipy.ndimage import gaussian_filter1d
            residuos_suaves = gaussian_filter1d(residuos, sigma=0.8)

            n_pred   = len(pred_fc)
            n_res    = len(residuos_suaves)
            indices  = [ii % n_res for ii in range(n_pred)]
            patron   = residuos_suaves[indices]

            std_hist = residuos_suaves.std()
            std_pred = pred_fc['yhat'].std()
            factor   = min(1.0, std_hist / (std_pred + 1e-6)) * 0.6

            pred_fc = pred_fc.copy()
            pred_fc['yhat']       = (pred_fc['yhat']       + patron * factor).round(2)
            pred_fc['yhat_lower'] = (pred_fc['yhat_lower'] + patron * factor).round(2)
            pred_fc['yhat_upper'] = (pred_fc['yhat_upper'] + patron * factor).round(2)

            bias_val = st.session_state.get(
                'bias_correccion', {}
            ).get((fundo, variable), 0.0)

            if bias_val != 0.0:
                pred_fc['yhat']       = (pred_fc['yhat']       + bias_val).round(2)
                pred_fc['yhat_lower'] = (pred_fc['yhat_lower'] + bias_val).round(2)
                pred_fc['yhat_upper'] = (pred_fc['yhat_upper'] + bias_val).round(2)

        lg = fundo
        
        # ── Trazas climatología Guadalupe (fija) — sin cambios ────────
        fig.add_trace(go.Scatter(
            x=clim_q3['Fecha'], y=clim_q3['Valor'],
            mode='lines', line=dict(width=0),
            showlegend=False, hoverinfo='skip', legendgroup=lg
        ), row=row, col=1)
        fig.add_trace(go.Scatter(
            x=clim_q1['Fecha'], y=clim_q1['Valor'],
            mode='lines', line=dict(width=0),
            fill='tonexty', fillcolor=_hex_to_rgba(banda_color, 0.55),
            showlegend=(i == 0), legendgroup=lg,
            name=f'Q1-Q3 SENAMHI Guadalupe (±{DELTA_Q}°C)',
            hovertemplate='%{x|%d/%b}<br>Q1: %{y:.1f}°C<extra></extra>'
        ), row=row, col=1)
        
        fig.add_trace(go.Scatter(
            x=clim_media['Fecha'], y=clim_media['Valor'],
            mode='lines', line=dict(color=clim_color, width=2.8),
            showlegend=(i == 0), legendgroup=lg,
            name=f'{variable} climatología SENAMHI Guadalupe',
            hovertemplate='%{x|%d/%b}<br>Clim Guadalupe: %{y:.1f}°C<extra></extra>'
        ), row=row, col=1)

        # ── Segunda climatología dinámica (sin cambios) ───────────────
        if media_mensual2 is not None and label_clim2:
            clim_media2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(media_mensual2)
            )
            clim_q1_2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(q1_mensual2)
            )
            clim_q3_2 = spline_diario_cached(
                fecha_ini_clima.strftime('%Y-%m-%d'),
                fecha_fin_clima.strftime('%Y-%m-%d'),
                tuple(q3_mensual2)
            )

            fig.add_trace(go.Scatter(
                x=clim_q3_2['Fecha'], y=clim_q3_2['Valor'],
                mode='lines', line=dict(width=0),
                showlegend=False, hoverinfo='skip', legendgroup=lg
            ), row=row, col=1)
            fig.add_trace(go.Scatter(
                x=clim_q1_2['Fecha'], y=clim_q1_2['Valor'],
                mode='lines', line=dict(width=0),
                fill='tonexty',
                fillcolor=_hex_to_rgba(
                    BANDA2_TMAX if variable == 'Tmax' else BANDA2_TMIN, 0.15
                ),
                showlegend=(i == 0), legendgroup=lg,
                name=f'Q1-Q3 SENAMHI {label_clim2} (±{DELTA_Q}°C)',
                hovertemplate='%{x|%d/%b}<br>Q1: %{y:.1f}°C<extra></extra>'
            ), row=row, col=1)

            fig.add_trace(go.Scatter(
                x=clim_media2['Fecha'], y=clim_media2['Valor'],
                mode='lines',
                line=dict(color='rgba(255,111,0,0.70)', width=1.8, dash='dot'),
                showlegend=(i == 0), legendgroup=lg,
                name=f'{variable} climatología SENAMHI {label_clim2}',
                hovertemplate='%{x|%d/%b}<br>Clim ' + label_clim2 + ': %{y:.1f}°C<extra></extra>'
            ), row=row, col=1)

        # ── Datos reales — AHORA incluye datos del mes actual ─────────
        if tipo_viz == 'candlestick':
            if not pred_fc.empty:
                if dias_pred_mostrar is not None and dias_pred_mostrar < len(pred_fc):
                    pred_fc = pred_fc.iloc[:dias_pred_mostrar].copy()
                generar_range_plot(variable, sub, pred_fc, fundo, row, fig)
            else:
                generar_range_plot(variable, sub, pd.DataFrame(), fundo, row, fig)
        else:
            fig.add_trace(go.Scatter(
                x=sub['Fecha'], y=sub[f'{variable}_smooth'],
                mode='lines+markers',
                line=dict(color=color_real, width=1.5),
                marker=dict(size=5, color='white', line=dict(color=color_real, width=1.0)),
                customdata=sub['ET'],
                showlegend=(i == 0), legendgroup=lg,
                name=f'{variable} REAL',
                hovertemplate='%{x|%d/%b/%Y}<br>' + f'{variable}: %{{y:.1f}}°C<br>ET: %{{customdata:.1f}}<extra>{fundo}</extra>'
            ), row=row, col=1)

        # Anotación último valor real (ahora apunta a fecha_corte_real)
        ult = sub.iloc[-1]
        ult_real_val = ult[variable]
        fig.add_annotation(
            x=ult['Fecha'], y=ult_real_val,
            text=f"<b>{ult_real_val:.1f}°C</b>",
            showarrow=True, arrowhead=2, arrowcolor=color_real, arrowwidth=1.0,
            ax=14, ay=-18, font=dict(size=9, color=color_real),
            row=row, col=1
        )
        
        # ── CAMBIO CLAVE: línea vertical en el último dato real ───────
        #    (antes era fecha_fin_norm = 31 Mayo)
        fig.add_vline(
            x=fecha_corte_real,
            line=dict(color='#546E7A', width=1.2, dash='dot'),
            row=row, col=1
        )
        
        # ── Predicciones modo línea ────────────────────────────────────
        if tipo_viz == 'linea' and not pred_fc.empty:
            if dias_pred_mostrar is not None and dias_pred_mostrar < len(pred_fc):
                pred_fc = pred_fc.iloc[:dias_pred_mostrar].copy()
            
            # Conector desde último dato real → primer día predicho
            ult_smooth = float(ult[f'{variable}_smooth'])
            
            fig.add_trace(go.Scatter(
                x=[ult['Fecha'], pred_fc['ds'].iloc[0]],
                y=[ult_smooth, float(pred_fc['yhat'].iloc[0])],
                mode='lines',
                line=dict(
                    color=f'rgba({int(pred_color[1:3],16)},{int(pred_color[3:5],16)},{int(pred_color[5:7],16)},0.70)',
                    width=1.5, dash='dot'
                ),
                showlegend=False, hoverinfo='skip', legendgroup=lg
            ), row=row, col=1)
            
            # Banda IC 95%
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat_lower'].tolist(),
                mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'),
                showlegend=False, hoverinfo='skip', legendgroup=lg, name='_lower'
            ), row=row, col=1)
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat_upper'].tolist(),
                mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'),
                fill='tonexty', fillcolor=_hex_to_rgba(pred_color, 0.20),
                showlegend=(i == 0), name='IC 95%',
                legendgroup=lg, hoverinfo='skip'
            ), row=row, col=1)
            
            # Línea predicción
            fig.add_trace(go.Scatter(
                x=pred_fc['ds'].tolist(), y=pred_fc['yhat'].tolist(),
                mode='lines',
                line=dict(
                    color=f'rgba({int(pred_color[1:3],16)},{int(pred_color[3:5],16)},{int(pred_color[5:7],16)},0.70)',
                    width=1.8, dash='dot'
                ),
                showlegend=(i == 0), legendgroup=lg,
                name='Predicción mes siguiente',
                hovertemplate='%{x|%d/%b/%Y}<br>Pred: %{y:.1f}°C<extra>Predicción</extra>'
            ), row=row, col=1)
            
            # Anotación último valor predicho
            ult_pred_dt  = pred_fc['ds'].iloc[-1]
            ult_pred_val = float(pred_fc['yhat'].iloc[-1])
            fig.add_annotation(
                x=ult_pred_dt, y=ult_pred_val,
                text=f"<b>{ult_pred_val:.1f}°C</b>",
                showarrow=True, arrowhead=2, arrowcolor=pred_color, arrowwidth=1.0,
                ax=-28, ay=-18, font=dict(size=9, color=pred_color),
                row=row, col=1
            )
        
        # ── Título subplot (sin cambios) ───────────────────────────────
        bg_color = '#C62828' if variable == 'Tmax' else '#1565C0'
        fig.layout.annotations[i].update(
            text=f"  <b>{fundo}</b>  ",
            font=dict(size=12, color='#FFFFFF'),
            bgcolor=bg_color,
            borderpad=4,
            bordercolor=bg_color,
            borderwidth=2,
            x=0.0, xanchor='left'
        )
        
        xk = 'xaxis' if row == 1 else f'xaxis{row}'
        fig.layout[xk].update(
            range=[zoom_ini, zoom_fin],
            tickformat='%d/%b', ticklabelmode='period',
            gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
            autorange=False, fixedrange=False,
            rangeslider=dict(visible=False)
        )
        yk = 'yaxis' if row == 1 else f'yaxis{row}'
        fig.layout[yk].update(
            ticksuffix='°C', gridcolor=GRID_COLOR, gridwidth=0.5,
            showgrid=True, zeroline=False,
        )

    fig.update_layout(
        height=320 * n,
        paper_bgcolor=fig_bg,
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(family='Arial', size=9, color='#333333'),
        title=dict(
            text=f"",
            font=dict(size=14, color='#1A1A1A'),
            x=0.0, xanchor='left', pad=dict(t=10, l=10)
        ),
        margin=dict(l=60, r=40, t=60, b=40),
        legend=dict(
            orientation='h', yanchor='top', y=-0.08,
            xanchor='left', x=0, font=dict(size=8),
            bgcolor='rgba(0,0,0,0)', borderwidth=0
        ),
        hovermode='x unified'
    )
    
    df_hist = pd.concat(rows_export, ignore_index=True) if rows_export else pd.DataFrame()
    df_pred = pd.concat(rows_pred_export, ignore_index=True) if rows_pred_export else pd.DataFrame()
    
    return fig, df_hist, df_pred



# ══════════════════════════════════════════════════════════════
# EXPORTAR EXCEL (igual que antes)
# ══════════════════════════════════════════════════════════════
def _generar_pred_horaria(df_pred: pd.DataFrame) -> pd.DataFrame:
    if df_pred.empty:
        return pd.DataFrame()
    
    rows = []
    grupos = df_pred.groupby(['Empresa', 'Fundo', 'Variable'])
    
    for (empresa, fundo, variable), grp in grupos:
        grp = grp.sort_values('Fecha').reset_index(drop=True)
        fechas = pd.to_datetime(grp['Fecha'])
        preds = grp['Pred'].values.astype(float)
        lows = grp['Pred_Low'].values.astype(float)
        highs = grp['Pred_High'].values.astype(float)
        
        n = len(fechas)
        x_ctrl = np.arange(n, dtype=float) * 24.0
        
        if n < 2:
            x_ctrl = np.array([0.0, 24.0])
            preds = np.array([preds[0], preds[0]])
            lows = np.array([lows[0], lows[0]])
            highs = np.array([highs[0], highs[0]])
        
        cs_pred = CubicSpline(x_ctrl, preds)
        cs_low = CubicSpline(x_ctrl, lows)
        cs_high = CubicSpline(x_ctrl, highs)
        
        horas_rango = pd.date_range(
            start=fechas.iloc[0].normalize(),
            end=fechas.iloc[-1].normalize() + pd.Timedelta(hours=23),
            freq='h'
        )
        
        for h_dt in horas_rango:
            x_h = (h_dt - fechas.iloc[0].normalize()).total_seconds() / 3600.0
            x_h = float(np.clip(x_h, x_ctrl[0], x_ctrl[-1]))
            hora = h_dt.hour
            
            base = float(cs_pred(x_h))
            lo = float(cs_low(x_h))
            hi = float(cs_high(x_h))
            ampl = max((hi - lo) / 4.0, 0.0)
            
            if variable == 'Tmax':
                offset = ampl * np.sin(2 * np.pi * (hora - 5) / 24)
            else:
                offset = -ampl * np.sin(2 * np.pi * (hora - 5) / 24)
            
            rows.append({
                'Empresa': empresa,
                'Fundo': fundo,
                'Fecha': h_dt.strftime('%Y/%m/%d'),
                'Hora': h_dt.strftime('%H:00'),
                'FechaHora': h_dt,
                'Variable': variable,
                'max_pred': round(base + offset, 1) if variable == 'Tmax' else None,
                'min_pred': round(base + offset, 1) if variable == 'Tmin' else None,
                'Tipo': 'PREDICCION',
            })
    
    return pd.DataFrame(rows)

def exportar_excel(df_hist: pd.DataFrame, df_pred: pd.DataFrame) -> bytes:
    HDR_HIST = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
    HDR_HORA = PatternFill('solid', start_color='1B5E20', end_color='1B5E20')
    HDR_DET = PatternFill('solid', start_color='37474F', end_color='37474F')
    REAL_F = PatternFill('solid', start_color='FFFF00', end_color='FFFF00')
    PRED_F = PatternFill('solid', start_color='FFE0B2', end_color='FFE0B2')
    ALT_F = PatternFill('solid', start_color='F3F4F6', end_color='F3F4F6')
    
    HDR_FNT_W = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    REAL_FNT = Font(name='Arial', size=8, color='000000')
    PRED_FNT = Font(name='Arial', size=8, color='BF360C', bold=True)
    CENT = Alignment(horizontal='center', vertical='center')
    
    def _autowidth(ws, max_w=22):
        for col in ws.columns:
            maxl = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(maxl + 3, max_w)
    
    wb = Workbook()
    
    # HOJA 1: Por Fechas
    ws_dia = wb.active
    ws_dia.title = 'Por_Fechas'
    ws_dia.sheet_properties.tabColor = '1E3A5F'
    
    filas_dia = []
    # Obtener fundos de ambos DataFrames, validando que no estén vacíos
    fundos_hist = list(df_hist['Fundo'].unique()) if not df_hist.empty and 'Fundo' in df_hist.columns else []
    fundos_pred = list(df_pred['Fundo'].unique()) if not df_pred.empty and 'Fundo' in df_pred.columns else []
    fundos_todos = sorted(set(fundos_hist + fundos_pred))
    
    for fundo in fundos_todos:
        h = df_hist[df_hist['Fundo'] == fundo].copy() if not df_hist.empty and 'Fundo' in df_hist.columns else pd.DataFrame()
        p = df_pred[df_pred['Fundo'] == fundo].copy() if not df_pred.empty and 'Fundo' in df_pred.columns else pd.DataFrame()
        
        empresa = (
            h['Empresa'].iloc[0] if not h.empty and 'Empresa' in h.columns else
            (p['Empresa'].iloc[0] if not p.empty and 'Empresa' in p.columns else '')
        )
        
        if 'Tmax' in h.columns and 'Tmin' in h.columns:
            hist_piv = (
                h[['Fecha', 'Tmax', 'Tmin']].drop_duplicates()
                 .rename(columns={'Tmax': 'max', 'Tmin': 'min'})
            )
        else:
            hist_piv = pd.DataFrame()
        
        if 'Pred' in p.columns:
            p_max = p[p['Variable'] == 'Tmax'][['Fecha', 'Pred']].rename(columns={'Pred': 'max pred'})
            p_min = p[p['Variable'] == 'Tmin'][['Fecha', 'Pred']].rename(columns={'Pred': 'min pred'})
            pred_piv = (
                p_max.merge(p_min, on='Fecha', how='outer') if not p_max.empty else
                pd.DataFrame(columns=['Fecha', 'max pred', 'min pred'])
            )
        else:
            pred_piv = pd.DataFrame()
        
        # Merge seguro: validar que ambos tengan 'Fecha' antes de hacer merge
        if not hist_piv.empty and not pred_piv.empty and 'Fecha' in hist_piv.columns and 'Fecha' in pred_piv.columns:
            tabla = hist_piv.merge(pred_piv, on='Fecha', how='outer')
        elif not hist_piv.empty:
            tabla = hist_piv.copy()
        elif not pred_piv.empty:
            tabla = pred_piv.copy()
        else:
            tabla = pd.DataFrame()
        
        if not tabla.empty:
            tabla['Fecha'] = pd.to_datetime(tabla['Fecha'])
            tabla = tabla.sort_values('Fecha').reset_index(drop=True)
            tabla['Empresa'] = empresa
            tabla['Fundo'] = fundo
            
            def _status(r):
                v = r.get('max', np.nan)
                try:
                    return 'real' if pd.notna(v) and not np.isnan(float(v)) else 'predicha'
                except Exception:
                    return 'predicha'
            tabla['status'] = tabla.apply(_status, axis=1)
            filas_dia.append(tabla)
    
    tabla_dia = pd.concat(filas_dia, ignore_index=True) if filas_dia else pd.DataFrame()
    tabla_dia = tabla_dia.sort_values(['Fundo', 'Fecha']).reset_index(drop=True)
    
    cols_dia = ['Empresa', 'Fundo', 'Fecha', 'max', 'min', 'max pred', 'min pred', 'status']
    for ci, h_txt in enumerate(cols_dia, 1):
        c = ws_dia.cell(row=1, column=ci)
        c.value = h_txt; c.fill = HDR_HIST
        c.font = HDR_FNT_W; c.alignment = CENT; c.border = THIN
    
    if not tabla_dia.empty:
        for ri, row_data in tabla_dia.iterrows():
            er = ri + 2
            is_real = row_data.get('status', 'predicha') == 'real'
            fill = REAL_F if is_real else PRED_F
            font = REAL_FNT if is_real else PRED_FNT
            
            def _v(key):
                val = row_data.get(key, np.nan)
                try:
                    return round(float(val), 1) if pd.notna(val) else None
                except Exception:
                    return None
            
            vals = [
                row_data.get('Empresa', ''),
                row_data.get('Fundo', ''),
                row_data['Fecha'].strftime('%Y/%m/%d') if pd.notna(row_data.get('Fecha')) else '',
                _v('max'), _v('min'),
                _v('max pred'), _v('min pred'),
                row_data.get('status', 'predicha'),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws_dia.cell(row=er, column=ci)
                c.value = v; c.fill = fill; c.font = font
                c.alignment = CENT; c.border = THIN
    
    _autowidth(ws_dia)
    ws_dia.freeze_panes = 'A2'
    
    # HOJA 2: Por Horas (reducida)
    ws_hora = wb.create_sheet('Por_Horas')
    ws_hora.sheet_properties.tabColor = '1B5E20'
    
    pred_hora_df = _generar_pred_horaria(df_pred)
    
    if not pred_hora_df.empty:
        sub_max = pred_hora_df[pred_hora_df['Variable'] == 'Tmax'][
            ['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora', 'max_pred']
        ].copy()
        sub_min = pred_hora_df[pred_hora_df['Variable'] == 'Tmin'][
            ['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora', 'min_pred']
        ].copy()
        
        if not sub_max.empty and not sub_min.empty:
            tabla_h = sub_max.merge(
                sub_min, on=['Empresa', 'Fundo', 'Fecha', 'Hora', 'FechaHora'], how='outer'
            )
        elif not sub_max.empty:
            tabla_h = sub_max.copy()
            if 'min_pred' not in tabla_h.columns:
                tabla_h['min_pred'] = None
        elif not sub_min.empty:
            tabla_h = sub_min.copy()
            if 'max_pred' not in tabla_h.columns:
                tabla_h['max_pred'] = None
        else:
            tabla_h = pd.DataFrame()
        
        if not tabla_h.empty:
            tabla_h = tabla_h.sort_values(['Fundo', 'FechaHora']).reset_index(drop=True)
            
            cols_hora = ['Empresa', 'Fundo', 'Fecha', 'Hora', 'max pred', 'min pred', 'status']
            for ci, h_txt in enumerate(cols_hora, 1):
                c = ws_hora.cell(row=1, column=ci)
                c.value = h_txt; c.fill = HDR_HORA
                c.font = HDR_FNT_W; c.alignment = CENT; c.border = THIN
            
            for ri, row_data in tabla_h.iterrows():
                er = ri + 2
                max_v = row_data.get('max_pred')
                min_v = row_data.get('min_pred')
                vals_h = [
                    row_data.get('Empresa', ''),
                    row_data.get('Fundo', ''),
                    row_data.get('Fecha', ''),
                    row_data.get('Hora', ''),
                    round(float(max_v), 1) if max_v is not None and not pd.isna(max_v) else None,
                    round(float(min_v), 1) if min_v is not None and not pd.isna(min_v) else None,
                    'predicha',
                ]
                for ci, v in enumerate(vals_h, 1):
                    c = ws_hora.cell(row=er, column=ci)
                    c.value = v; c.fill = PRED_F; c.font = PRED_FNT
                    c.alignment = CENT; c.border = THIN
            
            _autowidth(ws_hora)
            ws_hora.freeze_panes = 'A2'
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
# ══════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def _leer_normales_desde_disco(path: str):
    with open(path, 'rb') as f:
        return f.read()

@st.cache_data(show_spinner=False)
def cargar_geojson_peru():
    """Descarga GeoJSON de distritos de Perú desde GitHub."""
    url = "https://raw.githubusercontent.com/juaneladio/peru-geojson/master/peru_distrital_simple.geojson"
    try:
        import urllib.request
        with urllib.request.urlopen(url) as r:
            return json.loads(r.read().decode('utf-8'))
    except Exception as e:
        st.warning(f"⚠️ No se pudo cargar el mapa de Perú: {e}")
        return None

def generar_mapa_distritos(geojson, distrito_fijo, distrito_din, variable, label_clim2=None):
    if geojson is None:
        return None

    import folium

    color_fijo = '#C62828' if variable == 'Tmax' else '#1565C0'
    color_din  = '#FF6F00'

    dist_fijo_norm = _normalizar(distrito_fijo)
    dist_din_norm  = _normalizar(distrito_din) if distrito_din else None

    # ── Mapa base Esri Satellite ───────────────────────────────
    m = folium.Map(
        location=[-7.27, -79.45],
        zoom_start=7,
        tiles=None,
        control_scale=False,
        zoom_control=False,
    )

    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri WorldImagery',
        name='Satellite',
        overlay=False,
        control=False,
    ).add_to(m)

    folium.TileLayer(
        tiles='https://services.arcgisonline.com/ArcGIS/rest/services/Reference/World_Boundaries_and_Places/MapServer/tile/{z}/{y}/{x}',
        attr='Esri Labels',
        name='Labels',
        overlay=True,
        control=False,
        opacity=0.7,
    ).add_to(m)

    from folium.plugins import Fullscreen
    Fullscreen(
        position='topright',
        title='Pantalla completa',
        title_cancel='Salir',
        force_separate_button=True,
    ).add_to(m)
    # ── Dibujar distritos relevantes ───────────────────────────
    def _style_fijo(f):
        return {
            'fillColor'  : color_fijo,
            'color'      : 'white',
            'weight'     : 1.5,
            'fillOpacity': 0.55,
        }

    def _style_din(f):
        return {
            'fillColor'  : color_din,
            'color'      : 'white',
            'weight'     : 1.5,
            'fillOpacity': 0.55,
        }

    for feat in geojson['features']:
        p      = feat['properties']
        nombre = _normalizar(p['NOMBDIST'])

        es_fijo = nombre == dist_fijo_norm
        es_din  = dist_din_norm and nombre == dist_din_norm

        if not es_fijo and not es_din:
            continue

        label_txt = p['NOMBDIST'] + ' — ' + p.get('NOMBDEP', '')

        folium.GeoJson(
            feat,
            style_function=_style_fijo if es_fijo else _style_din,
            tooltip=folium.Tooltip(label_txt, sticky=False),
        ).add_to(m)

        # Centroide para marcador
        geom   = feat['geometry']
        coords = []
        if geom['type'] == 'Polygon':
            coords = geom['coordinates'][0]
        elif geom['type'] == 'MultiPolygon':
            coords = geom['coordinates'][0][0]

        if coords:
            lon_c = sum(c[0] for c in coords) / len(coords)
            lat_c = sum(c[1] for c in coords) / len(coords)

            folium.Marker(
                location=[lat_c, lon_c],
                popup=folium.Popup(label_txt, max_width=200),
                icon=folium.Icon(
                    color='red'    if es_fijo else 'orange',
                    icon='tint'    if variable == 'Tmax' else 'info-sign',
                    prefix='glyphicon',
                ),
            ).add_to(m)
 # ── Fundos Aquanqa (puntos fijos) ─────────────────────────
    fundos_coords = [
        {'nombre': 'Ayllu Allpa', 'lat': -7.64792, 'lon': -79.38378},
        {'nombre': 'Vivadis',     'lat': -7.64375, 'lon': -79.36462},
        {'nombre':'Arena Azul',    'lat': -7.65512, 'lon': -79.35300},
    ]

    for fundo in fundos_coords:
        folium.Marker(
            location=[fundo['lat'], fundo['lon']],
            popup=folium.Popup(f"<b>{fundo['nombre']}</b><br>Fundo Aquanqa", max_width=150),
            tooltip=folium.Tooltip(fundo['nombre'], sticky=False),
            icon=folium.Icon(
                color='blue',
                icon='leaf',
                prefix='glyphicon',
            ),
        ).add_to(m)

    return m
def cargar_normales_dinamico(_file_bytes, distrito_sel: str):
    """
    Carga normales SENAMHI para el distrito seleccionado dinámicamente.
    Reemplaza completamente a cargar_normales_fijo().
    """
    try:
        # Función interna que usa distrito dinámico (no la constante DISTRITO_BUSCAR)
        def _cargar_hoja(hoja):
            raw = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=None)
            header_row = None
            for idx, row in raw.iterrows():
                vals = [_normalizar(str(v)) for v in row.values if pd.notna(v)]
                if 'DISTRITO' in vals:
                    header_row = idx
                    break
 
            if header_row is None:
                raise ValueError(f"No se encontró fila con 'DISTRITO' en hoja '{hoja}'.")
 
            df = pd.read_excel(io.BytesIO(_file_bytes), sheet_name=hoja, header=header_row)
            df.columns = [str(c).strip() for c in df.columns]
 
            col_distrito = next((c for c in df.columns if _normalizar(c) == 'DISTRITO'), None)
            if col_distrito is None:
                raise ValueError("Columna DISTRITO no encontrada.")
 
            # Filtrar por distrito seleccionado dinámicamente
            df['_dist_norm'] = df[col_distrito].astype(str).apply(_normalizar)
            df_f = df[df['_dist_norm'].str.contains(distrito_sel, na=False)].copy()
 
            if df_f.empty:
                raise ValueError(f"No se encontraron datos para el distrito '{distrito_sel}'.")
 
            # Extraer valores mensuales
            meses_encontrados = {}
            for col in df.columns:
                col_norm = _normalizar(col)
                for mes in MESES_RAW:
                    if col_norm == _normalizar(mes) and mes not in meses_encontrados:
                        meses_encontrados[mes] = col
                        break
 
            meses_faltantes = [m for m in MESES_RAW if m not in meses_encontrados]
            if meses_faltantes:
                raise ValueError(f"Columnas de meses no encontradas: {meses_faltantes}")
 
            cols_meses = [meses_encontrados[m] for m in MESES_RAW]
            valores = df_f[cols_meses].apply(pd.to_numeric, errors='coerce').mean(axis=0).values
 
            col_nombre = next((c for c in df.columns if _normalizar(c) == 'NOMBRE ESTACION'), col_distrito)
            col_prov   = next((c for c in df.columns if _normalizar(c) == 'PROVINCIA'), col_distrito)
 
            return (
                valores,
                valores - DELTA_Q,
                valores + DELTA_Q,
                df_f[[col_nombre, col_prov, col_distrito]].drop_duplicates()
            )
 
        MEDIA_TMAX, Q1_TMAX, Q3_TMAX, estaciones_tmax = _cargar_hoja('TMAX')
        MEDIA_TMIN, Q1_TMIN, Q3_TMIN, _               = _cargar_hoja('TMIN')
 
        return MEDIA_TMAX, Q1_TMAX, Q3_TMAX, MEDIA_TMIN, Q1_TMIN, Q3_TMIN, estaciones_tmax
 
    except Exception as e:
        st.error(f"Error al cargar normales para '{distrito_sel}': {e}")
        return None
 

# ──── SIDEBAR MEJORADO ────
norm_bytes_sidebar = _leer_normales_desde_disco(NORMALES_PATH) if os.path.exists(NORMALES_PATH) else None

with st.sidebar:
    st.markdown("## ⚙️ CONFIGURACIÓN")

    with st.expander("📂 **Archivo meteorológico**", expanded=True):
        st.markdown("### 📊 **Archivo meteorológico**")
        st.info("✅ Leyendo: `assets/Metereologia_Prize.xlsx`")

    st.divider()

    st.markdown("### 🌿 **Fundos**")
    fundos_disponibles_placeholder = st.empty()

    st.divider()
    # Estado corrección BIAS
    bias_dict = st.session_state.get('bias_correccion', {})
    if bias_dict:
        st.markdown("### 🎯 Corrección BIAS")
        for (f, v), b in bias_dict.items():
            icono = '🔺' if b > 0 else '🔻'
            st.caption(f"{icono} {f} {v}: {b:+.2f}°C")
    else:
        st.caption("⚠️ Abre validación para activar BIAS")
    st.markdown("### 🌦️ **Climatología SENAMHI**")

    # ── Guadalupe siempre fijo ─────────────────────────────────
    st.info(f"📌 Referencia fija: **{DISTRITO_BUSCAR}**")

    # ── Selector dinámico (sin Guadalupe) ─────────────────────
    st.markdown("**Comparar con otra estación:**")

    if norm_bytes_sidebar is not None:
        catalogo = cargar_catalogo_normales(norm_bytes_sidebar, 'TMAX')

        if catalogo:
            sectores_lista = sorted(catalogo.keys())

            sector_sel = st.selectbox(
                "Sector",
                options=sectores_lista,
                index=0,
                key='sector_sel'
            )

            deptos_lista = sorted(catalogo.get(sector_sel, {}).keys())

            depto_sel = st.selectbox(
                "Departamento",
                options=deptos_lista,
                index=0,
                key='depto_sel'
            )

            distritos_lista = [DISTRITO_NINGUNA] + catalogo.get(sector_sel, {}).get(depto_sel, [])

            distrito_sel2 = st.selectbox(
                "Estación / Distrito",
                options=distritos_lista,
                index=0,   # ← "— Ninguna —" por defecto
                key='distrito_sel2'
            )

            if distrito_sel2 != DISTRITO_NINGUNA:
                st.success(f"✅ Comparando: **{distrito_sel2}**")
            else:
                st.caption("Sin comparación activa.")
        else:
            st.error("❌ No se pudo leer el catálogo.")
            distrito_sel2 = DISTRITO_NINGUNA
    else:
        st.error("❌ Normales SENAMHI no encontradas")
        distrito_sel2 = DISTRITO_NINGUNA

# ── GeoJSON Perú ──────────────────────────────────────────────
geojson_peru = cargar_geojson_peru()  # ← AQUÍ
# ──── HEADER ────
st.markdown("""
<div class="header-bar">
    <h1>🌡️ Temperatura Fundos — Aquanqa</h1>
    <p>Climatología SENAMHI · Predicción Prophet · OPTIMIZADO</p>
</div>
""", unsafe_allow_html=True)

# ──── VALORES POR DEFECTO ────
dias_vista_num = 30  # Mostrar últimos 30 días
variable_sel = 'Ambas'  # Mostrar Tmax y Tmin
min_reg_ui = MIN_REGISTROS  # Registros mínimos por día (valor de constante)
dias_pred_ui = 30  # Días de predicción Prophet

# ──── CARGAR NORMALES ────────────────────────────────────────

# Guadalupe siempre fijo
normales_guadalupe = cargar_normales_dinamico(norm_bytes_sidebar, DISTRITO_BUSCAR)
if normales_guadalupe is None:
    st.stop()

MEDIA_TMAX, Q1_TMAX, Q3_TMAX, MEDIA_TMIN, Q1_TMIN, Q3_TMIN, estaciones_tmax = normales_guadalupe

# Normales dinámicas (solo si se seleccionó algo)
normales_din = None
if distrito_sel2 != DISTRITO_NINGUNA:
    normales_din = cargar_normales_dinamico(norm_bytes_sidebar, distrito_sel2)

if normales_din is not None:
    MEDIA_TMAX2, Q1_TMAX2, Q3_TMAX2, MEDIA_TMIN2, Q1_TMIN2, Q3_TMIN2, _ = normales_din
else:
    MEDIA_TMAX2 = Q1_TMAX2 = Q3_TMAX2 = None
    MEDIA_TMIN2 = Q1_TMIN2 = Q3_TMIN2 = None

# ──── LECTURA DIRECTA DEL ARCHIVO ────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
METEO_PATH = os.path.join(SCRIPT_DIR, "assets", "Metereologia_Prize.xlsx")

# Intentar rutas alternativas si la principal no existe
if not os.path.exists(METEO_PATH):
    alt_paths = [
        "assets/Metereologia_Prize.xlsx",
        "./assets/Metereologia_Prize.xlsx",
        "../assets/Metereologia_Prize.xlsx",
    ]
    for alt_path in alt_paths:
        if os.path.exists(alt_path):
            METEO_PATH = alt_path
            break

with st.spinner(f"Leyendo meteorología (optimizado)..."):
    try:
        if not os.path.exists(METEO_PATH):
            st.error(f"❌ Archivo no encontrado en: {METEO_PATH}")
            st.info("📂 Rutas buscadas:")
            st.code(f"{SCRIPT_DIR}/assets/Metereologia_Prize.xlsx\n./assets/Metereologia_Prize.xlsx")
            st.stop()
        
        # Leer archivo directamente desde disco
        with open(METEO_PATH, 'rb') as f:
            meteo_bytes = f.read()
        
        filename = "Metereologia_Prize.xlsx"
        df_preview = leer_meteo_bytes_optimizado(meteo_bytes, filename)
        
        if 'Fundo' not in df_preview.columns:
            st.error(f"Columna `Fundo` no encontrada.")
            st.stop()
        fundos_disponibles = sorted(df_preview['Fundo'].dropna().unique().tolist())
    except Exception as e:
        st.error(f"Error: {e}")
        st.stop()

with fundos_disponibles_placeholder:
    fundos_sel = st.multiselect(
        "Seleccionar fundos",
        options=fundos_disponibles,
        default=fundos_disponibles,
    )

fundos_activos = fundos_sel if fundos_sel else fundos_disponibles

with st.spinner("Procesando datos (optimizado)..."):
    try:
        dia_full = cargar_meteoro_optimizado(
            meteo_bytes, filename, SHEET_NAME,
            tuple(),
            min_reg_ui
        )
        dia = dia_full[dia_full['Fundo'].isin(fundos_activos)].reset_index(drop=True)
    except Exception as e:
        st.error(f"Error: {e}")
        st.stop()

# ──── PROPHET ────
CACHE_DIR  = Path(_APP_DIR) / "assets" / ".prophet_cache"
CACHE_PKL  = CACHE_DIR / "forecasts.pkl"
CACHE_HASH = CACHE_DIR / "data_hash.txt"

def _hash_meteo(file_bytes: bytes) -> str:
    """MD5 del archivo Excel completo — cambia solo si los datos cambian."""
    return hashlib.md5(file_bytes).hexdigest()

def guardar_cache_prophet(forecasts: dict, data_hash: str):
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    with open(CACHE_PKL, "wb") as f:
        pickle.dump(forecasts, f)
    CACHE_HASH.write_text(data_hash)

def cargar_cache_prophet(data_hash: str) -> dict | None:
    """Devuelve el dict de forecasts si el hash coincide, None si no."""
    if not CACHE_PKL.exists() or not CACHE_HASH.exists():
        return None
    if CACHE_HASH.read_text().strip() != data_hash:
        return None  # datos cambiaron → reentrenar
    with open(CACHE_PKL, "rb") as f:
        return pickle.load(f)
    
_data_hash = _hash_meteo(meteo_bytes)

forecasts_cache = cargar_cache_prophet(_data_hash)

if forecasts_cache is None:
    # Primera vez o datos cambiaron
    with st.spinner("Entrenando modelos Prophet (primera vez o datos nuevos)..."):
        forecasts_cache = entrenar_todos_optimizado(dia_full, dias_pred_ui)
    guardar_cache_prophet(forecasts_cache, _data_hash)
    st.success("Modelos entrenados y guardados en caché.")
else:
    st.info("Modelos cargados desde caché — sin reentrenar.")

# ──── MÉTRICAS ────
st.markdown("#### 📊 Resumen de datos")

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Registros</div>
        <div class="value">{len(dia):,}</div>
        <div class="sub">días-fundo</div>
    </div>""", unsafe_allow_html=True)
with col2:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Fundos</div>
        <div class="value">{dia['Fundo'].nunique()}</div>
        <div class="sub">activos</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Desde</div>
        <div class="value" style="font-size:1.15rem">{dia['Fecha'].min().strftime('%d/%m')}</div>
        <div class="sub">{dia['Fecha'].min().year}</div>
    </div>""", unsafe_allow_html=True)
with col4:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Hasta</div>
        <div class="value" style="font-size:1.15rem">{dia['Fecha'].max().strftime('%d/%m')}</div>
        <div class="sub">{dia['Fecha'].max().year}</div>
    </div>""", unsafe_allow_html=True)
with col5:
    st.markdown(f"""
    <div class="metric-card">
        <div class="label">Tmax prom</div>
        <div class="value">{dia['Tmax'].mean():.1f}°C</div>
        <div class="sub">todos</div>
    </div>""", unsafe_allow_html=True)

st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)

# ──── TABS ────
mostrar_tmax = variable_sel in ['Ambas', 'Solo Tmax']
mostrar_tmin = variable_sel in ['Ambas', 'Solo Tmin']

# Calcular días de predicción dinámicos
dias_pred_mitad = max(7, dias_pred_ui // 2)  # Mínimo 7 días

if mostrar_tmax and mostrar_tmin:
    tab_tmax, tab_tmin, tab_et, tab_datos = st.tabs(
        ["🔴 Temperatura Máxima", "🔵 Temperatura Mínima", "💧 ET", "📋 Datos"]
    )
elif mostrar_tmax:
    tab_tmax, tab_et, tab_datos = st.tabs(["🔴 Temperatura Máxima", "💧 ET", "📋 Datos"])
    tab_tmin = None
else:
    tab_tmin, tab_et, tab_datos = st.tabs(["🔵 Temperatura Mínima", "💧 ET", "📋 Datos"])
    tab_tmax = None
n_fundos        = len(dia['Fundo'].unique())
altura_graficos = 320 * n_fundos
if mostrar_tmax and tab_tmax is not None:
    with tab_tmax:
        col1, col2 = st.columns([2, 2])
        with col1:
            st.markdown("#### 📊 Predicción a mostrar")
            dias_pred_mostrar = st.radio(
                "Mostrar predicción de:",
                options=[f"{dias_pred_mitad}d", f"{dias_pred_ui}d"],
                index=1, horizontal=True,
                key='dias_pred_mostrar_tmax'
            )
            dias_pred_mostrar_num = int(dias_pred_mostrar.replace('d', ''))

        tipo_viz_tmax_lower = 'linea'

        with st.spinner("Generando Tmax..."):
            fig_tmax, df_tmax_hist, df_tmax_pred = generar_figura(
                'Tmax', dia, MEDIA_TMAX, Q1_TMAX, Q3_TMAX,
                dias_pred_ui, forecasts_cache, dias_vista_num,
                dias_pred_mostrar_num, tipo_viz_tmax_lower,
                media_mensual2=MEDIA_TMAX2, q1_mensual2=Q1_TMAX2,
                q3_mensual2=Q3_TMAX2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )

        col_graf_tmax, col_mapa_tmax = st.columns([3, 1])
        with col_graf_tmax:
            st.plotly_chart(fig_tmax, use_container_width=True)
        with col_mapa_tmax:
            from streamlit_folium import st_folium
            mapa_tmax = generar_mapa_distritos(
                geojson_peru,
                distrito_fijo=DISTRITO_BUSCAR,
                distrito_din=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                variable='Tmax',
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )
            if mapa_tmax:
                st_folium(mapa_tmax, width=None, height=altura_graficos,
                          returned_objects=[], key='mapa_tmax')

        # ← DENTRO del with tab_tmax
        generar_seccion_validacion('Tmax', dia, forecasts_cache, dias_pred_ui)

else:
    df_tmax_hist = pd.DataFrame()
    df_tmax_pred = pd.DataFrame()

if mostrar_tmin and tab_tmin is not None:
    with tab_tmin:
        col1, col2 = st.columns([2, 2])
        with col1:
            st.markdown("#### 📊 Predicción a mostrar")
            dias_pred_mostrar = st.radio(
                "Mostrar predicción de:",
                options=[f"{dias_pred_mitad}d", f"{dias_pred_ui}d"],
                index=1, horizontal=True,
                key='dias_pred_mostrar_tmin'
            )
            dias_pred_mostrar_num = int(dias_pred_mostrar.replace('d', ''))

        tipo_viz_tmin_lower = 'linea'

        with st.spinner("Generando Tmin..."):
            fig_tmin, df_tmin_hist, df_tmin_pred = generar_figura(
                'Tmin', dia, MEDIA_TMIN, Q1_TMIN, Q3_TMIN,
                dias_pred_ui, forecasts_cache, dias_vista_num,
                dias_pred_mostrar_num, tipo_viz_tmin_lower,
                media_mensual2=MEDIA_TMIN2, q1_mensual2=Q1_TMIN2,
                q3_mensual2=Q3_TMIN2,
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )

        col_graf_tmin, col_mapa_tmin = st.columns([3, 1])
        with col_graf_tmin:
            st.plotly_chart(fig_tmin, use_container_width=True)
        with col_mapa_tmin:
            from streamlit_folium import st_folium
            mapa_tmin = generar_mapa_distritos(
                geojson_peru,
                distrito_fijo=DISTRITO_BUSCAR,
                distrito_din=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None,
                variable='Tmin',
                label_clim2=distrito_sel2 if distrito_sel2 != DISTRITO_NINGUNA else None
            )
            if mapa_tmin:
                st_folium(mapa_tmin, width=None, height=altura_graficos,
                          returned_objects=[], key='mapa_tmin')

        # ← DENTRO del with tab_tmin
        generar_seccion_validacion('Tmin', dia, forecasts_cache, dias_pred_ui)

else:
    df_tmin_hist = pd.DataFrame()
    df_tmin_pred = pd.DataFrame()

# ── Tab ET — independiente ─────────────────────────────────────
if tab_et is not None:
    with tab_et:
        generar_tab_et(dia, forecasts_cache, dias_pred_ui)
else:   
    df_tmin_hist = pd.DataFrame()
    df_tmin_pred = pd.DataFrame()

# ──── TAB DATOS ────
with tab_datos:
    st.markdown("#### 📋 Tabla de datos diarios")
    
    col_f, col_var = st.columns([2, 2])
    with col_f:
        fundos_tabla = st.multiselect(
            "Filtrar fundo", dia['Fundo'].unique().tolist(),
            default=dia['Fundo'].unique().tolist(), key='tabla_fundo'
        )
    with col_var:
        var_tabla = st.selectbox("Variable", ['Tmax', 'Tmin', 'Ambas'], key='tabla_var')
    
    df_show = dia[dia['Fundo'].isin(fundos_tabla)].copy()
    df_show['Fecha'] = df_show['Fecha'].dt.strftime('%Y-%m-%d')
    
    if var_tabla == 'Tmax':
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmax', 'Tmax_smooth']
    elif var_tabla == 'Tmin':
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmin', 'Tmin_smooth']
    else:
        cols_show = ['Empresa', 'Fundo', 'Fecha', 'Tmax', 'Tmax_smooth', 'Tmin', 'Tmin_smooth']
    
    st.dataframe(df_show[cols_show], use_container_width=True, hide_index=True, height=340)
    
    st.markdown("<hr class='section-divider'>", unsafe_allow_html=True)
    st.markdown("#### ⬇️ Descargas")
    
    col_d1, col_d2, col_d3 = st.columns(3)
    
    df_hist_all = pd.concat([df_tmax_hist, df_tmin_hist], ignore_index=True)
    df_pred_all = pd.concat([df_tmax_pred, df_tmin_pred], ignore_index=True)
    
    with col_d1:
        if not df_hist_all.empty:
            excel_bytes = exportar_excel(df_hist_all, df_pred_all)
            st.download_button(
                label="📥 Excel completo",
                data=excel_bytes,
                file_name="Temperatura_Aquanqa_BI.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    with col_d2:
        if not df_hist_all.empty:
            st.download_button(
                label="📄 CSV histórico",
                data=df_hist_all.to_csv(index=False).encode('utf-8'),
                file_name="historico_temperatura.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col_d3:
        if not df_pred_all.empty:
            st.download_button(
                label="📄 CSV predicciones",
                data=df_pred_all.to_csv(index=False).encode('utf-8'),
                file_name="predicciones_prophet.csv",
                mime="text/csv",
                use_container_width=True
            )

st.markdown("""
<div class="footer-note">
    Elaborado por la Gerencia de Planificación, Control Operacional y Gestión
</div>
""", unsafe_allow_html=True)